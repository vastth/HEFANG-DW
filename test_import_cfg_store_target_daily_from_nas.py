# -*- coding: utf-8 -*-

import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook

from tools.import_cfg_store_target_daily_from_nas import (
    ParsedAssessmentAssignmentRow,
    SourceRow,
    _build_validation_error_message,
    _build_validation_warning_messages,
    _build_store_attr_rows,
    _expand_rows,
    _parse_optional_assessment_assignment_sheet,
    _parse_workbook,
    _resolve_assessment_assignments,
)


def _make_summary(**overrides):
    summary = {
        'ambiguous_store_names': [],
        'source_row_count': 10,
        'matched_store_count': 9,
        'missing_store_names': ['长沙运达汇店'],
        'missing_store_suggestions': {'长沙运达汇店': ['长沙IFS店']},
        'store_attr_candidate_duplicate_store_ids': [],
        'sync_store_report_attr': False,
        'store_attr_overlap_rows': [],
        'sync_assessment': False,
        'assessment_assignment_source_row_count': 0,
        'assessment_assignment_row_count': 0,
        'assessment_assignment_missing_store_keys': [],
        'assessment_assignment_missing_store_names': [],
        'assessment_assignment_missing_store_suggestions': {},
        'assessment_assignment_store_key_name_mismatch_rows': [],
        'assessment_assignment_missing_anchor_store_names': [],
        'assessment_assignment_missing_anchor_store_suggestions': {},
        'assessment_assignment_inconsistent_primary_anchor_rows': [],
        'assessment_assignment_unanchored_rows': [],
        'assessment_assignment_overlap_rows': [],
        'assessment_assignment_missing_subject_codes': [],
        'assessment_assignment_skipped_missing_primary_rows': [],
    }
    summary.update(overrides)
    return summary


def _write_target_workbook(
    workbook_path: Path,
    *,
    grade_header: str = '等级',
    grade_value: str | None = 'A',
    effective_start_date: str | None = None,
    effective_end_date: str | None = None,
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = '导入模板'

    header_row = ['目标月份', '目标版本', '门店名称', '门店类型']
    if grade_header:
        header_row.append(grade_header)
    if effective_start_date is not None:
        header_row.append('生效开始日')
    if effective_end_date is not None:
        header_row.append('生效结束日')
    header_row.append('月目标')
    header_row.extend([f'{day}日目标' for day in range(1, 32)])
    worksheet.append(header_row)

    data_row = ['2026-05', 'v1', '广州K11专卖店', '直营']
    if grade_header:
        data_row.append(grade_value)
    if effective_start_date is not None:
        data_row.append(effective_start_date)
    if effective_end_date is not None:
        data_row.append(effective_end_date)
    data_row.append(310000)
    data_row.extend([10000 + day for day in range(1, 32)])
    worksheet.append(data_row)

    workbook.save(workbook_path)


class TestImportCfgStoreTargetDailyFromNas(unittest.TestCase):
    def test_parse_workbook_reads_store_grade_from_excel_column(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / '202605考核数据配置表.xlsx'
            _write_target_workbook(workbook_path, grade_header='等级', grade_value='A')

            parsed_rows, _ = _parse_workbook(
                workbook_path,
                '导入模板',
                require_store_type=True,
                target_month_filter='2026-05',
            )

        self.assertEqual(len(parsed_rows), 1)
        self.assertEqual(parsed_rows[0].store_grade, 'A')

    def test_parse_workbook_reads_effective_date_range_from_excel_columns(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / '202605考核数据配置表.xlsx'
            _write_target_workbook(
                workbook_path,
                effective_start_date='2026-05-07',
                effective_end_date='2026-05-31',
            )

            parsed_rows, _ = _parse_workbook(
                workbook_path,
                '导入模板',
                require_store_type=True,
                target_month_filter='2026-05',
            )

        self.assertEqual(len(parsed_rows), 1)
        self.assertEqual(parsed_rows[0].effective_start_date, date(2026, 5, 7))
        self.assertEqual(parsed_rows[0].effective_end_date, date(2026, 5, 31))

    def test_expand_rows_only_keeps_dates_inside_effective_range(self):
        parsed_rows = [
            SourceRow(
                row_number=2,
                target_month='2026-05',
                target_month_start=date(2026, 5, 1),
                target_version='v1',
                store_name='广州K11专卖店',
                store_type='直营',
                store_grade='A',
                month_target=0,
                day_targets={day: 1 for day in range(1, 32)},
                blank_day_cell_count=0,
                effective_start_date=date(2026, 5, 7),
                effective_end_date=date(2026, 5, 8),
            )
        ]
        store_map = {
            '广州k11专卖店': {
                'store_id': 13,
                'store_code': 'RT001',
            }
        }

        expanded_rows, missing_store_names = _expand_rows(parsed_rows, store_map)

        self.assertEqual(missing_store_names, [])
        self.assertEqual([row.target_date for row in expanded_rows], [date(2026, 5, 7), date(2026, 5, 8)])

    def test_build_store_attr_rows_prefers_excel_store_grade_over_existing_value(self):
        parsed_rows = [
            SourceRow(
                row_number=2,
                target_month='2026-05',
                target_month_start=date(2026, 5, 1),
                target_version='v1',
                store_name='广州K11专卖店',
                store_type='直营',
                store_grade='A',
                month_target=0,
                day_targets={1: 0},
                blank_day_cell_count=0,
                effective_start_date=date(2026, 5, 1),
                effective_end_date=date(2026, 5, 31),
            )
        ]
        store_map = {
            '广州k11专卖店': {
                'store_id': 13,
                'store_code': 'RT001',
            }
        }
        latest_store_attr_map = {
            13: {
                'store_grade': 'B',
                'is_duty_free': 'N',
            }
        }

        store_attr_rows = _build_store_attr_rows(
            parsed_rows,
            store_map,
            latest_store_attr_map,
            date(2026, 5, 1),
        )

        self.assertEqual(len(store_attr_rows), 1)
        self.assertEqual(store_attr_rows[0].store_grade, 'A')

    def test_build_store_attr_rows_keeps_existing_store_grade_when_excel_blank(self):
        parsed_rows = [
            SourceRow(
                row_number=2,
                target_month='2026-05',
                target_month_start=date(2026, 5, 1),
                target_version='v1',
                store_name='广州K11专卖店',
                store_type='直营',
                store_grade=None,
                month_target=0,
                day_targets={1: 0},
                blank_day_cell_count=0,
                effective_start_date=date(2026, 5, 1),
                effective_end_date=date(2026, 5, 31),
            )
        ]
        store_map = {
            '广州k11专卖店': {
                'store_id': 13,
                'store_code': 'RT001',
            }
        }
        latest_store_attr_map = {
            13: {
                'store_grade': 'B',
                'is_duty_free': 'N',
            }
        }

        store_attr_rows = _build_store_attr_rows(
            parsed_rows,
            store_map,
            latest_store_attr_map,
            date(2026, 5, 1),
        )

        self.assertEqual(len(store_attr_rows), 1)
        self.assertEqual(store_attr_rows[0].store_grade, 'B')

    def test_build_store_attr_rows_marks_duty_free_when_channel_type_contains_keyword(self):
        parsed_rows = [
            SourceRow(
                row_number=2,
                target_month='2026-05',
                target_month_start=date(2026, 5, 1),
                target_version='v1',
                store_name='杭州萧山国际机场店',
                store_type='联营-免税',
                store_grade=None,
                month_target=0,
                day_targets={1: 0},
                blank_day_cell_count=0,
                effective_start_date=date(2026, 5, 1),
                effective_end_date=date(2026, 5, 31),
            )
        ]
        store_map = {
            '杭州萧山国际机场店': {
                'store_id': 708,
                'store_code': 'RT110',
            }
        }
        latest_store_attr_map = {
            708: {
                'store_grade': None,
                'is_duty_free': 'N',
            }
        }

        store_attr_rows = _build_store_attr_rows(
            parsed_rows,
            store_map,
            latest_store_attr_map,
            date(2026, 5, 1),
        )

        self.assertEqual(len(store_attr_rows), 1)
        self.assertEqual(store_attr_rows[0].is_duty_free, 'Y')

    def test_build_store_attr_rows_marks_non_duty_free_when_channel_type_not_contains_keyword(self):
        parsed_rows = [
            SourceRow(
                row_number=2,
                target_month='2026-05',
                target_month_start=date(2026, 5, 1),
                target_version='v1',
                store_name='苏州比斯特店',
                store_type='联营-奥莱',
                store_grade=None,
                month_target=0,
                day_targets={1: 0},
                blank_day_cell_count=0,
                effective_start_date=date(2026, 5, 31),
                effective_end_date=date(2026, 5, 31),
            )
        ]
        store_map = {
            '苏州比斯特店': {
                'store_id': 583,
                'store_code': 'RT091',
            }
        }
        latest_store_attr_map = {
            583: {
                'store_grade': 'B',
                'is_duty_free': 'Y',
            }
        }

        store_attr_rows = _build_store_attr_rows(
            parsed_rows,
            store_map,
            latest_store_attr_map,
            date(2026, 5, 31),
        )

        self.assertEqual(len(store_attr_rows), 1)
        self.assertEqual(store_attr_rows[0].is_duty_free, 'N')

    def test_build_store_attr_rows_uses_row_level_effective_start_date(self):
        parsed_rows = [
            SourceRow(
                row_number=2,
                target_month='2026-05',
                target_month_start=date(2026, 5, 1),
                target_version='v1',
                store_name='广州K11专卖店',
                store_type='直营',
                store_grade='A',
                month_target=0,
                day_targets={7: 0},
                blank_day_cell_count=0,
                effective_start_date=date(2026, 5, 7),
                effective_end_date=date(2026, 5, 31),
            )
        ]
        store_map = {
            '广州k11专卖店': {
                'store_id': 13,
                'store_code': 'RT001',
            }
        }

        store_attr_rows = _build_store_attr_rows(
            parsed_rows,
            store_map,
            latest_store_attr_map={},
            effective_start_date=date(2026, 5, 1),
        )

        self.assertEqual(len(store_attr_rows), 1)
        self.assertEqual(store_attr_rows[0].effective_start_date, date(2026, 5, 7))

    def test_store_attr_sync_no_longer_blocks_when_only_has_existing_previous_version(self):
        summary = _make_summary(
            sync_store_report_attr=True,
            store_attr_overlap_rows=[],
            existing_store_attr_rows_in_target_month=73,
            existing_store_attr_rows_same_start_date=0,
            store_attr_effective_start_source='existing_latest_in_target_month',
            store_attr_effective_start_date=date(2026, 5, 1),
        )

        error_message = _build_validation_error_message(summary)

        self.assertEqual(error_message, '')

    def test_store_attr_sync_still_fails_on_true_current_overlap(self):
        summary = _make_summary(
            sync_store_report_attr=True,
            store_attr_effective_start_date=date(2026, 5, 1),
            store_attr_overlap_rows=[
                {
                    'store_id': 13,
                    'store_code': 'RT001',
                    'store_name': '广州K11专卖店',
                    'report_channel_type': '直营',
                    'effective_start_date': '2026-04-01',
                    'effective_end_date': '9999-12-31',
                },
                {
                    'store_id': 13,
                    'store_code': 'RT001',
                    'store_name': '广州K11专卖店',
                    'report_channel_type': '联营',
                    'effective_start_date': '2026-04-15',
                    'effective_end_date': '9999-12-31',
                },
            ],
        )

        error_message = _build_validation_error_message(summary)

        self.assertIn('同店多条当前有效配置', error_message)
        self.assertIn('RT001', error_message)

    def test_missing_store_becomes_warning_when_still_has_matched_rows(self):
        summary = _make_summary()

        error_message = _build_validation_error_message(summary)
        warning_messages = _build_validation_warning_messages(summary)

        self.assertEqual(error_message, '')
        self.assertEqual(len(warning_messages), 1)
        self.assertIn('已跳过这些门店的目标/门店属性配置', warning_messages[0])
        self.assertIn('长沙运达汇店', warning_messages[0])

    def test_all_target_rows_missing_still_fails_safely(self):
        summary = _make_summary(
            source_row_count=1,
            matched_store_count=0,
        )

        error_message = _build_validation_error_message(summary)

        self.assertIn('全部未命中', error_message)
        self.assertIn('无法安全覆盖', error_message)

    def test_resolve_assessment_assignments_skips_rows_depending_on_missing_primary_store(self):
        parsed_rows = [
            ParsedAssessmentAssignmentRow(
                row_number=2,
                target_month='2026-05',
                target_month_start=date(2026, 5, 1),
                target_version='v1',
                store_key='999',
                store_name='缺失主店',
                subject_code='SUB001',
                assignment_role='主店',
                is_joint_assessment='Y',
                anchor_store_name=None,
                effective_start_date=date(2026, 5, 1),
                effective_end_date=date(9999, 12, 31),
                remark=None,
            ),
            ParsedAssessmentAssignmentRow(
                row_number=3,
                target_month='2026-05',
                target_month_start=date(2026, 5, 1),
                target_version='v1',
                store_key='1',
                store_name='深圳万象城店',
                subject_code='SUB001',
                assignment_role='快闪',
                is_joint_assessment='Y',
                anchor_store_name=None,
                effective_start_date=date(2026, 5, 1),
                effective_end_date=date(9999, 12, 31),
                remark=None,
            ),
        ]
        store_map = {
            '深圳万象城店'.lower(): {
                'store_id': 1,
                'store_code': 'S001',
                'store_name': '深圳万象城店',
            }
        }
        store_map_by_id = {
            1: {
                'store_id': 1,
                'store_code': 'S001',
                'store_name': '深圳万象城店',
            }
        }
        store_map_by_code = {
            's001': {
                'store_id': 1,
                'store_code': 'S001',
                'store_name': '深圳万象城店',
            }
        }

        resolved_rows, validation = _resolve_assessment_assignments(parsed_rows, store_map, store_map_by_id, store_map_by_code)

        self.assertEqual(resolved_rows, [])
        self.assertEqual(validation['missing_store_keys'], ['999'])
        self.assertEqual(validation['missing_store_names'], ['缺失主店'])
        self.assertEqual(validation['unanchored_rows'], [])
        self.assertEqual(len(validation['skipped_missing_primary_rows']), 1)
        self.assertEqual(validation['skipped_missing_primary_rows'][0]['store_name'], '深圳万象城店')

    def test_parse_assessment_assignment_sheet_requires_store_id_header(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / '202605考核数据配置表.xlsx'
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = '门店考核归属'
            worksheet.append(['目标月份', '目标版本', '门店名称', '主体编码', '归属角色', '是否共同考核'])
            worksheet.append(['2026-05', 'v1', '深圳万象城店', 'SUB001', '主店', 'Y'])
            workbook.save(workbook_path)

            with self.assertRaises(ValueError) as cm:
                _parse_optional_assessment_assignment_sheet(workbook_path, '2026-05')

        self.assertIn('门店ID', str(cm.exception))

    def test_parse_assessment_assignment_sheet_accepts_store_code_key(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            workbook_path = Path(temp_dir) / '202605考核数据配置表.xlsx'
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = '门店考核归属'
            worksheet.append(['目标月份', '目标版本', '门店ID', '门店名称', '主体编码', '归属角色', '是否共同考核'])
            worksheet.append(['2026-05', 'v1', 'RT050', '北京国贸店', 'SUBJ_BJGM', '主店', 'Y'])
            workbook.save(workbook_path)

            parsed_rows, _ = _parse_optional_assessment_assignment_sheet(workbook_path, '2026-05')

        self.assertEqual(len(parsed_rows), 1)
        self.assertEqual(parsed_rows[0].store_key, 'RT050')

    def test_resolve_assessment_assignments_accepts_store_code_when_name_changed(self):
        parsed_rows = [
            ParsedAssessmentAssignmentRow(
                row_number=2,
                target_month='2026-05',
                target_month_start=date(2026, 5, 1),
                target_version='v1',
                store_key='RT050',
                store_name='北京国贸经营体主店',
                subject_code='SUBJ_BJGM',
                assignment_role='主店',
                is_joint_assessment='Y',
                anchor_store_name=None,
                effective_start_date=date(2026, 5, 1),
                effective_end_date=date(2026, 5, 31),
                remark=None,
            )
        ]
        store_map = {
            '北京国贸店'.lower(): {
                'store_id': 318,
                'store_code': 'RT050',
                'store_name': '北京国贸店',
            }
        }
        store_map_by_id = {
            318: {
                'store_id': 318,
                'store_code': 'RT050',
                'store_name': '北京国贸店',
            }
        }
        store_map_by_code = {
            'rt050': {
                'store_id': 318,
                'store_code': 'RT050',
                'store_name': '北京国贸店',
            }
        }

        resolved_rows, validation = _resolve_assessment_assignments(parsed_rows, store_map, store_map_by_id, store_map_by_code)

        self.assertEqual(len(resolved_rows), 1)
        self.assertEqual(resolved_rows[0].store_id, 318)
        self.assertEqual(resolved_rows[0].store_name, '北京国贸店')
        self.assertEqual(resolved_rows[0].source_store_name, '北京国贸经营体主店')
        self.assertEqual(resolved_rows[0].source_store_key, 'RT050')
        self.assertEqual(validation['missing_store_keys'], [])
        self.assertEqual(validation['store_key_name_mismatch_rows'][0]['store_id'], 318)
        self.assertEqual(validation['store_key_name_mismatch_rows'][0]['source_store_key'], 'RT050')
        self.assertEqual(validation['store_key_name_mismatch_rows'][0]['excel_store_name'], '北京国贸经营体主店')
        self.assertEqual(validation['store_key_name_mismatch_rows'][0]['dim_store_name'], '北京国贸店')


if __name__ == '__main__':
    unittest.main()