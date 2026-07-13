# -*- coding: utf-8 -*-
"""从 NAS 云雀订单管理 Excel 提取达播候选集。"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from tools.nas_access import ensure_nas_path_access

DEFAULT_NAS_DIR = Path(r"\\192.168.0.151\hefang总部\14-数据中台\达播数据同步\云雀达播订单筛选")
DEFAULT_FILE_PATTERN = "订单管理*.xlsx"
DEFAULT_SHEET_NAME = "T_V_OMSONLINEORDER"
DABO_SOURCE_NAME = "yunque_order_management"
ORDER_ID_SUFFIX_RE = re.compile(r"-C\d+$", re.IGNORECASE)
REQUIRED_HEADERS = (
    "平台单号",
    "系统单号",
    "平台",
    "状态",
    "商品数量",
    "商品编码",
    "主播名称",
)


@dataclass(frozen=True)
class CandidateRow:
    source_row_number: int
    system_order_id: str
    platform_order_id: str
    platform_name: str
    platform_code: str
    order_status: str
    influencer_id: str | None
    influencer_name: str
    platform_ship_time: str | None
    product_alias_code: str
    qty: int
    source_file: str
    source_sheet: str


@dataclass(frozen=True)
class OrderLabelRow:
    source_file: str
    source_sheet: str
    source_file_mtime: str
    first_source_row_number: int
    source_row_count: int
    system_order_id: str
    platform_order_id: str
    is_dabo_order: int
    dabo_source: str
    dabo_channel_code: str
    dabo_channel_name: str
    influencer_id: str | None
    influencer_name: str
    order_status: str
    platform_ship_time: str | None


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return re.sub(r"\s+", " ", str(value).strip())


def _normalize_order_id(value: object) -> str:
    return _normalize_text(value)


def split_order_id_tokens(order_id: str) -> list[str]:
    return [token for token in (_normalize_order_id(part) for part in str(order_id).split(",")) if token]


def normalize_order_id_token(token: str) -> str:
    return ORDER_ID_SUFFIX_RE.sub("", _normalize_order_id(token))


def build_order_id_token_key(order_id: str) -> tuple[str, ...]:
    normalized_tokens: list[str] = []
    for token in split_order_id_tokens(order_id):
        normalized_token = normalize_order_id_token(token)
        if normalized_token and normalized_token not in normalized_tokens:
            normalized_tokens.append(normalized_token)
    return tuple(sorted(normalized_tokens))


def _normalize_sku(value: object) -> str:
    return _normalize_text(value).replace("\t", "")


def _parse_int(value: object) -> tuple[int, str | None]:
    text_value = _normalize_text(value)
    if not text_value:
        return 0, None

    cleaned = text_value.replace(",", "")
    try:
        return int(Decimal(cleaned)), None
    except (InvalidOperation, ValueError):
        return 0, f"商品数量非法: {text_value}"


def _normalize_datetime(value: object) -> str | None:
    text_value = _normalize_text(value)
    return text_value or None


def _pick_single_value(rows: list[CandidateRow], attr_name: str) -> str | None:
    values = []
    for row in rows:
        value = getattr(row, attr_name)
        if value in (None, ""):
            continue
        if value not in values:
            values.append(value)
    return values[0] if values else None


def _resolve_input_file(explicit_file: str | None, nas_dir: Path, pattern: str) -> Path:
    if explicit_file:
        file_path = Path(explicit_file)
        if not file_path.is_absolute():
            file_path = REPO_ROOT / file_path
        ensure_nas_path_access(file_path)
        if not file_path.exists() or not file_path.is_file():
            raise FileNotFoundError(f"未找到 Excel 文件: {file_path}")
        return file_path

    ensure_nas_path_access(nas_dir)
    if not nas_dir.exists() or not nas_dir.is_dir():
        raise FileNotFoundError(f"未找到 NAS 目录: {nas_dir}")

    candidates = [
        path for path in nas_dir.glob(pattern)
        if path.is_file() and not path.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError(f"目录 {nas_dir} 下未找到匹配文件: {pattern}")

    return max(candidates, key=lambda path: (path.stat().st_mtime, path.name))


def _resolve_output_path(path_value: str) -> Path:
    path = Path(path_value)
    if not path.is_absolute():
        path = REPO_ROOT / path
    path.parent.mkdir(parents=True, exist_ok=True)
    return path


def _build_header_map(header_row: tuple[object, ...]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, value in enumerate(header_row):
        header = _normalize_text(value)
        if header:
            header_map[header] = index

    missing_headers = [header for header in REQUIRED_HEADERS if header not in header_map]
    if missing_headers:
        raise ValueError(f"工作表缺少必填列: {', '.join(missing_headers)}")

    return header_map


def _get_cell(row: tuple[object, ...], index: int) -> object | None:
    if index >= len(row):
        return None
    return row[index]


def _normalize_platform_code(platform_name: str) -> str:
    text_value = platform_name.strip()
    if not text_value:
        return "unknown"
    if text_value in {"抖音", "抖店"}:
        return "dy"
    if text_value.startswith("小红书"):
        return "xhs"
    if text_value in {"视频号", "视频号小店"}:
        return "sph"
    if text_value in {"淘宝", "天猫"}:
        return "tm"
    if "京东" in text_value:
        return "jd"
    if "唯品" in text_value:
        return "vip"
    return "unknown"


def _evaluate_row(row_data: dict[str, Any]) -> tuple[list[str], int]:
    reasons: list[str] = []
    qty, qty_error = _parse_int(row_data["商品数量"])

    if row_data["状态"] != "平台发货":
        reasons.append("状态非平台发货")
    if not row_data["主播名称"]:
        reasons.append("主播名称为空")
    elif row_data["主播名称"].upper().startswith("HEFANG"):
        reasons.append("主播名称以HEFANG开头")
    if not row_data["系统单号"]:
        reasons.append("缺少系统单号")
    if not row_data["商品编码"]:
        reasons.append("缺少商品编码")
    if qty_error:
        reasons.append(qty_error)

    return reasons, qty


def _build_summary(
    file_path: Path,
    sheet_name: str,
    selected_rows: list[CandidateRow],
    order_labels: list[OrderLabelRow],
    label_conflicts: list[dict[str, Any]],
    total_rows: int,
    mismatch_rows: int,
    selected_rows_missing_ship_time: int,
    reason_counter: Counter[str],
    platform_counter: Counter[str],
    preview_limit: int,
) -> dict[str, Any]:
    selected_system_orders = {row.system_order_id for row in selected_rows}
    selected_platform_orders = {row.platform_order_id for row in selected_rows if row.platform_order_id}
    selected_system_sku_pairs = {
        (row.system_order_id, row.product_alias_code)
        for row in selected_rows
    }
    unknown_platform_rows = sum(1 for row in selected_rows if row.platform_code == "unknown")

    summary: dict[str, Any] = {
        "source_file": str(file_path),
        "source_file_name": file_path.name,
        "source_sheet": sheet_name,
        "source_file_mtime": datetime.fromtimestamp(file_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "total_rows": total_rows,
        "mismatch_rows": mismatch_rows,
        "selected_rows": len(selected_rows),
        "selected_system_orders": len(selected_system_orders),
        "selected_platform_orders": len(selected_platform_orders),
        "selected_system_sku_pairs": len(selected_system_sku_pairs),
        "selected_order_labels": len(order_labels),
        "label_conflict_order_count": len(label_conflicts),
        "selected_rows_missing_ship_time": selected_rows_missing_ship_time,
        "selected_rows_unknown_platform": unknown_platform_rows,
        "platform_distribution": dict(platform_counter.most_common()),
        "reject_reason_distribution": dict(reason_counter.most_common()),
    }
    if preview_limit > 0:
        summary["preview_rows"] = [asdict(row) for row in selected_rows[:preview_limit]]
        summary["preview_order_labels"] = [asdict(row) for row in order_labels[:preview_limit]]
        summary["label_conflict_order_preview"] = label_conflicts[:preview_limit]
    return summary


def build_order_labels(selected_rows: list[CandidateRow], source_file_mtime: str) -> tuple[list[OrderLabelRow], list[dict[str, Any]]]:
    grouped_rows: dict[str, list[CandidateRow]] = {}
    for row in selected_rows:
        grouped_rows.setdefault(row.system_order_id, []).append(row)

    order_labels: list[OrderLabelRow] = []
    label_conflicts: list[dict[str, Any]] = []
    conflict_attrs = (
        "platform_order_id",
        "platform_code",
        "platform_name",
        "influencer_id",
        "influencer_name",
        "order_status",
    )

    for system_order_id in sorted(grouped_rows):
        rows = grouped_rows[system_order_id]
        first_row = min(rows, key=lambda item: item.source_row_number)

        conflict_fields: dict[str, list[str]] = {}
        for attr_name in conflict_attrs:
            values = []
            for row in rows:
                value = getattr(row, attr_name)
                if value in (None, ""):
                    continue
                if value not in values:
                    values.append(value)
            if len(values) > 1:
                conflict_fields[attr_name] = values

        if conflict_fields:
            label_conflicts.append(
                {
                    "system_order_id": system_order_id,
                    "conflict_fields": conflict_fields,
                    "source_row_numbers": [row.source_row_number for row in rows],
                }
            )

        order_labels.append(
            OrderLabelRow(
                source_file=first_row.source_file,
                source_sheet=first_row.source_sheet,
                source_file_mtime=source_file_mtime,
                first_source_row_number=min(row.source_row_number for row in rows),
                source_row_count=len(rows),
                system_order_id=system_order_id,
                platform_order_id=_pick_single_value(rows, "platform_order_id") or "",
                is_dabo_order=1,
                dabo_source=DABO_SOURCE_NAME,
                dabo_channel_code=_pick_single_value(rows, "platform_code") or "unknown",
                dabo_channel_name=_pick_single_value(rows, "platform_name") or "",
                influencer_id=_pick_single_value(rows, "influencer_id"),
                influencer_name=_pick_single_value(rows, "influencer_name") or "",
                order_status=_pick_single_value(rows, "order_status") or "",
                platform_ship_time=_pick_single_value(rows, "platform_ship_time"),
            )
        )

    return order_labels, label_conflicts


def extract_candidates(file_path: Path, sheet_name: str, preview_limit: int) -> tuple[dict[str, Any], list[CandidateRow], list[OrderLabelRow]]:
    ensure_nas_path_access(file_path)
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"未找到工作表 {sheet_name}，可选 sheet: {', '.join(workbook.sheetnames)}")

    worksheet = workbook[sheet_name]
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise ValueError("Excel 缺少表头")

    header_map = _build_header_map(header_row)
    selected_rows: list[CandidateRow] = []
    reason_counter: Counter[str] = Counter()
    platform_counter: Counter[str] = Counter()
    total_rows = 0
    mismatch_rows = 0
    selected_rows_missing_ship_time = 0

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if row is None:
            continue
        if all(cell is None or _normalize_text(cell) == "" for cell in row):
            continue

        total_rows += 1
        row_data = {
            "平台单号": _normalize_order_id(_get_cell(row, header_map["平台单号"])),
            "系统单号": _normalize_order_id(_get_cell(row, header_map["系统单号"])),
            "平台": _normalize_text(_get_cell(row, header_map["平台"])),
            "状态": _normalize_text(_get_cell(row, header_map["状态"])),
            "商品数量": _get_cell(row, header_map["商品数量"]),
            "商品编码": _normalize_sku(_get_cell(row, header_map["商品编码"])),
            "主播名称": _normalize_text(_get_cell(row, header_map["主播名称"])),
            "主播ID": _normalize_text(_get_cell(row, header_map["主播ID"])) if "主播ID" in header_map else "",
            "平台发货时间": _normalize_datetime(_get_cell(row, header_map["平台发货时间"])) if "平台发货时间" in header_map else None,
            "订单总额": _normalize_text(_get_cell(row, header_map["订单总额"])) if "订单总额" in header_map else "",
        }

        if row_data["平台单号"] and row_data["系统单号"] and row_data["平台单号"] != row_data["系统单号"]:
            mismatch_rows += 1

        reasons, qty = _evaluate_row(row_data)
        if reasons:
            reason_counter.update(reasons)
            continue

        platform_code = _normalize_platform_code(row_data["平台"])
        if not row_data["平台发货时间"]:
            selected_rows_missing_ship_time += 1

        candidate_row = CandidateRow(
            source_row_number=row_number,
            system_order_id=row_data["系统单号"],
            platform_order_id=row_data["平台单号"],
            platform_name=row_data["平台"],
            platform_code=platform_code,
            order_status=row_data["状态"],
            influencer_id=row_data["主播ID"] or None,
            influencer_name=row_data["主播名称"],
            platform_ship_time=row_data["平台发货时间"],
            product_alias_code=row_data["商品编码"],
            qty=qty,
            source_file=file_path.name,
            source_sheet=sheet_name,
        )
        selected_rows.append(candidate_row)
        platform_counter.update([candidate_row.platform_name])

    source_file_mtime = datetime.fromtimestamp(file_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
    order_labels, label_conflicts = build_order_labels(selected_rows, source_file_mtime)

    summary = _build_summary(
        file_path=file_path,
        sheet_name=sheet_name,
        selected_rows=selected_rows,
        order_labels=order_labels,
        label_conflicts=label_conflicts,
        total_rows=total_rows,
        mismatch_rows=mismatch_rows,
        selected_rows_missing_ship_time=selected_rows_missing_ship_time,
        reason_counter=reason_counter,
        platform_counter=platform_counter,
        preview_limit=preview_limit,
    )
    return summary, selected_rows, order_labels


def _write_json(path: Path, payload: dict[str, Any]) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_csv(path: Path, rows: list[CandidateRow]) -> None:
    fieldnames = list(CandidateRow.__dataclass_fields__.keys())
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def _write_order_label_csv(path: Path, rows: list[OrderLabelRow]) -> None:
    fieldnames = list(OrderLabelRow.__dataclass_fields__.keys())
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(asdict(row))


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="从 NAS 云雀订单管理 Excel 提取达播候选集与订单标签")
    parser.add_argument("--file", help="指定 Excel 文件路径；不传时默认扫描 NAS 最新订单管理*.xlsx")
    parser.add_argument("--nas-dir", default=str(DEFAULT_NAS_DIR), help="NAS 目录，默认读取云雀达播订单筛选目录")
    parser.add_argument("--pattern", default=DEFAULT_FILE_PATTERN, help="Excel 文件匹配模式")
    parser.add_argument("--sheet", default=DEFAULT_SHEET_NAME, help="工作表名称")
    parser.add_argument("--preview-limit", type=int, default=5, help="输出预览行数，默认 5")
    parser.add_argument("--report-json", help="将汇总 JSON 写入指定路径")
    parser.add_argument("--export-csv", help="将筛选后的候选集写入指定 CSV 路径")
    parser.add_argument("--export-order-label-csv", help="将去重后的订单标签写入指定 CSV 路径")
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    file_path = _resolve_input_file(
        explicit_file=args.file,
        nas_dir=Path(args.nas_dir),
        pattern=args.pattern,
    )
    summary, selected_rows, order_labels = extract_candidates(
        file_path=file_path,
        sheet_name=args.sheet,
        preview_limit=max(args.preview_limit, 0),
    )

    if args.report_json:
        report_path = _resolve_output_path(args.report_json)
        _write_json(report_path, summary)

    if args.export_csv:
        export_path = _resolve_output_path(args.export_csv)
        _write_csv(export_path, selected_rows)

    if args.export_order_label_csv:
        export_path = _resolve_output_path(args.export_order_label_csv)
        _write_order_label_csv(export_path, order_labels)

    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()