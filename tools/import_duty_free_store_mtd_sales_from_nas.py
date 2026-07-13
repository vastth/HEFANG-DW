# -*- coding: utf-8 -*-
"""从 NAS 导入免税门店外部月累计销售额快照。"""

from __future__ import annotations

import argparse
import hashlib
import json
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from pymysql.cursors import DictCursor


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from db_connections import connect_mysql
from tools.nas_access import ensure_nas_path_access


DEFAULT_NAS_FILE_PATH = Path(r"\\192.168.0.151\hefang总部\14-数据中台\销售部\目标配置表\免税门店月累计销售.xlsx")
DEFAULT_SHEET_NAME = '免税月累计'
DEFAULT_CREATED_BY = 'duty_free_mtd_sales_import'
TARGET_TABLE_NAME = 'cfg_duty_free_store_mtd_sales'
LOG_TABLE_NAME = 'log_duty_free_store_mtd_sales_import'
TARGET_TABLE_SQL_PATH = REPO_ROOT / 'SQL' / 'create_cfg_duty_free_store_mtd_sales.sql'
LOG_TABLE_SQL_PATH = REPO_ROOT / 'SQL' / 'create_log_duty_free_store_mtd_sales_import.sql'
REQUIRED_HEADERS = ('目标月份', '数据版本', '门店ID', '门店名称', '渠道类型', '月累计')
DECIMAL_ZERO = Decimal('0.00')


@dataclass(frozen=True)
class DutyFreeSourceRow:
    row_number: int
    target_month_start: date
    data_version: str
    store_key: str
    store_name: str
    report_channel_type: str
    external_mtd_sales_amt: Decimal


def _connect():
    return connect_mysql(cursorclass=DictCursor, autocommit=False)


def _normalize_text(value: object) -> str:
    if value is None:
        return ''
    return ' '.join(str(value).strip().split())


def _parse_target_month(value: object, row_number: int) -> date:
    if isinstance(value, datetime):
        return value.date().replace(day=1)
    if isinstance(value, date):
        return value.replace(day=1)
    text_value = _normalize_text(value)
    if not text_value:
        raise ValueError(f'第 {row_number} 行缺少 目标月份')
    normalized_value = text_value.replace('/', '-')
    try:
        if len(normalized_value) == 7:
            normalized_value = normalized_value + '-01'
        return date.fromisoformat(normalized_value).replace(day=1)
    except ValueError as exc:
        raise ValueError(f'第 {row_number} 行 目标月份 格式非法: {text_value}') from exc


def _parse_decimal(value: object, row_number: int, header_name: str, *, blank_as_zero: bool = False) -> Decimal:
    text_value = _normalize_text(value)
    if not text_value:
        if blank_as_zero:
            return DECIMAL_ZERO
        raise ValueError(f'第 {row_number} 行 {header_name} 不能为空')
    try:
        return Decimal(text_value).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
    except InvalidOperation as exc:
        raise ValueError(f'第 {row_number} 行 {header_name} 不是合法数字: {text_value}') from exc


def _parse_store_key(value: object, row_number: int) -> str:
    text_value = _normalize_text(value)
    if not text_value:
        raise ValueError(f'第 {row_number} 行 门店ID 不能为空')
    try:
        numeric_value = Decimal(text_value)
        if numeric_value == numeric_value.to_integral_value():
            return str(int(numeric_value))
    except InvalidOperation:
        pass
    return text_value


def _split_store_keys(store_keys: list[str]) -> tuple[list[int], list[str]]:
    store_ids: list[int] = []
    store_codes: list[str] = []
    for store_key in store_keys:
        try:
            numeric_value = Decimal(store_key)
            if numeric_value == numeric_value.to_integral_value():
                store_ids.append(int(numeric_value))
                continue
        except InvalidOperation:
            pass
        store_codes.append(store_key)
    return store_ids, store_codes


def _compute_file_md5(file_path: Path) -> str:
    md5_hash = hashlib.md5()
    with file_path.open('rb') as file_obj:
        for chunk in iter(lambda: file_obj.read(1024 * 1024), b''):
            md5_hash.update(chunk)
    return md5_hash.hexdigest()


def _resolve_input_file(file_path_arg: str | None) -> Path:
    if file_path_arg:
        file_path = Path(file_path_arg)
    else:
        ensure_nas_path_access(str(DEFAULT_NAS_FILE_PATH.parent))
        file_path = DEFAULT_NAS_FILE_PATH

    if not file_path.exists():
        raise ValueError(f'未找到免税月累计文件: {file_path}')
    return file_path


def _load_workbook_sheetnames(file_path: Path) -> list[str]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _parse_workbook(file_path: Path, sheet_name: str) -> tuple[list[DutyFreeSourceRow], dict]:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f'未找到工作表 {sheet_name}；当前可用工作表: {", ".join(workbook.sheetnames)}'
            )
        worksheet = workbook[sheet_name]
        rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        raise ValueError(f'工作表 {sheet_name} 为空')

    header = [_normalize_text(value) for value in rows[0]]
    missing_headers = [header_name for header_name in REQUIRED_HEADERS if header_name not in header]
    if missing_headers:
        raise ValueError(f'工作表 {sheet_name} 缺少表头: {", ".join(missing_headers)}')

    header_index_map = {header_name: header.index(header_name) for header_name in REQUIRED_HEADERS}
    parsed_rows: list[DutyFreeSourceRow] = []
    seen_store_keys: set[str] = set()
    target_months: set[date] = set()
    data_versions: set[str] = set()

    for row_number, row in enumerate(rows[1:], start=2):
        if all(_normalize_text(value) == '' for value in row):
            continue

        target_month_start = _parse_target_month(row[header_index_map['目标月份']], row_number)
        data_version = _normalize_text(row[header_index_map['数据版本']])
        if not data_version:
            raise ValueError(f'第 {row_number} 行 数据版本 不能为空')

        store_key = _parse_store_key(row[header_index_map['门店ID']], row_number)
        if store_key in seen_store_keys:
            raise ValueError(f'第 {row_number} 行 门店ID={store_key} 与文件内其他行重复')
        seen_store_keys.add(store_key)

        store_name = _normalize_text(row[header_index_map['门店名称']])
        if not store_name:
            raise ValueError(f'第 {row_number} 行 门店名称 不能为空')

        report_channel_type = _normalize_text(row[header_index_map['渠道类型']])
        if not report_channel_type:
            raise ValueError(f'第 {row_number} 行 渠道类型 不能为空')

        external_mtd_sales_amt = _parse_decimal(
            row[header_index_map['月累计']],
            row_number,
            '月累计',
            blank_as_zero=True,
        )

        target_months.add(target_month_start)
        data_versions.add(data_version)
        parsed_rows.append(
            DutyFreeSourceRow(
                row_number=row_number,
                target_month_start=target_month_start,
                data_version=data_version,
                store_key=store_key,
                store_name=store_name,
                report_channel_type=report_channel_type,
                external_mtd_sales_amt=external_mtd_sales_amt,
            )
        )

    if not parsed_rows:
        raise ValueError(f'工作表 {sheet_name} 没有有效数据行')
    if len(target_months) != 1:
        raise ValueError('同一份免税月累计文件只能包含一个 目标月份')
    if len(data_versions) != 1:
        raise ValueError('同一份免税月累计文件只能包含一个 数据版本')

    target_month_start = next(iter(target_months))
    data_version = next(iter(data_versions))
    return parsed_rows, {
        'target_month': target_month_start.strftime('%Y-%m'),
        'target_month_start': target_month_start.isoformat(),
        'data_version': data_version,
        'source_row_count': len(parsed_rows),
    }


def _fetch_store_validation_map(conn, store_keys: list[str], target_month_start: date) -> dict[str, dict]:
    if not store_keys:
        return {}
    store_ids, store_codes = _split_store_keys(store_keys)
    store_conditions = []
    store_params: list[object] = []
    if store_ids:
        placeholders = ', '.join(['%s'] * len(store_ids))
        store_conditions.append(f'store_id IN ({placeholders})')
        store_params.extend(store_ids)
    if store_codes:
        placeholders = ', '.join(['%s'] * len(store_codes))
        store_conditions.append(f'store_code IN ({placeholders})')
        store_params.extend(store_codes)
    store_filter = ' OR '.join(store_conditions)
    sql = f"""
        WITH store_candidates AS (
            SELECT
                store_id,
                store_code,
                store_name
            FROM dim_store
            WHERE {store_filter}
        ),
        attr_candidates AS (
            SELECT
                sra.store_id,
                sra.store_name,
                sra.report_channel_type,
                sra.is_duty_free,
                ROW_NUMBER() OVER (
                    PARTITION BY sra.store_id
                    ORDER BY
                        CASE
                            WHEN %s BETWEEN sra.effective_start_date AND sra.effective_end_date THEN 0
                            ELSE 1
                        END,
                        sra.effective_end_date DESC,
                        sra.effective_start_date DESC
                ) AS attr_recency_rank
            FROM dim_store_report_attr sra
            INNER JOIN store_candidates sc
              ON sra.store_id = sc.store_id
            WHERE 1 = 1
              AND sra.effective_start_date <= %s
        )
        SELECT
            sc.store_id,
            sc.store_code,
            sc.store_name AS dim_store_name,
            ac.store_name AS attr_store_name,
            ac.report_channel_type,
            ac.is_duty_free
        FROM store_candidates sc
        LEFT JOIN attr_candidates ac
          ON sc.store_id = ac.store_id
         AND ac.attr_recency_rank = 1
    """
    params = [*store_params, target_month_start, target_month_start]
    with conn.cursor() as cursor:
        cursor.execute(sql, params)
        rows = cursor.fetchall()
    validation_map: dict[str, dict] = {}
    for row in rows:
        validation_map[str(int(row['store_id']))] = row
        store_code = _normalize_text(row.get('store_code'))
        if store_code:
            validation_map[store_code] = row
    return validation_map


def _validate_rows(conn, parsed_rows: list[DutyFreeSourceRow]) -> tuple[list[dict], dict]:
    target_month_start = parsed_rows[0].target_month_start
    validation_map = _fetch_store_validation_map(conn, [row.store_key for row in parsed_rows], target_month_start)
    missing_store_keys: list[str] = []
    missing_attr_store_ids: list[int] = []
    non_duty_free_store_ids: list[int] = []
    name_mismatches: list[str] = []
    channel_type_mismatches: list[str] = []
    resolved_rows: list[dict] = []

    for row in parsed_rows:
        store_info = validation_map.get(row.store_key)
        if store_info is None:
            missing_store_keys.append(row.store_key)
            continue
        resolved_store_id = int(store_info['store_id'])
        if store_info.get('report_channel_type') is None:
            missing_attr_store_ids.append(resolved_store_id)
            continue
        if _normalize_text(store_info.get('is_duty_free')) != 'Y':
            non_duty_free_store_ids.append(resolved_store_id)
        dim_store_name = _normalize_text(store_info.get('dim_store_name'))
        if dim_store_name and row.store_name != dim_store_name:
            name_mismatches.append(
                f'门店ID={row.store_key}, store_id={resolved_store_id}: 文件={row.store_name} / dim_store={dim_store_name}'
            )
        current_channel_type = _normalize_text(store_info.get('report_channel_type'))
        if current_channel_type and row.report_channel_type != current_channel_type:
            channel_type_mismatches.append(
                f'门店ID={row.store_key}, store_id={resolved_store_id}: 文件={row.report_channel_type} / dim_store_report_attr={current_channel_type}'
            )

        resolved_rows.append(
            {
                'target_month': row.target_month_start,
                'data_version': row.data_version,
                'store_id': resolved_store_id,
                'store_code': _normalize_text(store_info.get('store_code')),
                'store_name': row.store_name,
                'report_channel_type': row.report_channel_type,
                'external_mtd_sales_amt': row.external_mtd_sales_amt,
            }
        )

    validation_errors = []
    if missing_store_keys:
        validation_errors.append(f'未命中 dim_store 的门店ID/编码: {missing_store_keys}')
    if missing_attr_store_ids:
        validation_errors.append(f'未命中当前有效 dim_store_report_attr 的门店ID: {missing_attr_store_ids}')
    if non_duty_free_store_ids:
        validation_errors.append(f'以下门店当前 is_duty_free 不是 Y: {non_duty_free_store_ids}')
    if name_mismatches:
        validation_errors.append('门店名称不一致: ' + '；'.join(name_mismatches))
    if channel_type_mismatches:
        validation_errors.append('渠道类型不一致: ' + '；'.join(channel_type_mismatches))
    if validation_errors:
        raise ValueError(' | '.join(validation_errors))

    return resolved_rows, {
        'matched_store_count': len(resolved_rows),
    }


def _fetch_existing_snapshot_map(conn, target_month: date, data_version: str) -> dict[int, dict]:
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                store_id,
                store_name,
                report_channel_type,
                external_mtd_sales_amt
            FROM {TARGET_TABLE_NAME}
                        WHERE target_month = %s
              AND data_version = %s
            """,
                        (target_month, data_version),
        )
        rows = cursor.fetchall()
    return {int(row['store_id']): row for row in rows}


def _build_snapshot_diff_summary(resolved_rows: list[dict], existing_snapshot_map: dict[int, dict]) -> dict:
    changed_store_count = 0
    new_store_count = 0
    unchanged_store_count = 0
    next_store_ids = set()
    for row in resolved_rows:
        store_id = int(row['store_id'])
        next_store_ids.add(store_id)
        existing_row = existing_snapshot_map.get(store_id)
        if existing_row is None:
            new_store_count += 1
            continue
        if (
            _normalize_text(existing_row.get('store_name')) == row['store_name']
            and _normalize_text(existing_row.get('report_channel_type')) == row['report_channel_type']
            and Decimal(str(existing_row.get('external_mtd_sales_amt'))).quantize(Decimal('0.01')) == row['external_mtd_sales_amt']
        ):
            unchanged_store_count += 1
        else:
            changed_store_count += 1

    exited_store_count = len(set(existing_snapshot_map.keys()) - next_store_ids)
    return {
        'existing_snapshot_row_count': len(existing_snapshot_map),
        'new_store_count': new_store_count,
        'changed_store_count': changed_store_count,
        'unchanged_store_count': unchanged_store_count,
        'exited_store_count': exited_store_count,
        'has_changes': (new_store_count + changed_store_count + exited_store_count) > 0,
    }


def _ensure_required_tables_exist(conn) -> None:
    with conn.cursor() as cursor:
        cursor.execute(
            """
            SELECT table_name AS table_name_alias
            FROM information_schema.tables
            WHERE table_schema = DATABASE()
              AND table_name IN (%s, %s)
            """,
            (TARGET_TABLE_NAME, LOG_TABLE_NAME),
        )
        rows = cursor.fetchall()
    existing_table_names = {row['table_name_alias'] for row in rows}
    if TARGET_TABLE_NAME not in existing_table_names:
        ddl_hint = TARGET_TABLE_SQL_PATH.relative_to(REPO_ROOT).as_posix()
        raise ValueError(f'未找到目标表 {TARGET_TABLE_NAME}，请先执行 {ddl_hint}')
    if LOG_TABLE_NAME not in existing_table_names:
        ddl_hint = LOG_TABLE_SQL_PATH.relative_to(REPO_ROOT).as_posix()
        raise ValueError(f'未找到日志表 {LOG_TABLE_NAME}，请先执行 {ddl_hint}')


def _insert_log_row(
    conn,
    *,
    file_path: Path,
    file_md5: str,
    sheet_name: str,
    target_month: date | None,
    data_version: str | None,
    store_count: int,
    records_total: int,
    records_inserted: int,
    changed_store_count: int,
    new_store_count: int,
    exited_store_count: int,
    status: str,
    message: str,
    started_at: datetime,
    finished_at: datetime,
) -> None:
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            INSERT INTO {LOG_TABLE_NAME} (
                file_name,
                file_path,
                file_md5,
                source_sheet,
                target_month,
                data_version,
                store_count,
                records_total,
                records_inserted,
                changed_store_count,
                new_store_count,
                exited_store_count,
                status,
                message,
                started_at,
                finished_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """,
            (
                file_path.name,
                str(file_path),
                file_md5,
                sheet_name,
                target_month,
                data_version,
                store_count,
                records_total,
                records_inserted,
                changed_store_count,
                new_store_count,
                exited_store_count,
                status,
                message,
                started_at,
                finished_at,
            ),
        )


def _apply_rows(conn, resolved_rows: list[dict], file_path: Path, file_md5: str, created_by: str) -> int:
    target_month = resolved_rows[0]['target_month']
    data_version = resolved_rows[0]['data_version']
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            DELETE FROM {TARGET_TABLE_NAME}
            WHERE target_month = %s
              AND data_version = %s
            """,
            (target_month, data_version),
        )
        insert_sql = f"""
            INSERT INTO {TARGET_TABLE_NAME} (
                target_month,
                data_version,
                store_id,
                store_code,
                store_name,
                report_channel_type,
                external_mtd_sales_amt,
                source_file_name,
                source_file_md5,
                created_by
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        insert_rows = [
            (
                row['target_month'],
                row['data_version'],
                row['store_id'],
                row['store_code'],
                row['store_name'],
                row['report_channel_type'],
                row['external_mtd_sales_amt'],
                file_path.name,
                file_md5,
                created_by,
            )
            for row in resolved_rows
        ]
        cursor.executemany(insert_sql, insert_rows)
    return len(resolved_rows)


def _write_output_json(output_path: Path, payload: dict) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding='utf-8')


def _build_summary_payload(
    *,
    mode: str,
    file_path: Path,
    file_md5: str,
    sheet_name: str,
    workbook_summary: dict,
    validation_summary: dict,
    diff_summary: dict,
    records_inserted: int,
) -> dict:
    return {
        'mode': mode,
        'file_path': str(file_path),
        'file_name': file_path.name,
        'file_md5': file_md5,
        'sheet_name': sheet_name,
        'validation_status': 'SUCCESS',
        'target_month': workbook_summary['target_month'],
        'target_month_start': workbook_summary['target_month_start'],
        'data_version': workbook_summary['data_version'],
        'source_row_count': workbook_summary['source_row_count'],
        'matched_store_count': validation_summary['matched_store_count'],
        'existing_snapshot_row_count': diff_summary['existing_snapshot_row_count'],
        'unchanged_store_count': diff_summary['unchanged_store_count'],
        'changed_store_count': diff_summary['changed_store_count'],
        'new_store_count': diff_summary['new_store_count'],
        'exited_store_count': diff_summary['exited_store_count'],
        'has_changes': diff_summary['has_changes'],
        'records_inserted': records_inserted,
    }


def _write_failure_output_json(
    output_json: str | None,
    args: argparse.Namespace,
    error_message: str,
    validation_status: str,
) -> None:
    if not output_json:
        return
    payload = {
        'mode': 'apply' if args.apply else 'dry-run',
        'sheet_name': args.sheet_name,
        'file_path': args.file_path,
        'validation_status': validation_status,
        'error_message': error_message,
    }
    _write_output_json(Path(output_json), payload)


def main() -> int:
    parser = argparse.ArgumentParser(description='从 NAS 导入免税门店外部月累计销售额，默认只做 dry-run')
    parser.add_argument('--file-path', default=None, help='可选：显式指定免税月累计 Excel 路径')
    parser.add_argument('--sheet-name', default=DEFAULT_SHEET_NAME, help='导入工作表名称')
    parser.add_argument('--preview-limit', type=int, default=10, help='dry-run 输出预览行数')
    parser.add_argument('--output-json', help='可选：将 dry-run / apply 摘要写入 JSON 文件')
    parser.add_argument('--created-by', default=DEFAULT_CREATED_BY, help='正式导入时写入 created_by')
    parser.add_argument('--apply', action='store_true', help='显式启用写库模式；默认只做 dry-run')
    args = parser.parse_args()

    conn = None
    started_at = datetime.now()
    file_path: Path | None = None
    workbook_summary: dict | None = None
    try:
        file_path = _resolve_input_file(args.file_path)
        _load_workbook_sheetnames(file_path)
        file_md5 = _compute_file_md5(file_path)
        parsed_rows, workbook_summary = _parse_workbook(file_path, args.sheet_name)

        conn = _connect()
        resolved_rows, validation_summary = _validate_rows(conn, parsed_rows)
        existing_snapshot_map = _fetch_existing_snapshot_map(
            conn,
            parsed_rows[0].target_month_start,
            parsed_rows[0].data_version,
        )
        diff_summary = _build_snapshot_diff_summary(resolved_rows, existing_snapshot_map)

        preview_rows = [
            {
                'target_month': row['target_month'].strftime('%Y-%m'),
                'data_version': row['data_version'],
                'store_id': row['store_id'],
                'store_code': row['store_code'],
                'store_name': row['store_name'],
                'report_channel_type': row['report_channel_type'],
                'external_mtd_sales_amt': str(row['external_mtd_sales_amt']),
            }
            for row in resolved_rows[:max(args.preview_limit, 0)]
        ]

        if not args.apply:
            summary = _build_summary_payload(
                mode='dry-run',
                file_path=file_path,
                file_md5=file_md5,
                sheet_name=args.sheet_name,
                workbook_summary=workbook_summary,
                validation_summary=validation_summary,
                diff_summary=diff_summary,
                records_inserted=0,
            )
            summary['preview_rows'] = preview_rows
            if args.output_json:
                _write_output_json(Path(args.output_json), summary)
            print(json.dumps(summary, ensure_ascii=False, indent=2))
            return 0

        _ensure_required_tables_exist(conn)
        records_inserted = _apply_rows(conn, resolved_rows, file_path, file_md5, args.created_by)
        finished_at = datetime.now()
        summary = _build_summary_payload(
            mode='apply',
            file_path=file_path,
            file_md5=file_md5,
            sheet_name=args.sheet_name,
            workbook_summary=workbook_summary,
            validation_summary=validation_summary,
            diff_summary=diff_summary,
            records_inserted=records_inserted,
        )
        summary['preview_rows'] = preview_rows
        _insert_log_row(
            conn,
            file_path=file_path,
            file_md5=file_md5,
            sheet_name=args.sheet_name,
            target_month=parsed_rows[0].target_month_start,
            data_version=parsed_rows[0].data_version,
            store_count=validation_summary['matched_store_count'],
            records_total=workbook_summary['source_row_count'],
            records_inserted=records_inserted,
            changed_store_count=diff_summary['changed_store_count'],
            new_store_count=diff_summary['new_store_count'],
            exited_store_count=diff_summary['exited_store_count'],
            status='SUCCESS',
            message=(
                f"target_month={workbook_summary['target_month']}, data_version={workbook_summary['data_version']}, "
                f"matched={validation_summary['matched_store_count']}, inserted={records_inserted}, "
                f"changed={diff_summary['changed_store_count']}, new={diff_summary['new_store_count']}, "
                f"exited={diff_summary['exited_store_count']}"
            ),
            started_at=started_at,
            finished_at=finished_at,
        )
        conn.commit()
        if args.output_json:
            _write_output_json(Path(args.output_json), summary)
        print(json.dumps(summary, ensure_ascii=False, indent=2))
        return 0
    except Exception as exc:
        if conn is not None:
            conn.rollback()
        _write_failure_output_json(args.output_json, args, str(exc), 'FAILED')
        print(str(exc), file=sys.stderr)
        return 1
    finally:
        if conn is not None:
            conn.close()


if __name__ == '__main__':
    raise SystemExit(main())