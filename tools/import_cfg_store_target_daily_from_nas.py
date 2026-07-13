# -*- coding: utf-8 -*-
"""从 NAS 读取门店日报目标 Excel，并导入 cfg_store_target_daily。"""

from __future__ import annotations

import argparse
import calendar
import difflib
import hashlib
import json
import re
import sys
from typing import Any
from dataclasses import dataclass, replace
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

from openpyxl import load_workbook
from pymysql.cursors import DictCursor


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from db_connections import connect_mysql
from tools.nas_access import ensure_nas_path_access


DEFAULT_NAS_DIR = Path(r"\\192.168.0.151\hefang总部\14-数据中台\销售部\目标配置表")
DEFAULT_SHEET_NAME = "导入模板"
SUBJECT_TARGET_SHEET_NAME = "统计主体目标"
ASSESSMENT_ASSIGNMENT_SHEET_NAME = "门店考核归属"
INSTRUCTION_SHEET_NAME = "填写说明"
DEFAULT_CREATED_BY = "store_target_nas_import"
LOG_TABLE_NAME = "log_store_target_import"
TARGET_TABLE_NAME = "cfg_store_target_daily"
STORE_ATTR_TABLE_NAME = "dim_store_report_attr"
STORE_ATTR_SNAPSHOT_TABLE_NAME = "cfg_store_report_attr_snapshot"
SUBJECT_TARGET_TABLE_NAME = "cfg_store_assessment_subject_target_daily"
ASSESSMENT_ASSIGNMENT_TABLE_NAME = "cfg_store_assessment_assignment"
STORE_TABLE_NAME = "dim_store"
LOG_TABLE_SQL_PATH = REPO_ROOT / "SQL" / "create_log_store_target_import.sql"
STORE_ATTR_SNAPSHOT_SQL_PATH = REPO_ROOT / "SQL" / "create_store_report_attr_snapshot.sql"
DAY_HEADER_RE = re.compile(r"^(?P<day>[1-9]|[12]\d|3[01])日目标$")
MONTH_RE = re.compile(r"^(?P<year>\d{4})[-/年](?P<month>\d{1,2})(?:月)?$")
NAS_FILE_PATTERNS = (
    re.compile(
        r"^(?P<year>\d{4})年(?P<month>\d{1,2})月日目标配置表(?:_(?P<version>v[0-9A-Za-z._-]+))?\.xlsx$"
    ),
    re.compile(
        r"^(?P<year>\d{4})(?P<month>\d{2})考核数据配置表(?:_(?P<version>v[0-9A-Za-z._-]+))?\.xlsx$"
    ),
)
REQUIRED_HEADERS = ("目标月份", "目标版本", "门店名称", "月目标")
SUBJECT_TARGET_REQUIRED_HEADERS = ("目标月份", "目标版本", "主体编码", "主体名称", "考核模式", "月目标")
ASSESSMENT_ASSIGNMENT_REQUIRED_HEADERS = (
    "目标月份",
    "目标版本",
    "门店ID",
    "门店名称",
    "主体编码",
    "归属角色",
    "是否共同考核",
)
DECIMAL_ZERO = Decimal("0.00")
STORE_TYPE_HEADER = "门店类型"
STORE_GRADE_HEADER = "等级"
STORE_GRADE_HEADER_ALIASES = (STORE_GRADE_HEADER, "店铺等级", "门店等级")
SUBJECT_CODE_HEADER = "主体编码"
SUBJECT_NAME_HEADER = "主体名称"
ASSESSMENT_MODE_HEADER = "考核模式"
ASSIGNMENT_ROLE_HEADER = "归属角色"
IS_JOINT_ASSESSMENT_HEADER = "是否共同考核"
STORE_ID_HEADER = "门店ID"
ANCHOR_STORE_NAME_HEADER = "主店名称"
EFFECTIVE_START_DATE_HEADER = "生效开始日"
EFFECTIVE_END_DATE_HEADER = "生效结束日"
REMARK_HEADER = "备注"
REPORT_CHANNEL_TYPE_GROUP_COLUMN = "report_channel_type_group"
REPORT_CHANNEL_TYPE_GROUP_MAPPING = {
    "小程序": "小程序",
    "线上小程序": "小程序",
    "直营": "直营",
    "直营-奥莱": "直营",
    "联营": "联营",
    "联营-免税": "联营",
    "联营-奥莱": "联营",
}
ALLOWED_REPORT_CHANNEL_TYPES = tuple(REPORT_CHANNEL_TYPE_GROUP_MAPPING.keys())
ALLOWED_ASSESSMENT_MODES = ("独立", "合并")
ALLOWED_ASSIGNMENT_ROLES = ("主店", "快闪", "独立")
DUTY_FREE_KEYWORD = "免税"
DEFAULT_EFFECTIVE_END_DATE = date(9999, 12, 31)
OPTIONAL_SHEET_NAME_ALIASES = {
    ANCHOR_STORE_NAME_HEADER: (ANCHOR_STORE_NAME_HEADER, "挂靠主店名称", "挂靠门店名称"),
    EFFECTIVE_START_DATE_HEADER: (EFFECTIVE_START_DATE_HEADER, "开始日期"),
    EFFECTIVE_END_DATE_HEADER: (EFFECTIVE_END_DATE_HEADER, "结束日期"),
    REMARK_HEADER: (REMARK_HEADER, "说明"),
}
STORE_ATTR_COMPARE_FIELDS = (
    "store_code",
    "store_name",
    "report_channel_type",
    "store_grade",
    "is_duty_free",
    "is_include_in_daily_report",
)


@dataclass(frozen=True)
class NasTargetFile:
    file_path: Path
    target_month: str
    file_version: str | None
    last_modified_at: datetime


@dataclass(frozen=True)
class SourceRow:
    row_number: int
    target_month: str
    target_month_start: date
    target_version: str
    store_name: str
    store_type: str | None
    store_grade: str | None
    month_target: Decimal
    day_targets: dict[int, Decimal]
    blank_day_cell_count: int
    effective_start_date: date
    effective_end_date: date


@dataclass(frozen=True)
class ExpandedRow:
    target_date: date
    store_name: str
    store_id: int
    store_code: str | None
    month_target: Decimal
    day_target: Decimal
    target_version: str


@dataclass(frozen=True)
class StoreAttrRow:
    store_id: int
    store_code: str
    store_name: str
    report_channel_type: str
    report_channel_type_group: str
    store_grade: str | None
    is_duty_free: str
    is_include_in_daily_report: str
    remark: str | None
    effective_start_date: date
    effective_end_date: date


@dataclass(frozen=True)
class SubjectSourceRow:
    row_number: int
    target_month: str
    target_month_start: date
    target_version: str
    subject_code: str
    subject_name: str
    assessment_mode: str
    month_target: Decimal
    day_targets: dict[int, Decimal]
    blank_day_cell_count: int
    remark: str | None


@dataclass(frozen=True)
class ExpandedSubjectTargetRow:
    target_date: date
    target_month_start: date
    subject_code: str
    subject_name: str
    assessment_mode: str
    month_target: Decimal
    day_target: Decimal
    target_version: str
    remark: str | None


@dataclass(frozen=True)
class ParsedAssessmentAssignmentRow:
    row_number: int
    target_month: str
    target_month_start: date
    target_version: str
    store_key: str
    store_name: str
    subject_code: str
    assignment_role: str
    is_joint_assessment: str
    anchor_store_name: str | None
    effective_start_date: date
    effective_end_date: date
    remark: str | None


@dataclass(frozen=True)
class AssessmentAssignmentRow:
    row_number: int
    target_month: str
    target_month_start: date
    target_version: str
    source_store_key: str
    source_store_name: str
    store_name: str
    store_id: int
    store_code: str | None
    subject_code: str
    assignment_role: str
    is_joint_assessment: str
    anchor_store_name: str | None
    anchor_store_id: int | None
    effective_start_date: date
    effective_end_date: date
    remark: str | None


def _connect():
    return connect_mysql(
        cursorclass=DictCursor,
        autocommit=False,
    )


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _normalize_store_key(value: object) -> str:
    return _normalize_text(value).lower()


def _parse_store_key_value(value: object, row_number: int, field_name: str) -> str:
    if value is None:
        raise ValueError(f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 第 {row_number} 行缺少 {field_name}")

    text_value = _normalize_text(value)
    if not text_value:
        raise ValueError(f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 第 {row_number} 行缺少 {field_name}")

    normalized_value = text_value.replace(',', '')
    if re.fullmatch(r"\d+\.0+", normalized_value):
        return str(int(normalized_value.split('.', 1)[0]))
    if re.fullmatch(r"\d+", normalized_value):
        return str(int(normalized_value))
    return text_value


def _derive_report_channel_type_group(report_channel_type: str | None, allow_unknown: bool = False) -> str:
    normalized_type = _normalize_text(report_channel_type)
    if normalized_type in REPORT_CHANNEL_TYPE_GROUP_MAPPING:
        return REPORT_CHANNEL_TYPE_GROUP_MAPPING[normalized_type]
    if allow_unknown:
        return normalized_type
    raise ValueError(
        f"未识别的日报渠道类型: {normalized_type or '<空>'}；"
        f"当前仅支持 {', '.join(ALLOWED_REPORT_CHANNEL_TYPES)}"
    )


def _parse_month_value(value: object, row_number: int) -> tuple[str, date, int]:
    if isinstance(value, datetime):
        parsed_date = value.date()
    elif isinstance(value, date):
        parsed_date = value
    else:
        text_value = _normalize_text(value)
        if not text_value:
            raise ValueError(f"第 {row_number} 行缺少目标月份")
        matched = MONTH_RE.match(text_value)
        if not matched:
            raise ValueError(f"第 {row_number} 行目标月份格式不支持: {text_value}")
        parsed_date = date(int(matched.group("year")), int(matched.group("month")), 1)

    month_start = date(parsed_date.year, parsed_date.month, 1)
    month_text = month_start.strftime("%Y-%m")
    month_days = calendar.monthrange(month_start.year, month_start.month)[1]
    return month_text, month_start, month_days


def _parse_target_month_arg(value: str) -> str:
    text_value = _normalize_text(value)
    matched = MONTH_RE.match(text_value)
    if not matched:
        raise argparse.ArgumentTypeError(f"--target-month 格式不支持: {value}；请使用 YYYY-MM")

    return f"{int(matched.group('year')):04d}-{int(matched.group('month')):02d}"


def _parse_decimal_value(value: object, row_number: int, field_name: str, allow_blank: bool) -> Decimal:
    if value is None:
        if allow_blank:
            return DECIMAL_ZERO
        raise ValueError(f"第 {row_number} 行 {field_name} 不能为空")

    if isinstance(value, str):
        text_value = value.strip().replace(",", "")
        if not text_value:
            if allow_blank:
                return DECIMAL_ZERO
            raise ValueError(f"第 {row_number} 行 {field_name} 不能为空")
        raw_value = text_value
    else:
        raw_value = str(value)

    try:
        decimal_value = Decimal(raw_value)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"第 {row_number} 行 {field_name} 不是合法数字: {value}") from exc

    return decimal_value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def _parse_date_value(value: object, row_number: int, field_name: str, allow_blank: bool) -> date | None:
    if value is None:
        if allow_blank:
            return None
        raise ValueError(f"第 {row_number} 行 {field_name} 不能为空")

    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text_value = _normalize_text(value)
    if not text_value:
        if allow_blank:
            return None
        raise ValueError(f"第 {row_number} 行 {field_name} 不能为空")

    normalized = (
        text_value.replace("年", "-")
        .replace("月", "-")
        .replace("日", "")
        .replace("/", "-")
        .replace(".", "-")
    )
    try:
        return date.fromisoformat(normalized)
    except ValueError as exc:
        raise ValueError(f"第 {row_number} 行 {field_name} 不是合法日期: {value}") from exc


def _parse_yes_no_flag(value: object, row_number: int, field_name: str) -> str:
    normalized = _normalize_text(value).upper()
    mapping = {
        "Y": "Y",
        "N": "N",
        "是": "Y",
        "否": "N",
        "TRUE": "Y",
        "FALSE": "N",
        "1": "Y",
        "0": "N",
    }
    if normalized not in mapping:
        raise ValueError(f"第 {row_number} 行 {field_name} 仅支持 Y/N/是/否/1/0，当前值: {value}")
    return mapping[normalized]


def _build_named_header_map(
    header_row: tuple[object, ...],
    required_headers: tuple[str, ...],
    optional_aliases: dict[str, tuple[str, ...]] | None = None,
) -> tuple[dict[str, int], dict[int, int]]:
    normalized_map: dict[str, int] = {}
    day_column_map: dict[int, int] = {}
    for index, header in enumerate(header_row):
        normalized = _normalize_text(header)
        if not normalized:
            continue
        normalized_map[normalized] = index
        matched = DAY_HEADER_RE.match(normalized)
        if matched:
            day_column_map[int(matched.group("day"))] = index

    header_index_map: dict[str, int] = {}
    for header in required_headers:
        if header not in normalized_map:
            raise ValueError(f"工作表首行缺少必填表头: {header}")
        header_index_map[header] = normalized_map[header]

    for canonical_name, aliases in (optional_aliases or {}).items():
        for alias in aliases:
            if alias in normalized_map:
                header_index_map[canonical_name] = normalized_map[alias]
                break

    return header_index_map, day_column_map


def _load_workbook_sheetnames(file_path: Path) -> list[str]:
    ensure_nas_path_access(file_path)
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _compute_file_md5(file_path: Path) -> str:
    ensure_nas_path_access(file_path)
    md5_hash = hashlib.md5()
    with file_path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            md5_hash.update(chunk)
    return md5_hash.hexdigest()


def _extract_nas_target_file(file_path: Path) -> NasTargetFile | None:
    matched = None
    for pattern in NAS_FILE_PATTERNS:
        matched = pattern.match(file_path.name)
        if matched:
            break
    if matched is None:
        return None

    target_month = f"{int(matched.group('year')):04d}-{int(matched.group('month')):02d}"
    return NasTargetFile(
        file_path=file_path,
        target_month=target_month,
        file_version=matched.group("version"),
        last_modified_at=datetime.fromtimestamp(file_path.stat().st_mtime),
    )


def _discover_nas_target_files(nas_dir: Path) -> list[NasTargetFile]:
    ensure_nas_path_access(nas_dir)
    if not nas_dir.exists():
        raise FileNotFoundError(f"未找到 NAS 目录: {nas_dir}")

    candidates: list[NasTargetFile] = []
    for file_path in nas_dir.iterdir():
        if not file_path.is_file():
            continue
        parsed = _extract_nas_target_file(file_path)
        if parsed is None:
            continue
        candidates.append(parsed)

    return sorted(candidates, key=lambda item: (item.target_month, item.file_version or "", item.file_path.name))


def _resolve_input_file(file_path_arg: str | None, target_month_filter: str | None) -> Path:
    if file_path_arg:
        file_path = Path(file_path_arg)
        ensure_nas_path_access(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"未找到目标文件: {file_path}")
        return file_path

    candidates = _discover_nas_target_files(DEFAULT_NAS_DIR)
    if not candidates:
        raise FileNotFoundError(
            f"NAS 目录 {DEFAULT_NAS_DIR} 下未找到符合命名规则的目标文件；"
            "当前支持命名形如 2026年04月日目标配置表_v1.xlsx 或 202604考核数据配置表.xlsx"
        )

    if target_month_filter is None:
        available = [f"{item.target_month}:{item.file_path.name}" for item in candidates]
        if len(candidates) == 1:
            return candidates[0].file_path
        raise FileNotFoundError(
            "NAS 目录下存在多个目标月份文件，请显式传入 --target-month YYYY-MM 选择本次导入月份；"
            f"当前可选文件: {available}"
        )

    matched_files = [item for item in candidates if item.target_month == target_month_filter]
    if not matched_files:
        available = [f"{item.target_month}:{item.file_path.name}" for item in candidates]
        raise FileNotFoundError(
            f"NAS 目录下未找到目标月份 {target_month_filter} 的目标文件；"
            f"当前可选文件: {available}"
        )

    if len(matched_files) > 1:
        duplicated = [item.file_path.name for item in matched_files]
        raise FileNotFoundError(
            f"目标月份 {target_month_filter} 在 NAS 目录下匹配到多个文件: {duplicated}；"
            "请使用 --file-path 显式指定本次导入文件"
        )

    return matched_files[0].file_path


def _build_header_map(header_row: tuple[object, ...]) -> tuple[dict[str, int], dict[int, int]]:
    header_index_map: dict[str, int] = {}
    day_column_map: dict[int, int] = {}
    for index, header in enumerate(header_row):
        normalized = _normalize_text(header)
        if not normalized:
            continue
        header_index_map[normalized] = index
        matched = DAY_HEADER_RE.match(normalized)
        if matched:
            day_column_map[int(matched.group("day"))] = index

    missing_headers = [header for header in REQUIRED_HEADERS if header not in header_index_map]
    if missing_headers:
        raise ValueError(f"导入模板首行缺少必填表头: {', '.join(missing_headers)}")

    return header_index_map, day_column_map


def _resolve_optional_header_index(
    header_index_map: dict[str, int],
    header_aliases: tuple[str, ...],
) -> int | None:
    for header_name in header_aliases:
        if header_name in header_index_map:
            return header_index_map[header_name]
    return None


def _parse_workbook(
    file_path: Path,
    sheet_name: str,
    require_store_type: bool = False,
    target_month_filter: str | None = None,
) -> tuple[list[SourceRow], dict]:
    ensure_nas_path_access(file_path)
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"未找到工作表 {sheet_name}，可选 sheet: {', '.join(workbook.sheetnames)}")

        worksheet = workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            raise ValueError("导入模板为空，缺少表头")

        header_index_map, day_column_map = _build_header_map(header_row)
        store_grade_column_index = _resolve_optional_header_index(header_index_map, STORE_GRADE_HEADER_ALIASES)
        effective_start_column_index = _resolve_optional_header_index(
            header_index_map,
            OPTIONAL_SHEET_NAME_ALIASES[EFFECTIVE_START_DATE_HEADER],
        )
        effective_end_column_index = _resolve_optional_header_index(
            header_index_map,
            OPTIONAL_SHEET_NAME_ALIASES[EFFECTIVE_END_DATE_HEADER],
        )
        if require_store_type and STORE_TYPE_HEADER not in header_index_map:
            raise ValueError(f"导入模板缺少 {STORE_TYPE_HEADER} 列，无法同步 {STORE_ATTR_TABLE_NAME}")

        parsed_rows: list[SourceRow] = []
        duplicate_keys: set[tuple[str, str, str]] = set()
        seen_keys: set[tuple[str, str, str]] = set()
        blank_day_cell_count = 0
        last_month_value: object | None = None
        last_version_value: object | None = None
        available_target_months: set[str] = set()

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if row is None:
                continue
            if all(cell is None or _normalize_text(cell) == "" for cell in row):
                continue

            month_cell = row[header_index_map["目标月份"]]
            version_cell = row[header_index_map["目标版本"]]

            if month_cell is not None and _normalize_text(month_cell) != "":
                last_month_value = month_cell
            if version_cell is not None and _normalize_text(version_cell) != "":
                last_version_value = version_cell

            month_text, month_start, month_days = _parse_month_value(last_month_value, row_number)
            month_end = date(month_start.year, month_start.month, month_days)
            available_target_months.add(month_text)
            if target_month_filter is not None and month_text != target_month_filter:
                continue

            store_name = _normalize_text(row[header_index_map["门店名称"]])
            month_target = _parse_decimal_value(row[header_index_map["月目标"]], row_number, "月目标", allow_blank=False)
            target_version = _normalize_text(last_version_value)
            store_type = _normalize_text(row[header_index_map[STORE_TYPE_HEADER]]) if STORE_TYPE_HEADER in header_index_map else ""
            store_grade = _normalize_text(row[store_grade_column_index]) if store_grade_column_index is not None else ""
            if not target_version:
                raise ValueError(f"第 {row_number} 行缺少目标版本")
            if not store_name:
                raise ValueError(f"第 {row_number} 行缺少门店名称")
            if require_store_type and not store_type:
                raise ValueError(f"第 {row_number} 行缺少 {STORE_TYPE_HEADER}，无法同步 {STORE_ATTR_TABLE_NAME}")
            if store_type and store_type not in ALLOWED_REPORT_CHANNEL_TYPES:
                raise ValueError(
                    f"第 {row_number} 行 {STORE_TYPE_HEADER} 不在允许范围内: {store_type}；"
                    f"当前仅支持 {', '.join(ALLOWED_REPORT_CHANNEL_TYPES)}"
                )

            effective_start_date = _parse_date_value(
                row[effective_start_column_index] if effective_start_column_index is not None else None,
                row_number,
                EFFECTIVE_START_DATE_HEADER,
                allow_blank=True,
            ) or month_start
            effective_end_date = _parse_date_value(
                row[effective_end_column_index] if effective_end_column_index is not None else None,
                row_number,
                EFFECTIVE_END_DATE_HEADER,
                allow_blank=True,
            ) or month_end
            if effective_start_date > effective_end_date:
                raise ValueError(f"第 {row_number} 行生效开始日不能晚于生效结束日")
            if effective_start_date < month_start or effective_end_date > month_end:
                raise ValueError(
                    f"第 {row_number} 行生效区间必须落在 {month_text} 自然月内"
                )

            day_targets: dict[int, Decimal] = {}
            row_blank_day_count = 0
            for day in range(1, month_days + 1):
                if day not in day_column_map:
                    raise ValueError(f"导入模板缺少 {day}日目标 列，无法覆盖 {month_text} 全月数据")
                cell_value = row[day_column_map[day]]
                if cell_value is None or _normalize_text(cell_value) == "":
                    row_blank_day_count += 1
                day_targets[day] = _parse_decimal_value(cell_value, row_number, f"{day}日目标", allow_blank=True)

            for day, column_index in day_column_map.items():
                if day <= month_days:
                    continue
                extra_value = row[column_index]
                if extra_value is not None and _normalize_text(extra_value) != "":
                    raise ValueError(f"第 {row_number} 行 {day}日目标 超出 {month_text} 自然月天数，需留空")

            row_key = (month_text, target_version, store_name)
            if row_key in seen_keys:
                duplicate_keys.add(row_key)
            seen_keys.add(row_key)

            blank_day_cell_count += row_blank_day_count
            parsed_rows.append(
                SourceRow(
                    row_number=row_number,
                    target_month=month_text,
                    target_month_start=month_start,
                    target_version=target_version,
                    store_name=store_name,
                    store_type=store_type or None,
                    store_grade=store_grade or None,
                    month_target=month_target,
                    day_targets=day_targets,
                    blank_day_cell_count=row_blank_day_count,
                    effective_start_date=effective_start_date,
                    effective_end_date=effective_end_date,
                )
            )

        if not parsed_rows:
            if target_month_filter is not None and available_target_months:
                raise ValueError(
                    f"模板中未找到目标月份 {target_month_filter} 的有效门店数据；"
                    f"当前可选月份: {sorted(available_target_months)}"
                )
            raise ValueError("导入模板没有有效的门店数据行")

        if duplicate_keys:
            duplicates = [" | ".join(item) for item in sorted(duplicate_keys)]
            raise ValueError(f"模板中存在重复门店行: {duplicates}")

        if len(available_target_months) != 1 and target_month_filter is None:
            suggested_month = sorted(available_target_months)[-1]
            raise ValueError(
                f"同一文件检测到多个目标月份: {sorted(available_target_months)}；"
                f"请使用 --target-month YYYY-MM 显式选择需导入的月份，例如 --target-month {suggested_month}"
            )

        target_months = sorted({row.target_month for row in parsed_rows})
        target_versions = sorted({row.target_version for row in parsed_rows})
        if len(target_months) != 1:
            raise ValueError(f"筛选后的数据仍包含多个目标月份，当前检测到: {target_months}")
        if len(target_versions) != 1:
            raise ValueError(f"同一文件只允许一个目标版本，当前检测到: {target_versions}")

        workbook_summary = {
            "target_month": target_months[0],
            "target_version": target_versions[0],
            "blank_day_cell_count": blank_day_cell_count,
            "source_row_count": len(parsed_rows),
            "available_target_months": sorted(available_target_months),
            "target_month_filter": target_month_filter,
        }
        return parsed_rows, workbook_summary
    finally:
        workbook.close()


def _parse_optional_subject_target_sheet(
    file_path: Path,
    target_month_filter: str | None,
) -> tuple[list[SubjectSourceRow], dict[str, Any]]:
    ensure_nas_path_access(file_path)
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    try:
        if SUBJECT_TARGET_SHEET_NAME not in workbook.sheetnames:
            return [], {
                "sheet_name": SUBJECT_TARGET_SHEET_NAME,
                "sheet_present": False,
                "source_row_count": 0,
                "blank_day_cell_count": 0,
                "available_target_months": [],
            }

        worksheet = workbook[SUBJECT_TARGET_SHEET_NAME]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None or all(_normalize_text(cell) == "" for cell in header_row):
            return [], {
                "sheet_name": SUBJECT_TARGET_SHEET_NAME,
                "sheet_present": True,
                "source_row_count": 0,
                "blank_day_cell_count": 0,
                "available_target_months": [],
            }

        header_index_map, day_column_map = _build_named_header_map(
            header_row,
            SUBJECT_TARGET_REQUIRED_HEADERS,
            optional_aliases={REMARK_HEADER: OPTIONAL_SHEET_NAME_ALIASES[REMARK_HEADER]},
        )

        parsed_rows: list[SubjectSourceRow] = []
        seen_keys: set[tuple[str, str, str]] = set()
        duplicate_keys: set[tuple[str, str, str]] = set()
        blank_day_cell_count = 0
        available_target_months: set[str] = set()
        last_month_value: object | None = None
        last_version_value: object | None = None

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if row is None:
                continue
            if all(cell is None or _normalize_text(cell) == "" for cell in row):
                continue

            month_cell = row[header_index_map["目标月份"]]
            version_cell = row[header_index_map["目标版本"]]
            if month_cell is not None and _normalize_text(month_cell) != "":
                last_month_value = month_cell
            if version_cell is not None and _normalize_text(version_cell) != "":
                last_version_value = version_cell

            month_text, month_start, month_days = _parse_month_value(last_month_value, row_number)
            available_target_months.add(month_text)
            if target_month_filter is not None and month_text != target_month_filter:
                continue

            target_version = _normalize_text(last_version_value)
            subject_code = _normalize_text(row[header_index_map[SUBJECT_CODE_HEADER]])
            subject_name = _normalize_text(row[header_index_map[SUBJECT_NAME_HEADER]])
            assessment_mode = _normalize_text(row[header_index_map[ASSESSMENT_MODE_HEADER]])
            remark = _normalize_text(row[header_index_map[REMARK_HEADER]]) if REMARK_HEADER in header_index_map else ""
            month_target = _parse_decimal_value(
                row[header_index_map["月目标"]],
                row_number,
                "月目标",
                allow_blank=False,
            )

            if not target_version:
                raise ValueError(f"工作表 {SUBJECT_TARGET_SHEET_NAME} 第 {row_number} 行缺少目标版本")
            if not subject_code:
                raise ValueError(f"工作表 {SUBJECT_TARGET_SHEET_NAME} 第 {row_number} 行缺少主体编码")
            if not subject_name:
                raise ValueError(f"工作表 {SUBJECT_TARGET_SHEET_NAME} 第 {row_number} 行缺少主体名称")
            if assessment_mode not in ALLOWED_ASSESSMENT_MODES:
                raise ValueError(
                    f"工作表 {SUBJECT_TARGET_SHEET_NAME} 第 {row_number} 行考核模式仅支持 "
                    f"{', '.join(ALLOWED_ASSESSMENT_MODES)}，当前值: {assessment_mode or '<空>'}"
                )

            day_targets: dict[int, Decimal] = {}
            row_blank_day_count = 0
            for day in range(1, month_days + 1):
                if day not in day_column_map:
                    raise ValueError(
                        f"工作表 {SUBJECT_TARGET_SHEET_NAME} 缺少 {day}日目标 列，无法覆盖 {month_text} 全月数据"
                    )
                cell_value = row[day_column_map[day]]
                if cell_value is None or _normalize_text(cell_value) == "":
                    row_blank_day_count += 1
                day_targets[day] = _parse_decimal_value(cell_value, row_number, f"{day}日目标", allow_blank=True)

            for day, column_index in day_column_map.items():
                if day <= month_days:
                    continue
                extra_value = row[column_index]
                if extra_value is not None and _normalize_text(extra_value) != "":
                    raise ValueError(
                        f"工作表 {SUBJECT_TARGET_SHEET_NAME} 第 {row_number} 行 {day}日目标 超出 {month_text} 自然月天数，需留空"
                    )

            row_key = (month_text, target_version, subject_code)
            if row_key in seen_keys:
                duplicate_keys.add(row_key)
            seen_keys.add(row_key)

            blank_day_cell_count += row_blank_day_count
            parsed_rows.append(
                SubjectSourceRow(
                    row_number=row_number,
                    target_month=month_text,
                    target_month_start=month_start,
                    target_version=target_version,
                    subject_code=subject_code,
                    subject_name=subject_name,
                    assessment_mode=assessment_mode,
                    month_target=month_target,
                    day_targets=day_targets,
                    blank_day_cell_count=row_blank_day_count,
                    remark=remark or None,
                )
            )

        if not parsed_rows:
            if target_month_filter is not None and available_target_months:
                raise ValueError(
                    f"工作表 {SUBJECT_TARGET_SHEET_NAME} 未找到目标月份 {target_month_filter} 的有效主体数据；"
                    f"当前可选月份: {sorted(available_target_months)}"
                )
            return [], {
                "sheet_name": SUBJECT_TARGET_SHEET_NAME,
                "sheet_present": True,
                "source_row_count": 0,
                "blank_day_cell_count": 0,
                "available_target_months": sorted(available_target_months),
            }

        if duplicate_keys:
            duplicates = [" | ".join(item) for item in sorted(duplicate_keys)]
            raise ValueError(f"工作表 {SUBJECT_TARGET_SHEET_NAME} 存在重复主体行: {duplicates}")

        if len(available_target_months) != 1 and target_month_filter is None:
            suggested_month = sorted(available_target_months)[-1]
            raise ValueError(
                f"工作表 {SUBJECT_TARGET_SHEET_NAME} 检测到多个目标月份: {sorted(available_target_months)}；"
                f"请使用 --target-month YYYY-MM 显式选择月份，例如 --target-month {suggested_month}"
            )

        target_months = sorted({row.target_month for row in parsed_rows})
        target_versions = sorted({row.target_version for row in parsed_rows})
        if len(target_months) != 1:
            raise ValueError(f"工作表 {SUBJECT_TARGET_SHEET_NAME} 筛选后仍包含多个目标月份: {target_months}")
        if len(target_versions) != 1:
            raise ValueError(f"工作表 {SUBJECT_TARGET_SHEET_NAME} 只允许一个目标版本，当前检测到: {target_versions}")

        return parsed_rows, {
            "sheet_name": SUBJECT_TARGET_SHEET_NAME,
            "sheet_present": True,
            "target_month": target_months[0],
            "target_version": target_versions[0],
            "source_row_count": len(parsed_rows),
            "blank_day_cell_count": blank_day_cell_count,
            "available_target_months": sorted(available_target_months),
        }
    finally:
        workbook.close()


def _expand_subject_target_rows(parsed_rows: list[SubjectSourceRow]) -> list[ExpandedSubjectTargetRow]:
    expanded_rows: list[ExpandedSubjectTargetRow] = []
    for parsed_row in parsed_rows:
        for day, day_target in sorted(parsed_row.day_targets.items()):
            expanded_rows.append(
                ExpandedSubjectTargetRow(
                    target_date=date(
                        parsed_row.target_month_start.year,
                        parsed_row.target_month_start.month,
                        day,
                    ),
                    target_month_start=parsed_row.target_month_start,
                    subject_code=parsed_row.subject_code,
                    subject_name=parsed_row.subject_name,
                    assessment_mode=parsed_row.assessment_mode,
                    month_target=parsed_row.month_target,
                    day_target=day_target,
                    target_version=parsed_row.target_version,
                    remark=parsed_row.remark,
                )
            )
    return expanded_rows


def _parse_optional_assessment_assignment_sheet(
    file_path: Path,
    target_month_filter: str | None,
) -> tuple[list[ParsedAssessmentAssignmentRow], dict[str, Any]]:
    ensure_nas_path_access(file_path)
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    try:
        if ASSESSMENT_ASSIGNMENT_SHEET_NAME not in workbook.sheetnames:
            return [], {
                "sheet_name": ASSESSMENT_ASSIGNMENT_SHEET_NAME,
                "sheet_present": False,
                "source_row_count": 0,
                "available_target_months": [],
            }

        worksheet = workbook[ASSESSMENT_ASSIGNMENT_SHEET_NAME]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None or all(_normalize_text(cell) == "" for cell in header_row):
            return [], {
                "sheet_name": ASSESSMENT_ASSIGNMENT_SHEET_NAME,
                "sheet_present": True,
                "source_row_count": 0,
                "available_target_months": [],
            }

        header_index_map, _ = _build_named_header_map(
            header_row,
            ASSESSMENT_ASSIGNMENT_REQUIRED_HEADERS,
            optional_aliases=OPTIONAL_SHEET_NAME_ALIASES,
        )

        parsed_rows: list[ParsedAssessmentAssignmentRow] = []
        seen_keys: set[tuple[str, str, str, date]] = set()
        duplicate_keys: set[tuple[str, str, str, date]] = set()
        available_target_months: set[str] = set()
        last_month_value: object | None = None
        last_version_value: object | None = None

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if row is None:
                continue
            if all(cell is None or _normalize_text(cell) == "" for cell in row):
                continue

            month_cell = row[header_index_map["目标月份"]]
            version_cell = row[header_index_map["目标版本"]]
            if month_cell is not None and _normalize_text(month_cell) != "":
                last_month_value = month_cell
            if version_cell is not None and _normalize_text(version_cell) != "":
                last_version_value = version_cell

            month_text, month_start, month_days = _parse_month_value(last_month_value, row_number)
            month_end = date(month_start.year, month_start.month, month_days)
            available_target_months.add(month_text)
            if target_month_filter is not None and month_text != target_month_filter:
                continue

            target_version = _normalize_text(last_version_value)
            store_key = _parse_store_key_value(row[header_index_map[STORE_ID_HEADER]], row_number, STORE_ID_HEADER)
            store_name = _normalize_text(row[header_index_map["门店名称"]])
            subject_code = _normalize_text(row[header_index_map[SUBJECT_CODE_HEADER]])
            assignment_role = _normalize_text(row[header_index_map[ASSIGNMENT_ROLE_HEADER]])
            is_joint_assessment = _parse_yes_no_flag(
                row[header_index_map[IS_JOINT_ASSESSMENT_HEADER]],
                row_number,
                IS_JOINT_ASSESSMENT_HEADER,
            )
            anchor_store_name = _normalize_text(row[header_index_map[ANCHOR_STORE_NAME_HEADER]]) if ANCHOR_STORE_NAME_HEADER in header_index_map else ""
            remark = _normalize_text(row[header_index_map[REMARK_HEADER]]) if REMARK_HEADER in header_index_map else ""
            effective_start_date = _parse_date_value(
                row[header_index_map[EFFECTIVE_START_DATE_HEADER]] if EFFECTIVE_START_DATE_HEADER in header_index_map else None,
                row_number,
                EFFECTIVE_START_DATE_HEADER,
                allow_blank=True,
            ) or month_start
            effective_end_date = _parse_date_value(
                row[header_index_map[EFFECTIVE_END_DATE_HEADER]] if EFFECTIVE_END_DATE_HEADER in header_index_map else None,
                row_number,
                EFFECTIVE_END_DATE_HEADER,
                allow_blank=True,
            ) or month_end

            if not target_version:
                raise ValueError(f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 第 {row_number} 行缺少目标版本")
            if not store_name:
                raise ValueError(f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 第 {row_number} 行缺少门店名称")
            if not subject_code:
                raise ValueError(f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 第 {row_number} 行缺少主体编码")
            if assignment_role not in ALLOWED_ASSIGNMENT_ROLES:
                raise ValueError(
                    f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 第 {row_number} 行归属角色仅支持 "
                    f"{', '.join(ALLOWED_ASSIGNMENT_ROLES)}，当前值: {assignment_role or '<空>'}"
                )
            if effective_start_date > effective_end_date:
                raise ValueError(
                    f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 第 {row_number} 行生效开始日不能晚于生效结束日"
                )
            if effective_start_date < month_start or effective_end_date > month_end:
                raise ValueError(
                    f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 第 {row_number} 行生效区间必须落在 {month_text} 自然月内"
                )

            row_key = (month_text, target_version, store_key, effective_start_date)
            if row_key in seen_keys:
                duplicate_keys.add(row_key)
            seen_keys.add(row_key)

            parsed_rows.append(
                ParsedAssessmentAssignmentRow(
                    row_number=row_number,
                    target_month=month_text,
                    target_month_start=month_start,
                    target_version=target_version,
                    store_key=store_key,
                    store_name=store_name,
                    subject_code=subject_code,
                    assignment_role=assignment_role,
                    is_joint_assessment=is_joint_assessment,
                    anchor_store_name=anchor_store_name or None,
                    effective_start_date=effective_start_date,
                    effective_end_date=effective_end_date,
                    remark=remark or None,
                )
            )

        if not parsed_rows:
            if target_month_filter is not None and available_target_months:
                raise ValueError(
                    f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 未找到目标月份 {target_month_filter} 的有效归属数据；"
                    f"当前可选月份: {sorted(available_target_months)}"
                )
            return [], {
                "sheet_name": ASSESSMENT_ASSIGNMENT_SHEET_NAME,
                "sheet_present": True,
                "source_row_count": 0,
                "available_target_months": sorted(available_target_months),
            }

        if duplicate_keys:
            duplicates = [" | ".join((item[0], item[1], item[2], item[3].isoformat())) for item in sorted(duplicate_keys)]
            raise ValueError(f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 存在重复门店归属行: {duplicates}")

        if len(available_target_months) != 1 and target_month_filter is None:
            suggested_month = sorted(available_target_months)[-1]
            raise ValueError(
                f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 检测到多个目标月份: {sorted(available_target_months)}；"
                f"请使用 --target-month YYYY-MM 显式选择月份，例如 --target-month {suggested_month}"
            )

        target_months = sorted({row.target_month for row in parsed_rows})
        target_versions = sorted({row.target_version for row in parsed_rows})
        if len(target_months) != 1:
            raise ValueError(f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 筛选后仍包含多个目标月份: {target_months}")
        if len(target_versions) != 1:
            raise ValueError(f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 只允许一个目标版本，当前检测到: {target_versions}")

        return parsed_rows, {
            "sheet_name": ASSESSMENT_ASSIGNMENT_SHEET_NAME,
            "sheet_present": True,
            "target_month": target_months[0],
            "target_version": target_versions[0],
            "source_row_count": len(parsed_rows),
            "available_target_months": sorted(available_target_months),
        }
    finally:
        workbook.close()


def _resolve_assessment_assignments(
    parsed_rows: list[ParsedAssessmentAssignmentRow],
    store_map: dict[str, dict[str, Any]],
    store_map_by_id: dict[int, dict[str, Any]],
    store_map_by_code: dict[str, dict[str, Any]],
) -> tuple[list[AssessmentAssignmentRow], dict[str, Any]]:
    resolved_rows: list[AssessmentAssignmentRow] = []
    missing_store_names: list[str] = []
    missing_store_keys: list[str] = []
    missing_anchor_store_names: list[str] = []
    inconsistent_primary_anchor_rows: list[dict[str, Any]] = []
    store_key_name_mismatch_rows: list[dict[str, Any]] = []
    missing_primary_subject_keys: set[tuple[str, str, str]] = set()

    for row in parsed_rows:
        store_info: dict[str, Any] | None = None
        matched_by = "store_code"
        if re.fullmatch(r"\d+", row.store_key):
            store_info = store_map_by_id.get(int(row.store_key))
            matched_by = "store_id"
        if store_info is None:
            store_info = store_map_by_code.get(_normalize_store_key(row.store_key))
            matched_by = "store_code"
        if store_info is None:
            missing_store_keys.append(row.store_key)
            missing_store_names.append(row.store_name)
            if row.assignment_role == "主店":
                missing_primary_subject_keys.add((row.target_month, row.target_version, row.subject_code))
            continue

        normalized_excel_store_name = _normalize_store_key(row.store_name)
        normalized_dim_store_name = _normalize_store_key(store_info["store_name"])
        if normalized_excel_store_name and normalized_excel_store_name != normalized_dim_store_name:
            store_key_name_mismatch_rows.append(
                {
                    "row_number": row.row_number,
                    "source_store_key": row.store_key,
                    "matched_by": matched_by,
                    "store_id": int(store_info["store_id"]),
                    "store_code": store_info.get("store_code"),
                    "subject_code": row.subject_code,
                    "excel_store_name": row.store_name,
                    "dim_store_name": store_info["store_name"],
                }
            )

        anchor_store_id: int | None = None
        anchor_store_name: str | None = row.anchor_store_name
        if row.anchor_store_name:
            anchor_store_info = store_map.get(_normalize_store_key(row.anchor_store_name))
            if anchor_store_info is None:
                missing_anchor_store_names.append(row.anchor_store_name)
                continue
            anchor_store_id = int(anchor_store_info["store_id"])
            anchor_store_name = anchor_store_info["store_name"]

        resolved_rows.append(
            AssessmentAssignmentRow(
                row_number=row.row_number,
                target_month=row.target_month,
                target_month_start=row.target_month_start,
                target_version=row.target_version,
                source_store_key=row.store_key,
                source_store_name=row.store_name,
                store_name=store_info["store_name"],
                store_id=int(store_info["store_id"]),
                store_code=store_info.get("store_code"),
                subject_code=row.subject_code,
                assignment_role=row.assignment_role,
                is_joint_assessment=row.is_joint_assessment,
                anchor_store_name=anchor_store_name,
                anchor_store_id=anchor_store_id,
                effective_start_date=row.effective_start_date,
                effective_end_date=row.effective_end_date,
                remark=row.remark,
            )
        )

    rows_by_subject: dict[tuple[str, str, str], list[AssessmentAssignmentRow]] = {}
    for row in resolved_rows:
        subject_key = (row.target_month, row.target_version, row.subject_code)
        rows_by_subject.setdefault(subject_key, []).append(row)

    finalized_rows: list[AssessmentAssignmentRow] = []
    unanchored_rows: list[dict[str, Any]] = []
    skipped_missing_primary_rows: list[dict[str, Any]] = []
    for subject_key, subject_rows in rows_by_subject.items():
        primary_rows = [row for row in subject_rows if row.assignment_role == "主店"]
        primary_row = primary_rows[0] if len(primary_rows) == 1 else None
        for row in subject_rows:
            updated_row = row
            if row.assignment_role == "主店":
                if row.anchor_store_id is None:
                    updated_row = replace(
                        row,
                        anchor_store_id=row.store_id,
                        anchor_store_name=row.store_name,
                    )
                elif row.anchor_store_id != row.store_id:
                    inconsistent_primary_anchor_rows.append(
                        {
                            "row_number": row.row_number,
                            "subject_code": row.subject_code,
                            "store_name": row.store_name,
                            "anchor_store_name": row.anchor_store_name,
                        }
                    )

            if (
                updated_row.is_joint_assessment == "Y"
                and updated_row.assignment_role != "主店"
                and updated_row.anchor_store_id is None
            ):
                if primary_row is not None:
                    updated_row = replace(
                        updated_row,
                        anchor_store_id=primary_row.store_id,
                        anchor_store_name=primary_row.store_name,
                    )
                elif subject_key in missing_primary_subject_keys:
                    skipped_missing_primary_rows.append(
                        {
                            "row_number": updated_row.row_number,
                            "subject_code": updated_row.subject_code,
                            "store_name": updated_row.store_name,
                        }
                    )
                    continue
                else:
                    unanchored_rows.append(
                        {
                            "row_number": updated_row.row_number,
                            "subject_code": updated_row.subject_code,
                            "store_name": updated_row.store_name,
                        }
                    )

            finalized_rows.append(updated_row)

    return sorted(finalized_rows, key=lambda row: (row.target_month, row.subject_code, row.store_id, row.effective_start_date)), {
        "missing_store_keys": sorted(set(missing_store_keys)),
        "missing_store_names": sorted(set(missing_store_names)),
        "store_key_name_mismatch_rows": store_key_name_mismatch_rows,
        "missing_anchor_store_names": sorted(set(missing_anchor_store_names)),
        "inconsistent_primary_anchor_rows": inconsistent_primary_anchor_rows,
        "unanchored_rows": unanchored_rows,
        "skipped_missing_primary_rows": skipped_missing_primary_rows,
    }


def _find_assignment_overlap_rows(rows: list[AssessmentAssignmentRow], limit: int = 20) -> list[dict[str, Any]]:
    overlap_rows: list[dict[str, Any]] = []
    rows_by_store: dict[tuple[str, str, int], list[AssessmentAssignmentRow]] = {}
    for row in rows:
        rows_by_store.setdefault((row.target_month, row.target_version, row.store_id), []).append(row)

    for store_key, store_rows in rows_by_store.items():
        sorted_rows = sorted(store_rows, key=lambda item: (item.effective_start_date, item.effective_end_date))
        for previous_row, current_row in zip(sorted_rows, sorted_rows[1:]):
            if current_row.effective_start_date <= previous_row.effective_end_date:
                overlap_rows.append(
                    {
                        "target_month": store_key[0],
                        "target_version": store_key[1],
                        "store_id": store_key[2],
                        "store_name": current_row.store_name,
                        "previous_range": [
                            previous_row.effective_start_date.isoformat(),
                            previous_row.effective_end_date.isoformat(),
                        ],
                        "current_range": [
                            current_row.effective_start_date.isoformat(),
                            current_row.effective_end_date.isoformat(),
                        ],
                    }
                )
                if len(overlap_rows) >= limit:
                    return overlap_rows
    return overlap_rows


def _fetch_existing_subject_target_count(conn: Any, month_start: date, month_end: date, target_version: str) -> int:
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT COUNT(*) AS row_count_alias
            FROM {SUBJECT_TARGET_TABLE_NAME}
            WHERE target_date BETWEEN %s AND %s
              AND target_version = %s
            """,
            (month_start, month_end, target_version),
        )
        row = cursor.fetchone()
    return int(dict(row)["row_count_alias"] or 0) if row else 0


def _fetch_existing_assessment_assignment_count(conn: Any, month_start: date, target_version: str) -> int:
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT COUNT(*) AS row_count_alias
            FROM {ASSESSMENT_ASSIGNMENT_TABLE_NAME}
            WHERE target_month = %s
              AND target_version = %s
            """,
            (month_start, target_version),
        )
        row = cursor.fetchone()
    return int(dict(row)["row_count_alias"] or 0) if row else 0


def _ensure_required_tables(
    conn: Any,
    sync_store_report_attr: bool = False,
    sync_assessment: bool = False,
) -> None:
    required_tables = [STORE_TABLE_NAME, TARGET_TABLE_NAME]
    if sync_store_report_attr:
        required_tables.extend([STORE_ATTR_TABLE_NAME, STORE_ATTR_SNAPSHOT_TABLE_NAME])
    if sync_assessment:
        required_tables.extend([SUBJECT_TARGET_TABLE_NAME, ASSESSMENT_ASSIGNMENT_TABLE_NAME])

    placeholders = ", ".join(["%s"] * len(required_tables))
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT table_name AS table_name_alias
            FROM information_schema.tables
            WHERE table_schema = DATABASE()
              AND table_name IN ({placeholders})
            """,
            required_tables,
        )
        existing = {dict(row)["table_name_alias"] for row in cursor.fetchall()}

    missing_tables = [table_name for table_name in required_tables if table_name not in existing]
    if missing_tables:
        missing_table_messages: list[str] = []
        for table_name in missing_tables:
            if table_name == STORE_ATTR_SNAPSHOT_TABLE_NAME:
                ddl_hint = STORE_ATTR_SNAPSHOT_SQL_PATH.relative_to(REPO_ROOT).as_posix()
                missing_table_messages.append(f"{table_name}(请先执行 {ddl_hint})")
            else:
                missing_table_messages.append(table_name)
        raise RuntimeError(f"缺少依赖表: {', '.join(missing_table_messages)}")


def _load_store_mapping(
    conn: Any,
) -> tuple[dict[str, dict[str, Any]], dict[int, dict[str, Any]], dict[str, dict[str, Any]], list[dict[str, Any]]]:
    with conn.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                store_id AS store_id_alias,
                store_code AS store_code_alias,
                store_name AS store_name_alias
            FROM dim_store
            WHERE store_name IS NOT NULL
              AND TRIM(store_name) <> ''
            ORDER BY store_id
            """
        )
        rows = [dict(row) for row in cursor.fetchall()]

    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        normalized_name = _normalize_store_key(row["store_name_alias"])
        grouped.setdefault(normalized_name, []).append(
            {
                "store_id": row["store_id_alias"],
                "store_code": row["store_code_alias"],
                "store_name": row["store_name_alias"],
            }
        )

    store_map: dict[str, dict[str, Any]] = {}
    ambiguous: list[dict[str, Any]] = []
    for normalized_name, matched_rows in grouped.items():
        unique_store_ids = {row["store_id"] for row in matched_rows}
        if len(unique_store_ids) > 1:
            ambiguous.append({
                "store_name": normalized_name,
                "matched_store_ids": sorted(unique_store_ids),
            })
            continue
        store_map[normalized_name] = matched_rows[0]

    store_map_by_id = {
        int(row["store_id_alias"]): {
            "store_id": row["store_id_alias"],
            "store_code": row["store_code_alias"],
            "store_name": row["store_name_alias"],
        }
        for row in rows
    }
    store_map_by_code = {
        _normalize_store_key(row["store_code_alias"]): {
            "store_id": row["store_id_alias"],
            "store_code": row["store_code_alias"],
            "store_name": row["store_name_alias"],
        }
        for row in rows
        if _normalize_text(row["store_code_alias"])
    }
    return store_map, store_map_by_id, store_map_by_code, ambiguous


def _load_latest_store_attr(conn: Any) -> dict[int, dict[str, Any]]:
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                t.store_id AS store_id_alias,
                t.report_channel_type AS report_channel_type_alias,
                t.store_grade AS store_grade_alias,
                t.is_duty_free AS is_duty_free_alias,
                t.remark AS remark_alias,
                t.effective_start_date AS effective_start_date_alias
            FROM {STORE_ATTR_TABLE_NAME} t
            INNER JOIN (
                SELECT store_id, MAX(effective_start_date) AS max_effective_start_date
                FROM {STORE_ATTR_TABLE_NAME}
                GROUP BY store_id
            ) latest
                ON latest.store_id = t.store_id
               AND latest.max_effective_start_date = t.effective_start_date
            """
        )
        rows = [dict(row) for row in cursor.fetchall()]

    return {
        int(row["store_id_alias"]): {
            "report_channel_type": row["report_channel_type_alias"],
            "store_grade": row["store_grade_alias"],
            "is_duty_free": row["is_duty_free_alias"],
            "remark": row["remark_alias"],
            "effective_start_date": row["effective_start_date_alias"],
        }
        for row in rows
    }


def _fetch_store_attr_row_count(conn: Any, effective_start_date: date) -> int:
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT COUNT(*) AS row_count_alias
            FROM {STORE_ATTR_TABLE_NAME}
            WHERE effective_start_date = %s
            """,
            (effective_start_date,),
        )
        row = cursor.fetchone()
    return int(dict(row)["row_count_alias"] or 0) if row else 0


def _resolve_store_attr_effective_start_date(
    conn: Any,
    month_start: date,
    month_end: date,
    override_effective_start_date: date | None,
) -> tuple[date, str, int, int]:
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                MAX(effective_start_date) AS latest_effective_start_date_alias,
                COUNT(*) AS row_count_alias
            FROM {STORE_ATTR_TABLE_NAME}
            WHERE effective_start_date BETWEEN %s AND %s
            """,
            (month_start, month_end),
        )
        latest_row = cursor.fetchone()

    month_row_count = int(dict(latest_row)["row_count_alias"] or 0) if latest_row else 0
    if override_effective_start_date is not None:
        return (
            override_effective_start_date,
            "cli_override",
            _fetch_store_attr_row_count(conn, override_effective_start_date),
            month_row_count,
        )

    latest_effective_start_date = dict(latest_row)["latest_effective_start_date_alias"] if latest_row else None
    if latest_effective_start_date is not None:
        return (
            latest_effective_start_date,
            "existing_latest_in_target_month",
            _fetch_store_attr_row_count(conn, latest_effective_start_date),
            month_row_count,
        )

    return month_start, "target_month_start_fallback", 0, 0


def _load_store_attr_snapshot_rows(
    conn: Any,
    target_month: date,
    target_version: str,
) -> list[StoreAttrRow]:
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                store_id AS store_id_alias,
                store_code AS store_code_alias,
                store_name AS store_name_alias,
                report_channel_type AS report_channel_type_alias,
                report_channel_type_group AS report_channel_type_group_alias,
                store_grade AS store_grade_alias,
                is_duty_free AS is_duty_free_alias,
                is_include_in_daily_report AS is_include_in_daily_report_alias,
                remark AS remark_alias,
                effective_start_date AS effective_start_date_alias,
                effective_end_date AS effective_end_date_alias
            FROM {STORE_ATTR_SNAPSHOT_TABLE_NAME}
            WHERE target_month = %s
              AND target_version = %s
            ORDER BY store_id
            """,
            (target_month, target_version),
        )
        rows = [dict(row) for row in cursor.fetchall()]

    return [
        StoreAttrRow(
            store_id=int(row["store_id_alias"]),
            store_code=row["store_code_alias"],
            store_name=row["store_name_alias"],
            report_channel_type=row["report_channel_type_alias"],
            report_channel_type_group=row["report_channel_type_group_alias"]
            or _derive_report_channel_type_group(row["report_channel_type_alias"]),
            store_grade=row["store_grade_alias"],
            is_duty_free=row["is_duty_free_alias"],
            is_include_in_daily_report=row["is_include_in_daily_report_alias"],
            remark=row["remark_alias"],
            effective_start_date=row["effective_start_date_alias"],
            effective_end_date=row["effective_end_date_alias"],
        )
        for row in rows
    ]


def _replace_store_attr_snapshot_rows(
    cursor: Any,
    target_month: date,
    target_version: str,
    store_attr_rows: list[StoreAttrRow],
    file_name: str,
    file_md5: str,
    updated_by: str,
) -> int:
    cursor.execute(
        f"""
        DELETE FROM {STORE_ATTR_SNAPSHOT_TABLE_NAME}
        WHERE target_month = %s
          AND target_version = %s
        """,
        (target_month, target_version),
    )

    if not store_attr_rows:
        return 0

    snapshot_payload = [
        (
            target_month,
            target_version,
            row.store_id,
            row.store_code,
            row.store_name,
            row.report_channel_type,
            row.report_channel_type_group,
            row.store_grade,
            row.is_duty_free,
            row.is_include_in_daily_report,
            row.remark,
            row.effective_start_date,
            row.effective_end_date,
            file_name,
            file_md5,
            updated_by,
        )
        for row in store_attr_rows
    ]
    cursor.executemany(
        f"""
        INSERT INTO {STORE_ATTR_SNAPSHOT_TABLE_NAME} (
            target_month,
            target_version,
            store_id,
            store_code,
            store_name,
            report_channel_type,
            report_channel_type_group,
            store_grade,
            is_duty_free,
            is_include_in_daily_report,
            remark,
            effective_start_date,
            effective_end_date,
            source_file_name,
            source_file_md5,
            updated_by
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        snapshot_payload,
    )
    return cursor.rowcount


def _fetch_store_attr_overlap_rows(conn: Any, effective_start_date: date, limit: int = 20) -> list[dict[str, Any]]:
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                store_id AS store_id_alias,
                store_code AS store_code_alias,
                store_name AS store_name_alias,
                report_channel_type AS report_channel_type_alias,
                effective_start_date AS effective_start_date_alias,
                effective_end_date AS effective_end_date_alias
            FROM {STORE_ATTR_TABLE_NAME}
            WHERE %s BETWEEN effective_start_date AND effective_end_date
            ORDER BY store_id, effective_start_date
            LIMIT {limit}
            """,
            (effective_start_date,),
        )
        rows = [dict(row) for row in cursor.fetchall()]

    grouped: dict[int, list[dict[str, Any]]] = {}
    for row in rows:
        grouped.setdefault(int(row["store_id_alias"]), []).append(row)

    overlap_rows: list[dict[str, Any]] = []
    for matched_rows in grouped.values():
        if len(matched_rows) <= 1:
            continue
        for row in matched_rows:
            overlap_rows.append(
                {
                    "store_id": int(row["store_id_alias"]),
                    "store_code": row["store_code_alias"],
                    "store_name": row["store_name_alias"],
                    "report_channel_type": row["report_channel_type_alias"],
                    "effective_start_date": row["effective_start_date_alias"].isoformat(),
                    "effective_end_date": row["effective_end_date_alias"].isoformat(),
                }
            )
            if len(overlap_rows) >= limit:
                return overlap_rows
    return overlap_rows


def _serialize_current_store_attr_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "store_id": int(row["store_id"]),
        "store_code": row["store_code"],
        "store_name": row["store_name"],
        "report_channel_type": row["report_channel_type"],
        REPORT_CHANNEL_TYPE_GROUP_COLUMN: row.get(REPORT_CHANNEL_TYPE_GROUP_COLUMN)
        or _derive_report_channel_type_group(row.get("report_channel_type"), allow_unknown=True),
        "store_grade": row.get("store_grade"),
        "is_duty_free": row.get("is_duty_free"),
        "is_include_in_daily_report": row.get("is_include_in_daily_report"),
        "remark": row.get("remark"),
        "effective_start_date": row["effective_start_date"].isoformat(),
        "effective_end_date": row["effective_end_date"].isoformat(),
    }


def _serialize_candidate_store_attr_row(row: StoreAttrRow) -> dict[str, Any]:
    return {
        "store_id": row.store_id,
        "store_code": row.store_code,
        "store_name": row.store_name,
        "report_channel_type": row.report_channel_type,
        REPORT_CHANNEL_TYPE_GROUP_COLUMN: row.report_channel_type_group,
        "store_grade": row.store_grade,
        "is_duty_free": row.is_duty_free,
        "is_include_in_daily_report": row.is_include_in_daily_report,
        "remark": row.remark,
        "effective_start_date": row.effective_start_date.isoformat(),
        "effective_end_date": row.effective_end_date.isoformat(),
    }


def _load_current_effective_store_attr(
    conn: Any,
    compare_date: date,
) -> tuple[list[dict[str, Any]], dict[int, dict[str, Any]], list[dict[str, Any]]]:
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                store_id AS store_id_alias,
                store_code AS store_code_alias,
                store_name AS store_name_alias,
                report_channel_type AS report_channel_type_alias,
                store_grade AS store_grade_alias,
                is_duty_free AS is_duty_free_alias,
                is_include_in_daily_report AS is_include_in_daily_report_alias,
                remark AS remark_alias,
                effective_start_date AS effective_start_date_alias,
                effective_end_date AS effective_end_date_alias
            FROM {STORE_ATTR_TABLE_NAME}
            WHERE is_include_in_daily_report = 'Y'
              AND %s BETWEEN effective_start_date AND effective_end_date
            ORDER BY store_id, effective_start_date
            """,
            (compare_date,),
        )
        rows = [dict(row) for row in cursor.fetchall()]

    grouped: dict[int, list[dict[str, Any]]] = {}
    for row in rows:
        current_row = {
            "store_id": int(row["store_id_alias"]),
            "store_code": row["store_code_alias"],
            "store_name": row["store_name_alias"],
            "report_channel_type": row["report_channel_type_alias"],
            REPORT_CHANNEL_TYPE_GROUP_COLUMN: _derive_report_channel_type_group(
                row["report_channel_type_alias"],
                allow_unknown=True,
            ),
            "store_grade": row["store_grade_alias"],
            "is_duty_free": row["is_duty_free_alias"],
            "is_include_in_daily_report": row["is_include_in_daily_report_alias"],
            "remark": row["remark_alias"],
            "effective_start_date": row["effective_start_date_alias"],
            "effective_end_date": row["effective_end_date_alias"],
        }
        grouped.setdefault(current_row["store_id"], []).append(current_row)

    current_rows: list[dict[str, Any]] = []
    overlap_rows: list[dict[str, Any]] = []
    current_map: dict[int, dict[str, Any]] = {}
    for store_id, matched_rows in grouped.items():
        if len(matched_rows) > 1:
            overlap_rows.extend(matched_rows)
            continue
        current_rows.append(matched_rows[0])
        current_map[store_id] = matched_rows[0]

    return current_rows, current_map, overlap_rows


def _build_candidate_store_attr_map(
    store_attr_rows: list[StoreAttrRow],
) -> tuple[dict[int, StoreAttrRow], list[dict[str, Any]]]:
    candidate_map: dict[int, StoreAttrRow] = {}
    duplicated_store_ids: list[dict[str, Any]] = []
    for row in store_attr_rows:
        if row.store_id in candidate_map:
            duplicated_store_ids.append(
                {
                    "store_id": row.store_id,
                    "first_store_name": candidate_map[row.store_id].store_name,
                    "duplicate_store_name": row.store_name,
                }
            )
            continue
        candidate_map[row.store_id] = row

    return candidate_map, duplicated_store_ids


def _detect_store_attr_changed_fields(current_row: dict[str, Any], candidate_row: StoreAttrRow) -> list[str]:
    changed_fields: list[str] = []
    for field_name in STORE_ATTR_COMPARE_FIELDS:
        current_value = _normalize_text(current_row.get(field_name))
        candidate_value = _normalize_text(getattr(candidate_row, field_name))
        if current_value != candidate_value:
            changed_fields.append(field_name)
    return changed_fields


def _classify_store_attr_rows(
    current_map: dict[int, dict[str, Any]],
    candidate_map: dict[int, StoreAttrRow],
) -> dict[str, list[dict[str, Any]]]:
    unchanged_rows: list[dict[str, Any]] = []
    changed_rows: list[dict[str, Any]] = []
    new_rows: list[dict[str, Any]] = []
    exited_rows: list[dict[str, Any]] = []

    for store_id in sorted(set(current_map) | set(candidate_map)):
        current_row = current_map.get(store_id)
        candidate_row = candidate_map.get(store_id)
        if current_row is not None and candidate_row is not None:
            changed_fields = _detect_store_attr_changed_fields(current_row, candidate_row)
            diff_payload = {
                "store_id": store_id,
                "changed_fields": changed_fields,
                "current": current_row,
                "candidate": candidate_row,
            }
            if changed_fields:
                changed_rows.append(diff_payload)
            else:
                unchanged_rows.append(diff_payload)
            continue

        if candidate_row is not None:
            new_rows.append({"store_id": store_id, "candidate": candidate_row})
            continue

        if current_row is not None:
            exited_rows.append({"store_id": store_id, "current": current_row})

    return {
        "unchanged_rows": unchanged_rows,
        "changed_rows": changed_rows,
        "new_rows": new_rows,
        "exited_rows": exited_rows,
    }


def _derive_duty_free_flag(
    store_name: str,
    report_channel_type: str | None,
    latest_store_attr: dict[str, Any] | None,
) -> str:
    if DUTY_FREE_KEYWORD in _normalize_text(report_channel_type):
        return "Y"
    return "N"


def _build_store_attr_rows(
    parsed_rows: list[SourceRow],
    store_map: dict[str, dict[str, Any]],
    latest_store_attr_map: dict[int, dict[str, Any]],
    effective_start_date: date,
) -> list[StoreAttrRow]:
    store_attr_rows: list[StoreAttrRow] = []
    for parsed_row in parsed_rows:
        store_info = store_map.get(_normalize_store_key(parsed_row.store_name))
        if store_info is None or parsed_row.store_type is None:
            continue

        store_id = int(store_info["store_id"])
        latest_store_attr = latest_store_attr_map.get(store_id)
        target_month = parsed_row.target_month
        target_version = parsed_row.target_version
        report_channel_type_group = _derive_report_channel_type_group(parsed_row.store_type)
        resolved_store_grade = (
            parsed_row.store_grade
            if parsed_row.store_grade is not None
            else latest_store_attr.get("store_grade") if latest_store_attr else None
        )
        row_effective_start_date = max(parsed_row.effective_start_date, effective_start_date)
        row_effective_end_date = parsed_row.effective_end_date
        if row_effective_start_date > row_effective_end_date:
            continue

        store_attr_rows.append(
            StoreAttrRow(
                store_id=store_id,
                store_code=str(store_info.get("store_code") or ""),
                store_name=parsed_row.store_name,
                report_channel_type=parsed_row.store_type,
                report_channel_type_group=report_channel_type_group,
                store_grade=resolved_store_grade,
                is_duty_free=_derive_duty_free_flag(parsed_row.store_name, parsed_row.store_type, latest_store_attr),
                is_include_in_daily_report="Y",
                remark=f"NAS导入:{target_month}/{target_version}/门店类型={parsed_row.store_type}",
                effective_start_date=row_effective_start_date,
                effective_end_date=DEFAULT_EFFECTIVE_END_DATE,
            )
        )

    return sorted(store_attr_rows, key=lambda row: (row.effective_start_date, row.store_id))


def _close_or_delete_store_attr_row(
    cursor: Any,
    current_row: dict[str, Any],
    effective_start_date: date,
    updated_by: str,
) -> str:
    current_start_date = current_row["effective_start_date"]
    current_end_date = current_row["effective_end_date"]
    if current_start_date == effective_start_date:
        cursor.execute(
            f"""
            DELETE FROM {STORE_ATTR_TABLE_NAME}
            WHERE store_id = %s
              AND effective_start_date = %s
              AND effective_end_date = %s
            """,
            (current_row["store_id"], current_start_date, current_end_date),
        )
        return "deleted"

    cursor.execute(
        f"""
        UPDATE {STORE_ATTR_TABLE_NAME}
        SET effective_end_date = %s,
            updated_by = %s
        WHERE store_id = %s
          AND effective_start_date = %s
          AND effective_end_date = %s
        """,
        (
            effective_start_date - date.resolution,
            updated_by,
            current_row["store_id"],
            current_start_date,
            current_end_date,
        ),
    )
    return "closed"


def _insert_store_attr_row(cursor: Any, row: StoreAttrRow, updated_by: str) -> int:
    cursor.execute(
        f"""
        INSERT INTO {STORE_ATTR_TABLE_NAME} (
            store_id,
            store_code,
            store_name,
            report_channel_type,
            store_grade,
            is_duty_free,
            is_include_in_daily_report,
            remark,
            effective_start_date,
            effective_end_date,
            updated_by
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """,
        (
            row.store_id,
            row.store_code,
            row.store_name,
            row.report_channel_type,
            row.store_grade,
            row.is_duty_free,
            row.is_include_in_daily_report,
            row.remark,
            row.effective_start_date,
            row.effective_end_date,
            updated_by,
        ),
    )
    return cursor.rowcount


def _expand_rows(parsed_rows: list[SourceRow], store_map: dict[str, dict[str, Any]]) -> tuple[list[ExpandedRow], list[str]]:
    expanded_rows: list[ExpandedRow] = []
    missing_store_names: list[str] = []

    for parsed_row in parsed_rows:
        store_info = store_map.get(_normalize_store_key(parsed_row.store_name))
        if store_info is None:
            missing_store_names.append(parsed_row.store_name)
            continue

        for day, day_target in sorted(parsed_row.day_targets.items()):
            target_date = date(
                parsed_row.target_month_start.year,
                parsed_row.target_month_start.month,
                day,
            )
            if target_date < parsed_row.effective_start_date or target_date > parsed_row.effective_end_date:
                continue
            expanded_rows.append(
                ExpandedRow(
                    target_date=target_date,
                    store_name=parsed_row.store_name,
                    store_id=int(store_info["store_id"]),
                    store_code=store_info.get("store_code"),
                    month_target=parsed_row.month_target,
                    day_target=day_target,
                    target_version=parsed_row.target_version,
                )
            )

    return expanded_rows, sorted(set(missing_store_names))


def _build_missing_store_suggestions(missing_store_names: list[str], store_map: dict[str, dict[str, Any]]) -> dict[str, list[str]]:
    available_names = sorted({row["store_name"] for row in store_map.values()})
    suggestions: dict[str, list[str]] = {}
    for missing_name in missing_store_names:
        matched = difflib.get_close_matches(missing_name, available_names, n=3, cutoff=0.45)
        suggestions[missing_name] = matched
    return suggestions


def _fetch_existing_target_count(conn: Any, month_start: date, month_end: date, target_version: str) -> int:
    with conn.cursor() as cursor:
        cursor.execute(
            """
            SELECT COUNT(*) AS row_count_alias
            FROM cfg_store_target_daily
            WHERE target_date BETWEEN %s AND %s
              AND target_version = %s
            """,
            (month_start, month_end, target_version),
        )
        row = cursor.fetchone()
    if row is None:
        return 0
    return int(dict(row)["row_count_alias"] or 0)


def _ensure_log_table_exists() -> None:
    ddl_hint = LOG_TABLE_SQL_PATH.relative_to(REPO_ROOT).as_posix()
    conn = _connect()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                """
                SELECT table_name AS table_name_alias
                FROM information_schema.tables
                WHERE table_schema = DATABASE()
                  AND table_name = %s
                """,
                (LOG_TABLE_NAME,),
            )
            row = cursor.fetchone()
    finally:
        conn.close()

    if row is None:
        raise RuntimeError(
            f"未找到日志表 {LOG_TABLE_NAME}，请先执行 {ddl_hint} 建表后再使用 --apply"
        )


def _truncate_message(message: str) -> str:
    normalized = message.strip()
    if len(normalized) <= 1000:
        return normalized
    return normalized[:997] + "..."


def _write_import_log(summary: dict, status: str, message: str, inserted_rows: int) -> None:
    conn = _connect()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                f"""
                INSERT INTO {LOG_TABLE_NAME} (
                    file_name,
                    file_path,
                    file_md5,
                    source_sheet,
                    target_month,
                    target_version,
                    store_count,
                    records_total,
                    records_after_filter,
                    records_inserted,
                    status,
                    message,
                    started_at,
                    finished_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    summary["file_name"],
                    summary["file_path"],
                    summary["file_md5"],
                    summary["sheet_name"],
                    summary["target_month_start"],
                    summary["target_version"],
                    summary["matched_store_count"],
                    summary["source_row_count"],
                    summary["expanded_row_count"],
                    inserted_rows,
                    status,
                    _truncate_message(message),
                    summary["started_at"],
                    datetime.now(),
                ),
            )
        conn.commit()
    finally:
        conn.close()


def _apply_import(
    summary: dict,
    expanded_rows: list[ExpandedRow],
    created_by: str,
    sync_store_report_attr: bool,
    store_attr_rows: list[StoreAttrRow],
    store_attr_classification: dict[str, list[dict[str, Any]]],
    sync_assessment: bool,
    expanded_subject_target_rows: list[ExpandedSubjectTargetRow],
    assessment_assignment_rows: list[AssessmentAssignmentRow],
) -> dict:
    inserted_rows = 0
    inserted_store_attr_snapshot_rows = 0
    inserted_store_attr_rows = 0
    closed_store_attr_rows = 0
    deleted_store_attr_rows = 0
    inserted_subject_target_rows = 0
    inserted_assessment_assignment_rows = 0
    conn = _connect()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                f"""
                DELETE FROM {TARGET_TABLE_NAME}
                WHERE target_date BETWEEN %s AND %s
                  AND target_version = %s
                """,
                (summary["target_month_start"], summary["target_month_end"], summary["target_version"]),
            )

            payload = [
                (
                    row.target_date,
                    row.store_id,
                    row.month_target,
                    row.day_target,
                    row.target_version,
                    created_by,
                    created_by,
                )
                for row in expanded_rows
            ]
            cursor.executemany(
                f"""
                INSERT INTO {TARGET_TABLE_NAME} (
                    target_date,
                    store_id,
                    month_target,
                    day_target,
                    target_version,
                    created_by,
                    updated_by
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                """,
                payload,
            )
            inserted_rows = cursor.rowcount

            if sync_assessment:
                cursor.execute(
                    f"""
                    DELETE FROM {SUBJECT_TARGET_TABLE_NAME}
                    WHERE target_date BETWEEN %s AND %s
                      AND target_version = %s
                    """,
                    (summary["target_month_start"], summary["target_month_end"], summary["target_version"]),
                )
                subject_target_payload = [
                    (
                        row.target_date,
                        row.target_month_start,
                        row.subject_code,
                        row.subject_name,
                        row.assessment_mode,
                        row.month_target,
                        row.day_target,
                        row.target_version,
                        row.remark,
                        created_by,
                        created_by,
                    )
                    for row in expanded_subject_target_rows
                ]
                if subject_target_payload:
                    cursor.executemany(
                        f"""
                        INSERT INTO {SUBJECT_TARGET_TABLE_NAME} (
                            target_date,
                            target_month,
                            subject_code,
                            subject_name,
                            assessment_mode,
                            month_target,
                            day_target,
                            target_version,
                            remark,
                            created_by,
                            updated_by
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        subject_target_payload,
                    )
                    inserted_subject_target_rows = cursor.rowcount

                cursor.execute(
                    f"""
                    DELETE FROM {ASSESSMENT_ASSIGNMENT_TABLE_NAME}
                    WHERE target_month = %s
                      AND target_version = %s
                    """,
                    (summary["target_month_start"], summary["target_version"]),
                )
                assignment_payload = [
                    (
                        row.target_month_start,
                        row.target_version,
                        row.store_id,
                        row.store_code,
                        row.store_name,
                        row.subject_code,
                        row.assignment_role,
                        row.is_joint_assessment,
                        row.anchor_store_id,
                        row.anchor_store_name,
                        row.effective_start_date,
                        row.effective_end_date,
                        row.remark,
                        created_by,
                        created_by,
                    )
                    for row in assessment_assignment_rows
                ]
                if assignment_payload:
                    cursor.executemany(
                        f"""
                        INSERT INTO {ASSESSMENT_ASSIGNMENT_TABLE_NAME} (
                            target_month,
                            target_version,
                            store_id,
                            store_code,
                            store_name,
                            subject_code,
                            assignment_role,
                            is_joint_assessment,
                            anchor_store_id,
                            anchor_store_name,
                            effective_start_date,
                            effective_end_date,
                            remark,
                            created_by,
                            updated_by
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        assignment_payload,
                    )
                    inserted_assessment_assignment_rows = cursor.rowcount

            if sync_store_report_attr:
                effective_start_date = summary["store_attr_effective_start_date"]
                inserted_store_attr_snapshot_rows = _replace_store_attr_snapshot_rows(
                    cursor,
                    summary["target_month_start"],
                    summary["target_version"],
                    store_attr_rows,
                    summary["file_name"],
                    summary["file_md5"],
                    created_by,
                )

                snapshot_store_attr_rows = _load_store_attr_snapshot_rows(
                    conn,
                    summary["target_month_start"],
                    summary["target_version"],
                )
                current_store_attr_rows, current_store_attr_map, current_store_attr_overlap_rows = _load_current_effective_store_attr(
                    conn,
                    effective_start_date,
                )
                if current_store_attr_overlap_rows:
                    raise RuntimeError(
                        f"{STORE_ATTR_TABLE_NAME} 在 {effective_start_date.isoformat()} 存在同店多条当前有效配置，"
                        f"当前不能安全同步历史表；请先清理真实重叠记录后再执行。样例: "
                        f"{[_serialize_current_store_attr_row(row) for row in current_store_attr_overlap_rows[:20]]}"
                    )
                snapshot_store_attr_map, snapshot_duplicate_store_ids = _build_candidate_store_attr_map(snapshot_store_attr_rows)
                if snapshot_duplicate_store_ids:
                    raise RuntimeError(
                        f"{STORE_ATTR_SNAPSHOT_TABLE_NAME} 中存在重复 store_id，无法安全同步历史表: "
                        f"{snapshot_duplicate_store_ids}"
                    )
                store_attr_classification = _classify_store_attr_rows(
                    current_store_attr_map,
                    snapshot_store_attr_map,
                )

                for changed_row in store_attr_classification["changed_rows"]:
                    action = _close_or_delete_store_attr_row(
                        cursor,
                        changed_row["current"],
                        effective_start_date,
                        created_by,
                    )
                    if action == "closed":
                        closed_store_attr_rows += 1
                    else:
                        deleted_store_attr_rows += 1
                    inserted_store_attr_rows += _insert_store_attr_row(
                        cursor,
                        changed_row["candidate"],
                        created_by,
                    )

                for new_row in store_attr_classification["new_rows"]:
                    inserted_store_attr_rows += _insert_store_attr_row(
                        cursor,
                        new_row["candidate"],
                        created_by,
                    )

                for exited_row in store_attr_classification["exited_rows"]:
                    action = _close_or_delete_store_attr_row(
                        cursor,
                        exited_row["current"],
                        effective_start_date,
                        created_by,
                    )
                    if action == "closed":
                        closed_store_attr_rows += 1
                    else:
                        deleted_store_attr_rows += 1
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    summary["records_inserted"] = inserted_rows
    summary["store_attr_snapshot_records_written"] = inserted_store_attr_snapshot_rows
    summary["store_attr_records_inserted"] = inserted_store_attr_rows
    summary["store_attr_records_closed"] = closed_store_attr_rows
    summary["store_attr_records_deleted_same_start_date"] = deleted_store_attr_rows
    summary["subject_target_records_inserted"] = inserted_subject_target_rows
    summary["assessment_assignment_records_inserted"] = inserted_assessment_assignment_rows
    return summary


def _build_summary(
    file_path: Path,
    sheet_name: str,
    sheet_names: list[str],
    file_md5: str,
    workbook_summary: dict,
    expanded_rows: list[ExpandedRow],
    missing_store_names: list[str],
    missing_store_suggestions: dict[str, list[str]],
    ambiguous_store_names: list[dict],
    existing_target_rows: int,
    preview_limit: int,
    sync_store_report_attr: bool,
    store_attr_rows: list[StoreAttrRow],
    current_store_attr_rows: list[dict[str, Any]],
    candidate_duplicate_store_ids: list[dict[str, Any]],
    store_attr_effective_start_date: date | None,
    store_attr_effective_start_source: str | None,
    existing_store_attr_rows_same_start_date: int,
    existing_store_attr_rows_in_month: int,
    store_attr_overlap_rows: list[dict[str, Any]],
    store_attr_classification: dict[str, list[dict[str, Any]]],
    sync_assessment: bool,
    subject_target_sheet_summary: dict[str, Any],
    expanded_subject_target_rows: list[ExpandedSubjectTargetRow],
    existing_subject_target_rows: int,
    assessment_assignment_sheet_summary: dict[str, Any],
    assessment_assignment_rows: list[AssessmentAssignmentRow],
    assessment_validation: dict[str, Any],
    existing_assessment_assignment_rows: int,
) -> dict:
    target_month_start = datetime.strptime(workbook_summary["target_month"], "%Y-%m").date().replace(day=1)
    target_month_end = date(
        target_month_start.year,
        target_month_start.month,
        calendar.monthrange(target_month_start.year, target_month_start.month)[1],
    )
    preview_rows = [
        {
            "target_date": row.target_date.isoformat(),
            "store_name": row.store_name,
            "store_id": row.store_id,
            "store_code": row.store_code,
            "month_target": format(row.month_target, ".2f"),
            "day_target": format(row.day_target, ".2f"),
            "target_version": row.target_version,
        }
        for row in expanded_rows[:preview_limit]
    ]
    store_type_counts: dict[str, int] = {}
    store_group_counts: dict[str, int] = {}
    for row in store_attr_rows:
        store_type_counts[row.report_channel_type] = store_type_counts.get(row.report_channel_type, 0) + 1
        store_group_counts[row.report_channel_type_group] = store_group_counts.get(row.report_channel_type_group, 0) + 1

    store_attr_preview_rows = [
        {
            "store_id": row.store_id,
            "store_code": row.store_code,
            "store_name": row.store_name,
            "report_channel_type": row.report_channel_type,
            REPORT_CHANNEL_TYPE_GROUP_COLUMN: row.report_channel_type_group,
            "store_grade": row.store_grade,
            "is_duty_free": row.is_duty_free,
            "effective_start_date": row.effective_start_date.isoformat(),
            "effective_end_date": row.effective_end_date.isoformat(),
            "remark": row.remark,
        }
        for row in store_attr_rows[:preview_limit]
    ]
    store_attr_unchanged_preview = [
        {
            "store_id": row["store_id"],
            "current": _serialize_current_store_attr_row(row["current"]),
            "candidate": _serialize_candidate_store_attr_row(row["candidate"]),
        }
        for row in store_attr_classification["unchanged_rows"][:preview_limit]
    ]
    store_attr_changed_preview = [
        {
            "store_id": row["store_id"],
            "changed_fields": row["changed_fields"],
            "current": _serialize_current_store_attr_row(row["current"]),
            "candidate": _serialize_candidate_store_attr_row(row["candidate"]),
        }
        for row in store_attr_classification["changed_rows"][:preview_limit]
    ]
    store_attr_new_preview = [
        {
            "store_id": row["store_id"],
            "candidate": _serialize_candidate_store_attr_row(row["candidate"]),
        }
        for row in store_attr_classification["new_rows"][:preview_limit]
    ]
    store_attr_exited_preview = [
        {
            "store_id": row["store_id"],
            "current": _serialize_current_store_attr_row(row["current"]),
        }
        for row in store_attr_classification["exited_rows"][:preview_limit]
    ]

    subject_target_subject_codes = sorted({row.subject_code for row in expanded_subject_target_rows})
    assessment_assignment_subject_codes = sorted({row.subject_code for row in assessment_assignment_rows})
    subject_target_orphan_subject_codes = sorted(
        set(subject_target_subject_codes) - set(assessment_assignment_subject_codes)
    )
    assessment_assignment_missing_subject_codes = sorted(
        set(assessment_assignment_subject_codes) - set(subject_target_subject_codes)
    )
    subject_target_preview_rows = [
        {
            "target_date": row.target_date.isoformat(),
            "subject_code": row.subject_code,
            "subject_name": row.subject_name,
            "assessment_mode": row.assessment_mode,
            "month_target": format(row.month_target, ".2f"),
            "day_target": format(row.day_target, ".2f"),
            "target_version": row.target_version,
            "remark": row.remark,
        }
        for row in expanded_subject_target_rows[:preview_limit]
    ]
    assessment_assignment_preview_rows = [
        {
            "source_store_key": row.source_store_key,
            "source_store_name": row.source_store_name,
            "store_id": row.store_id,
            "store_code": row.store_code,
            "store_name": row.store_name,
            "subject_code": row.subject_code,
            "assignment_role": row.assignment_role,
            "is_joint_assessment": row.is_joint_assessment,
            "anchor_store_id": row.anchor_store_id,
            "anchor_store_name": row.anchor_store_name,
            "effective_start_date": row.effective_start_date.isoformat(),
            "effective_end_date": row.effective_end_date.isoformat(),
            "remark": row.remark,
        }
        for row in assessment_assignment_rows[:preview_limit]
    ]

    return {
        "mode": "dry-run",
        "file_name": file_path.name,
        "file_path": str(file_path),
        "file_md5": file_md5,
        "sheet_name": sheet_name,
        "sheet_names": sheet_names,
        "target_month": workbook_summary["target_month"],
        "available_target_months": workbook_summary.get("available_target_months", [workbook_summary["target_month"]]),
        "target_month_filter": workbook_summary.get("target_month_filter"),
        "target_month_start": target_month_start,
        "target_month_end": target_month_end,
        "target_version": workbook_summary["target_version"],
        "source_row_count": workbook_summary["source_row_count"],
        "matched_store_count": len({row.store_id for row in expanded_rows}),
        "expanded_row_count": len(expanded_rows),
        "existing_target_rows": existing_target_rows,
        "blank_day_cell_count": workbook_summary["blank_day_cell_count"],
        "missing_store_names": missing_store_names,
        "missing_store_suggestions": missing_store_suggestions,
        "ambiguous_store_names": ambiguous_store_names,
        "sync_store_report_attr": sync_store_report_attr,
        "store_attr_current_effective_row_count": len(current_store_attr_rows),
        "store_attr_snapshot_row_count": len(store_attr_rows),
        "store_attr_effective_start_date": store_attr_effective_start_date,
        "store_attr_effective_start_source": store_attr_effective_start_source,
        "existing_store_attr_rows_same_start_date": existing_store_attr_rows_same_start_date,
        "existing_store_attr_rows_in_target_month": existing_store_attr_rows_in_month,
        "store_attr_overlap_rows": store_attr_overlap_rows,
        "store_attr_candidate_duplicate_store_ids": candidate_duplicate_store_ids,
        "store_attr_row_count": len(store_attr_rows),
        "store_attr_type_counts": store_type_counts,
        "store_attr_group_counts": store_group_counts,
        "store_attr_diff_counts": {
            "unchanged": len(store_attr_classification["unchanged_rows"]),
            "changed": len(store_attr_classification["changed_rows"]),
            "new": len(store_attr_classification["new_rows"]),
            "exited": len(store_attr_classification["exited_rows"]),
        },
        "store_attr_action_counts": {
            "no_action": len(store_attr_classification["unchanged_rows"]),
            "close_and_open": len(store_attr_classification["changed_rows"]),
            "open_only": len(store_attr_classification["new_rows"]),
            "close_only": len(store_attr_classification["exited_rows"]),
        },
        "store_attr_preview_rows": store_attr_preview_rows,
        "store_attr_unchanged_preview": store_attr_unchanged_preview,
        "store_attr_changed_preview": store_attr_changed_preview,
        "store_attr_new_preview": store_attr_new_preview,
        "store_attr_exited_preview": store_attr_exited_preview,
        "store_attr_snapshot_records_written": 0,
        "store_attr_records_inserted": 0,
        "store_attr_records_closed": 0,
        "store_attr_records_deleted_same_start_date": 0,
        "sync_assessment": sync_assessment,
        "subject_target_sheet_present": subject_target_sheet_summary.get("sheet_present", False),
        "subject_target_sheet_name": subject_target_sheet_summary.get("sheet_name", SUBJECT_TARGET_SHEET_NAME),
        "subject_target_source_row_count": subject_target_sheet_summary.get("source_row_count", 0),
        "subject_target_blank_day_cell_count": subject_target_sheet_summary.get("blank_day_cell_count", 0),
        "subject_target_expanded_row_count": len(expanded_subject_target_rows),
        "subject_target_subject_count": len(subject_target_subject_codes),
        "subject_target_existing_rows": existing_subject_target_rows,
        "subject_target_orphan_subject_codes": subject_target_orphan_subject_codes,
        "subject_target_preview_rows": subject_target_preview_rows,
        "assessment_assignment_sheet_present": assessment_assignment_sheet_summary.get("sheet_present", False),
        "assessment_assignment_sheet_name": assessment_assignment_sheet_summary.get(
            "sheet_name",
            ASSESSMENT_ASSIGNMENT_SHEET_NAME,
        ),
        "assessment_assignment_source_row_count": assessment_assignment_sheet_summary.get("source_row_count", 0),
        "assessment_assignment_row_count": len(assessment_assignment_rows),
        "assessment_assignment_subject_count": len(assessment_assignment_subject_codes),
        "assessment_assignment_store_count": len({row.store_id for row in assessment_assignment_rows}),
        "assessment_assignment_existing_rows": existing_assessment_assignment_rows,
        "assessment_assignment_missing_store_keys": assessment_validation.get("missing_store_keys", []),
        "assessment_assignment_missing_store_names": assessment_validation.get("missing_store_names", []),
        "assessment_assignment_missing_store_suggestions": assessment_validation.get("missing_store_suggestions", {}),
        "assessment_assignment_store_key_name_mismatch_rows": assessment_validation.get(
            "store_key_name_mismatch_rows",
            [],
        ),
        "assessment_assignment_missing_anchor_store_names": assessment_validation.get("missing_anchor_store_names", []),
        "assessment_assignment_missing_anchor_store_suggestions": assessment_validation.get(
            "missing_anchor_store_suggestions",
            {},
        ),
        "assessment_assignment_overlap_rows": assessment_validation.get("overlap_rows", []),
        "assessment_assignment_unanchored_rows": assessment_validation.get("unanchored_rows", []),
        "assessment_assignment_inconsistent_primary_anchor_rows": assessment_validation.get(
            "inconsistent_primary_anchor_rows",
            [],
        ),
        "assessment_assignment_skipped_missing_primary_rows": assessment_validation.get(
            "skipped_missing_primary_rows",
            [],
        ),
        "assessment_assignment_missing_subject_codes": assessment_assignment_missing_subject_codes,
        "assessment_assignment_preview_rows": assessment_assignment_preview_rows,
        "instruction_sheet_present": INSTRUCTION_SHEET_NAME in sheet_names,
        "subject_target_records_inserted": 0,
        "assessment_assignment_records_inserted": 0,
        "preview_rows": preview_rows,
        "started_at": datetime.now(),
    }


def _to_printable_summary(summary: dict) -> dict:
    printable = dict(summary)
    printable["target_month_start"] = summary["target_month_start"].isoformat()
    printable["target_month_end"] = summary["target_month_end"].isoformat()
    store_attr_effective_start_date = summary.get("store_attr_effective_start_date")
    if isinstance(store_attr_effective_start_date, date):
        printable["store_attr_effective_start_date"] = store_attr_effective_start_date.isoformat()
    started_at = summary.get("started_at")
    if isinstance(started_at, datetime):
        printable["started_at"] = started_at.isoformat(timespec="seconds")
    return printable


def _format_warning_preview(items: list[str], limit: int = 10) -> list[str]:
    normalized_items = [str(item) for item in items if _normalize_text(item)]
    if len(normalized_items) <= limit:
        return normalized_items
    remaining = len(normalized_items) - limit
    return normalized_items[:limit] + [f"...其余{remaining}项"]


def _format_assignment_warning_preview(rows: list[dict[str, Any]], limit: int = 10) -> list[str]:
    preview: list[str] = []
    for row in rows[:limit]:
        store_name = row.get('store_name') or row.get('excel_store_name') or row.get('dim_store_name') or '-'
        preview.append(
            f"row={row.get('row_number')} store={store_name} source_key={row.get('source_store_key')} "
            f"match={row.get('matched_by')} store_id={row.get('store_id')} subject={row.get('subject_code')}"
        )
    remaining = len(rows) - len(preview)
    if remaining > 0:
        preview.append(f"...其余{remaining}项")
    return preview


def _build_validation_error_message(summary: dict) -> str:
    if summary.get("ambiguous_store_names"):
        return (
            "dim_store 中存在门店名称歧义，无法按 store_name 精确映射: "
            f"{summary['ambiguous_store_names']}"
        )
    if summary.get("source_row_count", 0) > 0 and summary.get("matched_store_count", 0) == 0:
        return (
            f"导入模板中的门店全部未命中 {STORE_TABLE_NAME}，无法安全覆盖 {TARGET_TABLE_NAME}；"
            "请先核对配置表或补齐 Oracle/ERP 建店后重试。"
        )
    if summary.get("store_attr_candidate_duplicate_store_ids"):
        return (
            "候选门店属性快照存在重复 store_id，无法安全执行新增/退出/变更处理: "
            f"{summary['store_attr_candidate_duplicate_store_ids']}"
        )
    if summary.get("sync_store_report_attr") and summary.get("store_attr_overlap_rows"):
        return (
            f"{STORE_ATTR_TABLE_NAME} 在 {summary['store_attr_effective_start_date'].isoformat()} "
            f"存在同店多条当前有效配置，当前不能安全同步历史表；"
            f"请先清理真实重叠记录后再执行。样例: {summary['store_attr_overlap_rows']}"
        )
    if (
        summary.get("sync_assessment")
        and summary.get("assessment_assignment_source_row_count", 0) > 0
        and summary.get("assessment_assignment_row_count", 0) == 0
    ):
        return (
            f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 的门店归属行全部因门店ID/门店编码、门店名称或主店未命中 {STORE_TABLE_NAME} 被跳过，"
            f"无法安全覆盖 {ASSESSMENT_ASSIGNMENT_TABLE_NAME}；"
            "请先核对配置表或补齐 Oracle/ERP 建店后重试。"
        )
    if summary.get("assessment_assignment_inconsistent_primary_anchor_rows"):
        return (
            f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 存在主店行挂靠到其他门店的异常配置: "
            f"{summary['assessment_assignment_inconsistent_primary_anchor_rows']}"
        )
    if summary.get("assessment_assignment_unanchored_rows"):
        return (
            f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 存在共同考核但无法确定主店的门店行: "
            f"{summary['assessment_assignment_unanchored_rows']}"
        )
    if summary.get("assessment_assignment_overlap_rows"):
        return (
            f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 存在同门店生效区间重叠，不能安全导入: "
            f"{summary['assessment_assignment_overlap_rows']}"
        )
    if summary.get("assessment_assignment_missing_subject_codes"):
        return (
            f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 引用了未在 {SUBJECT_TARGET_SHEET_NAME} 配置的主体编码: "
            f"{summary['assessment_assignment_missing_subject_codes']}"
        )
    return ""


def _build_validation_warning_messages(summary: dict) -> list[str]:
    warning_messages: list[str] = []
    if summary.get("missing_store_names"):
        warning_messages.append(
            f"以下门店名称在 {STORE_TABLE_NAME} 中未命中，已跳过这些门店的目标/门店属性配置: "
            f"{_format_warning_preview(summary['missing_store_names'])}；候选建议: "
            f"{summary['missing_store_suggestions']}"
        )
    if summary.get("assessment_assignment_missing_store_keys"):
        warning_messages.append(
            f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 中以下门店ID/门店编码在 {STORE_TABLE_NAME} 未命中，"
            f"已跳过这些归属配置: {_format_warning_preview(summary['assessment_assignment_missing_store_keys'])}"
        )
    if summary.get("assessment_assignment_missing_store_names"):
        warning_messages.append(
            f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 中以下门店名称仅作辅助展示，但其对应门店ID/门店编码未成功命中 {STORE_TABLE_NAME}，"
            f"已跳过这些归属配置: {_format_warning_preview(summary['assessment_assignment_missing_store_names'])}；"
            f"候选建议: {summary['assessment_assignment_missing_store_suggestions']}"
        )
    if summary.get("assessment_assignment_store_key_name_mismatch_rows"):
        warning_messages.append(
            f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 中以下行的门店ID/门店编码已命中，但 Excel 门店名称与 {STORE_TABLE_NAME} 当前名称不一致，"
            f"本次已按该门店标识继续导入: {_format_assignment_warning_preview(summary['assessment_assignment_store_key_name_mismatch_rows'])}"
        )
    if summary.get("assessment_assignment_missing_anchor_store_names"):
        warning_messages.append(
            f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 中以下主店名称在 {STORE_TABLE_NAME} 未命中，"
            f"已跳过相关归属配置: {_format_warning_preview(summary['assessment_assignment_missing_anchor_store_names'])}；"
            f"候选建议: {summary['assessment_assignment_missing_anchor_store_suggestions']}"
        )
    if summary.get("assessment_assignment_skipped_missing_primary_rows"):
        warning_messages.append(
            f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 中以下门店因所属主店未命中 {STORE_TABLE_NAME} 被连带跳过: "
            f"{_format_assignment_warning_preview(summary['assessment_assignment_skipped_missing_primary_rows'])}"
        )
    return warning_messages


def _write_output_json(output_path: Path, payload: dict) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_failure_output_json(
    output_json: str | None,
    args: argparse.Namespace,
    error_message: str,
    validation_status: str,
) -> None:
    if not output_json:
        return

    payload = {
        "mode": "apply" if args.apply else "dry-run",
        "sheet_name": args.sheet_name,
        "file_path": args.file_path,
        "target_month_filter": args.target_month,
        "validation_status": validation_status,
        "error_message": error_message,
    }
    _write_output_json(Path(output_json), payload)


def main() -> int:
    parser = argparse.ArgumentParser(description="从 NAS 导入 cfg_store_target_daily，默认只做 dry-run")
    parser.add_argument(
        "--file-path",
        default=None,
        help="可选：显式指定目标 Excel 路径；不传时按 NAS 目录与 --target-month 自动解析文件",
    )
    parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME, help="导入模板工作表名称")
    parser.add_argument(
        "--target-month",
        type=_parse_target_month_arg,
        help="可选：显式指定需导入的目标月份；不传 --file-path 时会按该月份自动选择 NAS 文件，格式 YYYY-MM",
    )
    parser.add_argument("--preview-limit", type=int, default=10, help="dry-run 输出预览行数")
    parser.add_argument("--output-json", help="可选：将 dry-run / apply 摘要写入 JSON 文件")
    parser.add_argument("--created-by", default=DEFAULT_CREATED_BY, help="正式导入时写入 created_by / updated_by")
    parser.add_argument(
        "--sync-store-report-attr",
        action="store_true",
        help=f"同步 {STORE_ATTR_TABLE_NAME}；要求模板存在 {STORE_TYPE_HEADER} 列",
    )
    parser.add_argument(
        "--attr-effective-start-date",
        type=date.fromisoformat,
        help="可选：显式指定 dim_store_report_attr 生效开始日，格式 YYYY-MM-DD",
    )
    parser.add_argument("--apply", action="store_true", help="显式启用写库模式；默认只做 dry-run")
    args = parser.parse_args()
    try:
        file_path = _resolve_input_file(args.file_path, args.target_month)
        sheet_names = _load_workbook_sheetnames(file_path)

        file_md5 = _compute_file_md5(file_path)
        parsed_rows, workbook_summary = _parse_workbook(
            file_path,
            args.sheet_name,
            require_store_type=args.sync_store_report_attr,
            target_month_filter=args.target_month,
        )
        subject_source_rows, subject_target_sheet_summary = _parse_optional_subject_target_sheet(
            file_path,
            args.target_month,
        )
        parsed_assessment_assignment_rows, assessment_assignment_sheet_summary = _parse_optional_assessment_assignment_sheet(
            file_path,
            args.target_month,
        )

        subject_target_sheet_present = bool(subject_target_sheet_summary.get("sheet_present"))
        assessment_assignment_sheet_present = bool(assessment_assignment_sheet_summary.get("sheet_present"))
        if subject_target_sheet_present != assessment_assignment_sheet_present:
            raise ValueError(
                f"工作表 {SUBJECT_TARGET_SHEET_NAME} 与 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 必须同时存在；"
                "旧模板可两张都不存在，新模板需两张一起提供"
            )

        sync_assessment = subject_target_sheet_present and assessment_assignment_sheet_present
        subject_target_has_data = bool(subject_source_rows)
        assessment_assignment_has_data = bool(parsed_assessment_assignment_rows)
        if sync_assessment and subject_target_has_data != assessment_assignment_has_data:
            raise ValueError(
                f"工作表 {SUBJECT_TARGET_SHEET_NAME} 与 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 必须同时提供有效数据，"
                "或同时留空以清空当月共同考核配置"
            )
        if subject_target_has_data:
            if subject_target_sheet_summary.get("target_month") != workbook_summary["target_month"]:
                raise ValueError(
                    f"工作表 {SUBJECT_TARGET_SHEET_NAME} 的目标月份 {subject_target_sheet_summary.get('target_month')} "
                    f"与 {args.sheet_name} 的 {workbook_summary['target_month']} 不一致"
                )
            if subject_target_sheet_summary.get("target_version") != workbook_summary["target_version"]:
                raise ValueError(
                    f"工作表 {SUBJECT_TARGET_SHEET_NAME} 的目标版本 {subject_target_sheet_summary.get('target_version')} "
                    f"与 {args.sheet_name} 的 {workbook_summary['target_version']} 不一致"
                )
        if assessment_assignment_has_data:
            if assessment_assignment_sheet_summary.get("target_month") != workbook_summary["target_month"]:
                raise ValueError(
                    f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 的目标月份 "
                    f"{assessment_assignment_sheet_summary.get('target_month')} 与 {args.sheet_name} 的 "
                    f"{workbook_summary['target_month']} 不一致"
                )
            if assessment_assignment_sheet_summary.get("target_version") != workbook_summary["target_version"]:
                raise ValueError(
                    f"工作表 {ASSESSMENT_ASSIGNMENT_SHEET_NAME} 的目标版本 "
                    f"{assessment_assignment_sheet_summary.get('target_version')} 与 {args.sheet_name} 的 "
                    f"{workbook_summary['target_version']} 不一致"
                )

        target_month_start = datetime.strptime(workbook_summary["target_month"], "%Y-%m").date().replace(day=1)
        target_month_end = date(
            target_month_start.year,
            target_month_start.month,
            calendar.monthrange(target_month_start.year, target_month_start.month)[1],
        )

        conn = _connect()
        try:
            _ensure_required_tables(
                conn,
                sync_store_report_attr=args.sync_store_report_attr,
                sync_assessment=sync_assessment,
            )
            store_map, store_map_by_id, store_map_by_code, ambiguous_store_names = _load_store_mapping(conn)
            expanded_rows, missing_store_names = _expand_rows(parsed_rows, store_map)
            missing_store_suggestions = _build_missing_store_suggestions(missing_store_names, store_map)
            existing_target_rows = _fetch_existing_target_count(
                conn,
                target_month_start,
                target_month_end,
                workbook_summary["target_version"],
            )
            latest_store_attr_map: dict[int, dict[str, Any]] = {}
            store_attr_effective_start_date: date | None = None
            store_attr_effective_start_source: str | None = None
            existing_store_attr_rows_same_start_date = 0
            existing_store_attr_rows_in_month = 0
            current_store_attr_rows: list[dict[str, Any]] = []
            store_attr_overlap_rows: list[dict[str, Any]] = []
            store_attr_rows: list[StoreAttrRow] = []
            candidate_duplicate_store_ids: list[dict[str, Any]] = []
            store_attr_classification = {
                "unchanged_rows": [],
                "changed_rows": [],
                "new_rows": [],
                "exited_rows": [],
            }
            expanded_subject_target_rows: list[ExpandedSubjectTargetRow] = []
            assessment_assignment_rows: list[AssessmentAssignmentRow] = []
            existing_subject_target_rows = 0
            existing_assessment_assignment_rows = 0
            assessment_validation = {
                "missing_store_keys": [],
                "missing_store_names": [],
                "missing_store_suggestions": {},
                "store_key_name_mismatch_rows": [],
                "missing_anchor_store_names": [],
                "missing_anchor_store_suggestions": {},
                "overlap_rows": [],
                "unanchored_rows": [],
                "inconsistent_primary_anchor_rows": [],
                "skipped_missing_primary_rows": [],
            }
            if args.sync_store_report_attr:
                latest_store_attr_map = _load_latest_store_attr(conn)
                (
                    store_attr_effective_start_date,
                    store_attr_effective_start_source,
                    existing_store_attr_rows_same_start_date,
                    existing_store_attr_rows_in_month,
                ) = _resolve_store_attr_effective_start_date(
                    conn,
                    target_month_start,
                    target_month_end,
                    args.attr_effective_start_date,
                )
                current_store_attr_rows, current_store_attr_map, current_store_attr_overlap_rows = _load_current_effective_store_attr(
                    conn,
                    store_attr_effective_start_date,
                )
                store_attr_overlap_rows = [
                    _serialize_current_store_attr_row(row)
                    for row in current_store_attr_overlap_rows
                ]
                store_attr_rows = _build_store_attr_rows(
                    parsed_rows,
                    store_map,
                    latest_store_attr_map,
                    store_attr_effective_start_date,
                )
                candidate_store_attr_map, candidate_duplicate_store_ids = _build_candidate_store_attr_map(store_attr_rows)
                store_attr_classification = _classify_store_attr_rows(
                    current_store_attr_map,
                    candidate_store_attr_map,
                )

            if sync_assessment:
                expanded_subject_target_rows = _expand_subject_target_rows(subject_source_rows)
                if parsed_assessment_assignment_rows:
                    assessment_assignment_rows, resolved_validation = _resolve_assessment_assignments(
                        parsed_assessment_assignment_rows,
                        store_map,
                        store_map_by_id,
                        store_map_by_code,
                    )
                    assessment_validation.update(resolved_validation)
                    assessment_validation["missing_store_suggestions"] = _build_missing_store_suggestions(
                        assessment_validation["missing_store_names"],
                        store_map,
                    )
                    assessment_validation["missing_anchor_store_suggestions"] = _build_missing_store_suggestions(
                        assessment_validation["missing_anchor_store_names"],
                        store_map,
                    )
                    assessment_validation["overlap_rows"] = _find_assignment_overlap_rows(assessment_assignment_rows)

                existing_subject_target_rows = _fetch_existing_subject_target_count(
                    conn,
                    target_month_start,
                    target_month_end,
                    workbook_summary["target_version"],
                )
                existing_assessment_assignment_rows = _fetch_existing_assessment_assignment_count(
                    conn,
                    target_month_start,
                    workbook_summary["target_version"],
                )
        finally:
            conn.close()

        summary = _build_summary(
            file_path=file_path,
            sheet_name=args.sheet_name,
            sheet_names=sheet_names,
            file_md5=file_md5,
            workbook_summary=workbook_summary,
            expanded_rows=expanded_rows,
            missing_store_names=missing_store_names,
            missing_store_suggestions=missing_store_suggestions,
            ambiguous_store_names=ambiguous_store_names,
            existing_target_rows=existing_target_rows,
            preview_limit=max(args.preview_limit, 0),
            sync_store_report_attr=args.sync_store_report_attr,
            store_attr_rows=store_attr_rows,
            current_store_attr_rows=current_store_attr_rows,
            candidate_duplicate_store_ids=candidate_duplicate_store_ids,
            store_attr_effective_start_date=store_attr_effective_start_date,
            store_attr_effective_start_source=store_attr_effective_start_source,
            existing_store_attr_rows_same_start_date=existing_store_attr_rows_same_start_date,
            existing_store_attr_rows_in_month=existing_store_attr_rows_in_month,
            store_attr_overlap_rows=store_attr_overlap_rows,
            store_attr_classification=store_attr_classification,
            sync_assessment=sync_assessment,
            subject_target_sheet_summary=subject_target_sheet_summary,
            expanded_subject_target_rows=expanded_subject_target_rows,
            existing_subject_target_rows=existing_subject_target_rows,
            assessment_assignment_sheet_summary=assessment_assignment_sheet_summary,
            assessment_assignment_rows=assessment_assignment_rows,
            assessment_validation=assessment_validation,
            existing_assessment_assignment_rows=existing_assessment_assignment_rows,
        )

        validation_error_message = _build_validation_error_message(summary)
        warning_messages = _build_validation_warning_messages(summary)
        summary["warning_messages"] = warning_messages
        summary["warning_count"] = len(warning_messages)
        summary["validation_status"] = "FAILED" if validation_error_message else ("WARNING" if warning_messages else "PASSED")

        if not args.apply:
            printable = _to_printable_summary(summary)
            if validation_error_message:
                printable["error_message"] = validation_error_message
            if args.output_json:
                _write_output_json(Path(args.output_json), printable)
            print(json.dumps(printable, ensure_ascii=False, indent=2))
            return 1 if validation_error_message else 0

        if validation_error_message:
            raise ValueError(validation_error_message)

        _ensure_log_table_exists()
        summary["mode"] = "apply"
        summary["started_at"] = datetime.now()

        try:
            _apply_import(
                summary,
                expanded_rows,
                args.created_by,
                sync_store_report_attr=args.sync_store_report_attr,
                store_attr_rows=store_attr_rows,
                store_attr_classification=store_attr_classification,
                sync_assessment=sync_assessment,
                expanded_subject_target_rows=expanded_subject_target_rows,
                assessment_assignment_rows=assessment_assignment_rows,
            )
            success_message = (
                f"month={summary['target_month']}, version={summary['target_version']}, "
                f"stores={summary['matched_store_count']}, inserted={summary['records_inserted']}"
            )
            if warning_messages:
                success_message += f", warnings={len(warning_messages)}"
            if args.sync_store_report_attr:
                success_message += (
                    f", store_attr_snapshot_written={summary['store_attr_snapshot_records_written']}"
                    f", store_attr_start={summary['store_attr_effective_start_date'].isoformat()}, "
                    f"store_attr_inserted={summary['store_attr_records_inserted']}"
                )
            if sync_assessment:
                success_message += (
                    f", subject_targets_inserted={summary['subject_target_records_inserted']}, "
                    f"assessment_assignments_inserted={summary['assessment_assignment_records_inserted']}"
                )
            _write_import_log(summary, "SUCCESS", success_message, summary["records_inserted"])
            printable = _to_printable_summary(summary)
            if args.output_json:
                _write_output_json(Path(args.output_json), printable)
            print(json.dumps(printable, ensure_ascii=False, indent=2))
            return 0
        except Exception as exc:
            failure_message = str(exc)
            try:
                _write_import_log(summary, "FAILED", failure_message, 0)
            except Exception:
                pass
            raise RuntimeError(f"导入失败: {failure_message}") from exc
    except ValueError as exc:
        _write_failure_output_json(args.output_json, args, str(exc), "FAILED")
        print(str(exc), file=sys.stderr)
        return 1
    except Exception as exc:
        _write_failure_output_json(args.output_json, args, str(exc), "ERROR")
        raise


if __name__ == "__main__":
    raise SystemExit(main())