# -*- coding: utf-8 -*-
"""从 NAS 读取门店经营负责人快照，并维护快照表与 SCD2 历史表。"""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pymysql.cursors import DictCursor


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from db_connections import connect_mysql
from tools.nas_access import ensure_nas_path_access


DEFAULT_FILE_PATH = Path(r"\\192.168.0.151\hefang总部\14-数据中台\销售部\目标配置表\门店负责人映射表.xlsx")
DEFAULT_CREATED_BY = "store_operation_owner_import"
DEFAULT_SHEET_ALIASES = ("门店负责人映射表", "门店负责人映射模板")

SNAPSHOT_TABLE_NAME = "cfg_store_operation_owner_snapshot"
HISTORY_TABLE_NAME = "dim_store_operation_owner_assignment"
LOG_TABLE_NAME = "log_store_operation_owner_import"
STORE_ATTR_TABLE_NAME = "dim_store_report_attr"
TARGET_TABLE_NAME = "cfg_store_target_daily"
ASSESSMENT_ASSIGNMENT_TABLE_NAME = "cfg_store_assessment_assignment"
ASSESSMENT_SUBJECT_TARGET_TABLE_NAME = "cfg_store_assessment_subject_target_daily"
DDL_FILE_PATH = REPO_ROOT / "SQL" / "create_store_operation_owner_tables.sql"

REQUIRED_HEADERS = ("门店编码", "门店名称", "负责人")
OPTIONAL_HEADERS = {
    "备注": "备注",
    "序号": "序号",
    "生效日期": "生效日期",
    "失效日期": "失效日期",
}
MAX_MESSAGE_LENGTH = 1000


@dataclass(frozen=True)
class SourceRow:
    row_number: int
    entity_type: str
    entity_code: str
    entity_name: str
    owner_name: str | None
    remark: str | None
    effective_start_date: date | None
    effective_end_date: date | None
    has_explicit_effective_start_date: bool = False
    has_explicit_effective_end_date: bool = False


@dataclass(frozen=True)
class ExpectedEntity:
    entity_type: str
    entity_id: int | None
    entity_code: str
    entity_name: str


@dataclass(frozen=True)
class SnapshotRow:
    snapshot_date: date
    entity_type: str
    entity_id: int | None
    entity_code: str
    entity_name: str
    owner_name: str | None
    remark: str | None
    effective_start_date: date
    effective_end_date: date
    has_explicit_effective_start_date: bool = False
    has_explicit_effective_end_date: bool = False


def _connect() -> Any:
    return connect_mysql(
        cursorclass=DictCursor,
        autocommit=False,
    )


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def _normalize_code(value: object) -> str:
    return _normalize_text(value).upper()


def _normalize_owner_name(value: object) -> str | None:
    normalized = _normalize_text(value)
    return normalized or None


def _parse_optional_excel_date(value: object, row_number: int, field_name: str) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    normalized = _normalize_text(value)
    if not normalized:
        return None

    for date_format in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(normalized, date_format).date()
        except ValueError:
            continue

    raise ValueError(f"第 {row_number} 行的 {field_name} 不是合法日期: {value}")


def _classify_entity_type(entity_code: str) -> str:
    return "SUBJECT" if entity_code.startswith("SUBJ_") else "STORE"


def _truncate_message(message: str) -> str:
    normalized = message.strip()
    if len(normalized) <= MAX_MESSAGE_LENGTH:
        return normalized
    return normalized[: MAX_MESSAGE_LENGTH - 3] + "..."


def _resolve_output_path(path_value: str) -> Path:
    output_path = Path(path_value)
    if not output_path.is_absolute():
        output_path = REPO_ROOT / output_path
    output_path.parent.mkdir(parents=True, exist_ok=True)
    return output_path


def _resolve_input_file(file_path_arg: str | None) -> Path:
    file_path = Path(file_path_arg) if file_path_arg else DEFAULT_FILE_PATH
    if not file_path.is_absolute():
        file_path = REPO_ROOT / file_path
    ensure_nas_path_access(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"未找到负责人映射文件: {file_path}")
    return file_path


def _resolve_sheet_name(workbook: Any, requested_sheet_name: str | None) -> str:
    if requested_sheet_name:
        if requested_sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"未找到工作表 {requested_sheet_name}，可选 sheet: {', '.join(workbook.sheetnames)}"
            )
        return requested_sheet_name

    for sheet_name in DEFAULT_SHEET_ALIASES:
        if sheet_name in workbook.sheetnames:
            return sheet_name

    raise ValueError(
        f"未找到负责人映射工作表，当前可选 sheet: {', '.join(workbook.sheetnames)}"
    )


def _build_header_map(header_row: tuple[object, ...]) -> dict[str, int]:
    header_map: dict[str, int] = {}
    for index, raw_value in enumerate(header_row):
        header_name = _normalize_text(raw_value)
        if header_name:
            header_map[header_name] = index

    missing_headers = [header for header in REQUIRED_HEADERS if header not in header_map]
    if missing_headers:
        raise ValueError(f"负责人映射模板缺少必填表头: {', '.join(missing_headers)}")
    return header_map


def _parse_workbook(file_path: Path, requested_sheet_name: str | None) -> tuple[list[SourceRow], dict[str, Any]]:
    ensure_nas_path_access(file_path)
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    try:
        sheet_name = _resolve_sheet_name(workbook, requested_sheet_name)
        worksheet = workbook[sheet_name]
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            raise ValueError("负责人映射模板为空，缺少表头")

        header_map = _build_header_map(header_row)
        source_rows: list[SourceRow] = []
        duplicate_keys: set[tuple[str, str]] = set()
        seen_keys: set[tuple[str, str]] = set()
        blank_owner_count = 0
        explicit_start_date_row_count = 0
        explicit_end_date_row_count = 0
        explicit_interval_row_count = 0
        date_headers_present = [
            header_name
            for header_name in (OPTIONAL_HEADERS["生效日期"], OPTIONAL_HEADERS["失效日期"])
            if header_name in header_map
        ]

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if row is None:
                continue
            if all(cell is None or _normalize_text(cell) == "" for cell in row):
                continue

            entity_code = _normalize_code(row[header_map["门店编码"]])
            entity_name = _normalize_text(row[header_map["门店名称"]])
            owner_name = _normalize_owner_name(row[header_map["负责人"]])
            remark = None
            if OPTIONAL_HEADERS["备注"] in header_map:
                remark = _normalize_owner_name(row[header_map[OPTIONAL_HEADERS["备注"]]])
            effective_start_date = None
            if OPTIONAL_HEADERS["生效日期"] in header_map:
                effective_start_date = _parse_optional_excel_date(
                    row[header_map[OPTIONAL_HEADERS["生效日期"]]],
                    row_number,
                    OPTIONAL_HEADERS["生效日期"],
                )
            effective_end_date = None
            if OPTIONAL_HEADERS["失效日期"] in header_map:
                effective_end_date = _parse_optional_excel_date(
                    row[header_map[OPTIONAL_HEADERS["失效日期"]]],
                    row_number,
                    OPTIONAL_HEADERS["失效日期"],
                )

            if not entity_code:
                raise ValueError(f"第 {row_number} 行缺少 门店编码")
            if not entity_name:
                raise ValueError(f"第 {row_number} 行缺少 门店名称")
            if owner_name is None:
                blank_owner_count += 1
            if effective_start_date is not None:
                explicit_start_date_row_count += 1
            if effective_end_date is not None:
                explicit_end_date_row_count += 1
            if effective_start_date is not None or effective_end_date is not None:
                explicit_interval_row_count += 1
            if (
                effective_start_date is not None
                and effective_end_date is not None
                and effective_start_date > effective_end_date
            ):
                raise ValueError(
                    f"第 {row_number} 行的 生效日期 晚于 失效日期: "
                    f"{effective_start_date.isoformat()} > {effective_end_date.isoformat()}"
                )

            entity_type = _classify_entity_type(entity_code)
            row_key = (entity_type, entity_code)
            if row_key in seen_keys:
                duplicate_keys.add(row_key)
            seen_keys.add(row_key)
            source_rows.append(
                SourceRow(
                    row_number=row_number,
                    entity_type=entity_type,
                    entity_code=entity_code,
                    entity_name=entity_name,
                    owner_name=owner_name,
                    remark=remark,
                    effective_start_date=effective_start_date,
                    effective_end_date=effective_end_date,
                    has_explicit_effective_start_date=effective_start_date is not None,
                    has_explicit_effective_end_date=effective_end_date is not None,
                )
            )

        if not source_rows:
            raise ValueError("负责人映射模板没有有效数据行")
        if duplicate_keys:
            duplicated = [f"{entity_type}:{entity_code}" for entity_type, entity_code in sorted(duplicate_keys)]
            raise ValueError(f"负责人映射模板存在重复实体编码: {duplicated}")

        return source_rows, {
            "source_sheet": sheet_name,
            "source_row_count": len(source_rows),
            "blank_owner_count": blank_owner_count,
            "date_headers_present": date_headers_present,
            "explicit_start_date_row_count": explicit_start_date_row_count,
            "explicit_end_date_row_count": explicit_end_date_row_count,
            "explicit_interval_row_count": explicit_interval_row_count,
            "sheet_names": list(workbook.sheetnames),
        }
    finally:
        workbook.close()


def _ensure_required_tables(conn: Any) -> None:
    required_tables = (
        STORE_ATTR_TABLE_NAME,
        TARGET_TABLE_NAME,
        ASSESSMENT_ASSIGNMENT_TABLE_NAME,
        ASSESSMENT_SUBJECT_TARGET_TABLE_NAME,
        SNAPSHOT_TABLE_NAME,
        HISTORY_TABLE_NAME,
        LOG_TABLE_NAME,
    )
    placeholders = ", ".join(["%s"] * len(required_tables))
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT table_name AS table_name_alias
            FROM information_schema.tables
            WHERE table_schema = DATABASE()
              AND table_name IN ({placeholders})
            """,
            required_tables,
        )
        existing_tables = {
            (row.get("table_name_alias") or row.get("TABLE_NAME_ALIAS"))
            for row in cursor.fetchall()
        }

    missing_tables = [table_name for table_name in required_tables if table_name not in existing_tables]
    if missing_tables:
        ddl_hint = DDL_FILE_PATH.relative_to(REPO_ROOT).as_posix()
        raise RuntimeError(
            f"缺少依赖表: {', '.join(missing_tables)}；请先执行 {ddl_hint}"
        )


def _load_active_store_rows(conn: Any, snapshot_date: date) -> tuple[dict[str, dict[str, Any]], list[dict[str, Any]]]:
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                store_id AS store_id_alias,
                store_code AS store_code_alias,
                store_name AS store_name_alias
            FROM {STORE_ATTR_TABLE_NAME}
            WHERE is_include_in_daily_report = 'Y'
              AND %s BETWEEN effective_start_date AND effective_end_date
                            AND EXISTS (
                                    SELECT 1
                                    FROM {TARGET_TABLE_NAME} t
                                    WHERE t.store_id = {STORE_ATTR_TABLE_NAME}.store_id
                                        AND t.target_date = %s
                            )
            ORDER BY store_code, effective_start_date
            """,
                        (snapshot_date, snapshot_date),
        )
        rows = [dict(row) for row in cursor.fetchall()]

    grouped_rows: dict[str, list[dict[str, Any]]] = {}
    for row in rows:
        store_code = _normalize_code(row["store_code_alias"])
        grouped_rows.setdefault(store_code, []).append(
            {
                "store_id": int(row["store_id_alias"]),
                "store_code": store_code,
                "store_name": row["store_name_alias"],
            }
        )

    duplicate_rows: list[dict[str, Any]] = []
    store_map: dict[str, dict[str, Any]] = {}
    for store_code, matched_rows in grouped_rows.items():
        if len(matched_rows) > 1:
            duplicate_rows.append(
                {
                    "store_code": store_code,
                    "store_ids": [row["store_id"] for row in matched_rows],
                }
            )
            continue
        store_map[store_code] = matched_rows[0]
    return store_map, duplicate_rows


def _load_active_joint_assignment_rows(conn: Any, snapshot_date: date) -> list[dict[str, Any]]:
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                store_id AS store_id_alias,
                store_code AS store_code_alias,
                store_name AS store_name_alias,
                subject_code AS subject_code_alias,
                anchor_store_id AS anchor_store_id_alias,
                anchor_store_name AS anchor_store_name_alias
            FROM {ASSESSMENT_ASSIGNMENT_TABLE_NAME}
            WHERE is_joint_assessment = 'Y'
              AND %s BETWEEN effective_start_date AND effective_end_date
            ORDER BY subject_code, store_code
            """,
            (snapshot_date,),
        )
        return [dict(row) for row in cursor.fetchall()]


def _load_month_joint_assignment_rows(conn: Any, snapshot_date: date) -> list[dict[str, Any]]:
    target_month = snapshot_date.replace(day=1)
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                store_code AS store_code_alias,
                subject_code AS subject_code_alias
            FROM {ASSESSMENT_ASSIGNMENT_TABLE_NAME}
            WHERE is_joint_assessment = 'Y'
              AND target_month = %s
            ORDER BY subject_code, store_code
            """,
            (target_month,),
        )
        return [dict(row) for row in cursor.fetchall()]


def _resolve_joint_assignment_maps(
    assignment_rows: list[dict[str, Any]],
    active_store_codes: set[str],
) -> tuple[dict[str, str], dict[str, int | None], list[dict[str, Any]], list[dict[str, Any]]]:
    store_subject_candidates: dict[str, set[str]] = {}
    subject_anchor_candidates: dict[str, set[int | None]] = {}

    for row in assignment_rows:
        store_code = _normalize_code(row["store_code_alias"])
        if store_code not in active_store_codes:
            continue
        subject_code = _normalize_code(row["subject_code_alias"])
        store_subject_candidates.setdefault(store_code, set()).add(subject_code)
        subject_anchor_candidates.setdefault(subject_code, set()).add(row["anchor_store_id_alias"])

    ambiguous_store_subject_rows = [
        {"store_code": store_code, "subject_codes": sorted(subject_codes)}
        for store_code, subject_codes in store_subject_candidates.items()
        if len(subject_codes) > 1
    ]
    ambiguous_subject_anchor_rows = [
        {"subject_code": subject_code, "anchor_store_ids": sorted(anchor_store_ids)}
        for subject_code, anchor_store_ids in subject_anchor_candidates.items()
        if len(anchor_store_ids) > 1
    ]

    store_subject_map = {
        store_code: next(iter(subject_codes))
        for store_code, subject_codes in store_subject_candidates.items()
        if len(subject_codes) == 1
    }
    subject_anchor_map = {
        subject_code: next(iter(anchor_store_ids))
        for subject_code, anchor_store_ids in subject_anchor_candidates.items()
        if len(anchor_store_ids) == 1
    }
    return (
        store_subject_map,
        subject_anchor_map,
        ambiguous_store_subject_rows,
        ambiguous_subject_anchor_rows,
    )


def _build_month_joint_transition_maps(
    assignment_rows: list[dict[str, Any]],
) -> tuple[dict[str, str], dict[str, set[str]]]:
    store_subject_candidates: dict[str, set[str]] = {}
    for row in assignment_rows:
        store_code = _normalize_code(row["store_code_alias"])
        subject_code = _normalize_code(row["subject_code_alias"])
        if not store_code or not subject_code:
            continue
        store_subject_candidates.setdefault(store_code, set()).add(subject_code)

    store_subject_map = {
        store_code: next(iter(subject_codes))
        for store_code, subject_codes in store_subject_candidates.items()
        if len(subject_codes) == 1
    }
    subject_store_map: dict[str, set[str]] = {}
    for store_code, subject_code in store_subject_map.items():
        subject_store_map.setdefault(subject_code, set()).add(store_code)
    return store_subject_map, subject_store_map


def _load_subject_name_map(
    conn: Any,
    snapshot_date: date,
) -> tuple[dict[str, str], list[dict[str, Any]]]:
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                subject_code AS subject_code_alias,
                subject_name AS subject_name_alias
            FROM {ASSESSMENT_SUBJECT_TARGET_TABLE_NAME}
            WHERE target_date = %s
            ORDER BY subject_code
            """,
            (snapshot_date,),
        )
        rows = [dict(row) for row in cursor.fetchall()]

    grouped_rows: dict[str, set[str]] = {}
    for row in rows:
        subject_code = _normalize_code(row["subject_code_alias"])
        subject_name = _normalize_text(row["subject_name_alias"])
        if not subject_name:
            continue
        grouped_rows.setdefault(subject_code, set()).add(subject_name)

    ambiguous_subject_name_rows = [
        {"subject_code": subject_code, "subject_names": sorted(subject_names)}
        for subject_code, subject_names in grouped_rows.items()
        if len(subject_names) > 1
    ]
    subject_name_map = {
        subject_code: next(iter(subject_names))
        for subject_code, subject_names in grouped_rows.items()
        if len(subject_names) == 1
    }
    return subject_name_map, ambiguous_subject_name_rows


def _build_expected_entities(
    store_map: dict[str, dict[str, Any]],
    store_subject_map: dict[str, str],
    subject_anchor_map: dict[str, int | None],
    subject_name_map: dict[str, str],
) -> dict[tuple[str, str], ExpectedEntity]:
    expected_entities: dict[tuple[str, str], ExpectedEntity] = {}
    for store_code, store_row in sorted(store_map.items()):
        if store_code in store_subject_map:
            subject_code = store_subject_map[store_code]
            entity_key = ("SUBJECT", subject_code)
            if entity_key not in expected_entities:
                expected_entities[entity_key] = ExpectedEntity(
                    entity_type="SUBJECT",
                    entity_id=subject_anchor_map.get(subject_code),
                    entity_code=subject_code,
                    entity_name=subject_name_map.get(subject_code, subject_code),
                )
            continue

        entity_key = ("STORE", store_code)
        expected_entities[entity_key] = ExpectedEntity(
            entity_type="STORE",
            entity_id=store_row["store_id"],
            entity_code=store_code,
            entity_name=store_row["store_name"],
        )
    return expected_entities


def _serialize_entity_key(entity_type: str, entity_code: str) -> dict[str, str]:
    return {"entity_type": entity_type, "entity_code": entity_code}


def _resolve_source_row_effective_interval(
    source_row: SourceRow,
    snapshot_date: date,
) -> tuple[date, date, bool]:
    effective_start_date = source_row.effective_start_date or snapshot_date
    effective_end_date = source_row.effective_end_date or date(9999, 12, 31)
    covers_snapshot_date = effective_start_date <= snapshot_date <= effective_end_date
    return effective_start_date, effective_end_date, covers_snapshot_date


def _build_snapshot_rows(
    source_rows: list[SourceRow],
    expected_entities: dict[tuple[str, str], ExpectedEntity],
    snapshot_date: date,
    month_store_subject_map: dict[str, str],
    month_subject_store_map: dict[str, set[str]],
) -> tuple[list[SnapshotRow], dict[str, Any]]:
    source_map = {
        (row.entity_type, row.entity_code): row
        for row in source_rows
    }
    expected_keys = set(expected_entities)
    active_source_map: dict[tuple[str, str], SourceRow] = {}
    invalid_effective_entity_keys: set[tuple[str, str]] = set()

    for entity_key, source_row in source_map.items():
        effective_start_date, effective_end_date, covers_snapshot_date = _resolve_source_row_effective_interval(
            source_row,
            snapshot_date,
        )
        if covers_snapshot_date:
            active_source_map[entity_key] = source_row
            continue

        if entity_key in expected_entities:
            invalid_effective_entity_keys.add(entity_key)

    active_source_keys = set(active_source_map)

    tolerated_transition_entities: list[dict[str, Any]] = []
    tolerated_transition_keys: set[tuple[str, str]] = set()
    for entity_key in sorted(active_source_keys - expected_keys):
        source_row = active_source_map[entity_key]
        entity_type, entity_code = entity_key
        if entity_type == "STORE":
            subject_code = month_store_subject_map.get(entity_code)
            if subject_code and ("SUBJECT", subject_code) in active_source_keys:
                tolerated_transition_entities.append(
                    {
                        **_serialize_entity_key(entity_type, entity_code),
                        "row_number": source_row.row_number,
                        "entity_name": source_row.entity_name,
                        "reason": "store_absorbed_by_subject",
                        "related_subject_code": subject_code,
                    }
                )
                tolerated_transition_keys.add(entity_key)
                continue

        if entity_type == "SUBJECT":
            related_store_codes = sorted(
                store_code
                for store_code in month_subject_store_map.get(entity_code, set())
                if ("STORE", store_code) in active_source_keys
            )
            if related_store_codes:
                tolerated_transition_entities.append(
                    {
                        **_serialize_entity_key(entity_type, entity_code),
                        "row_number": source_row.row_number,
                        "entity_name": source_row.entity_name,
                        "reason": "subject_prepared_before_effective_date",
                        "related_store_codes": related_store_codes,
                    }
                )
                tolerated_transition_keys.add(entity_key)

    missing_entities = [
        {
            **_serialize_entity_key(entity_type, entity_code),
            "entity_name": expected_entities[(entity_type, entity_code)].entity_name,
        }
        for entity_type, entity_code in sorted(expected_keys - active_source_keys - invalid_effective_entity_keys)
    ]
    unexpected_entities = [
        {
            **_serialize_entity_key(entity_type, entity_code),
            "row_number": active_source_map[(entity_type, entity_code)].row_number,
            "entity_name": active_source_map[(entity_type, entity_code)].entity_name,
        }
        for entity_type, entity_code in sorted(active_source_keys - expected_keys - tolerated_transition_keys)
    ]

    entity_name_mismatch_rows: list[dict[str, Any]] = []
    invalid_effective_date_rows: list[dict[str, Any]] = []
    snapshot_rows: list[SnapshotRow] = []
    for entity_key in sorted(expected_keys & set(source_map)):
        expected_entity = expected_entities[entity_key]
        source_row = source_map[entity_key]
        effective_start_date, effective_end_date, covers_snapshot_date = _resolve_source_row_effective_interval(
            source_row,
            snapshot_date,
        )
        if not covers_snapshot_date:
            invalid_effective_date_rows.append(
                {
                    **_serialize_entity_key(expected_entity.entity_type, expected_entity.entity_code),
                    "row_number": source_row.row_number,
                    "snapshot_date": snapshot_date.isoformat(),
                    "effective_start_date": effective_start_date.isoformat(),
                    "effective_end_date": effective_end_date.isoformat(),
                }
            )
            continue

        if _normalize_text(source_row.entity_name) != _normalize_text(expected_entity.entity_name):
            entity_name_mismatch_rows.append(
                {
                    **_serialize_entity_key(expected_entity.entity_type, expected_entity.entity_code),
                    "row_number": source_row.row_number,
                    "expected_entity_name": expected_entity.entity_name,
                    "source_entity_name": source_row.entity_name,
                }
            )
            continue

        snapshot_rows.append(
            SnapshotRow(
                snapshot_date=snapshot_date,
                entity_type=expected_entity.entity_type,
                entity_id=expected_entity.entity_id,
                entity_code=expected_entity.entity_code,
                entity_name=expected_entity.entity_name,
                owner_name=source_row.owner_name,
                remark=source_row.remark,
                effective_start_date=effective_start_date,
                effective_end_date=effective_end_date,
                has_explicit_effective_start_date=source_row.has_explicit_effective_start_date,
                has_explicit_effective_end_date=source_row.has_explicit_effective_end_date,
            )
        )

    return snapshot_rows, {
        "missing_entities": missing_entities,
        "unexpected_entities": unexpected_entities,
        "tolerated_transition_entities": tolerated_transition_entities,
        "entity_name_mismatch_rows": entity_name_mismatch_rows,
        "invalid_effective_date_rows": invalid_effective_date_rows,
        "matched_entity_count": len(snapshot_rows),
    }


def _serialize_history_row(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "entity_type": row["entity_type"],
        "entity_id": row["entity_id"],
        "entity_code": row["entity_code"],
        "entity_name": row["entity_name"],
        "owner_name": row["owner_name"],
        "effective_start_date": row["effective_start_date"].isoformat(),
        "effective_end_date": row["effective_end_date"].isoformat(),
        "is_current": row["is_current"],
    }


def _serialize_snapshot_row(row: SnapshotRow) -> dict[str, Any]:
    return {
        "snapshot_date": row.snapshot_date.isoformat(),
        "entity_type": row.entity_type,
        "entity_id": row.entity_id,
        "entity_code": row.entity_code,
        "entity_name": row.entity_name,
        "owner_name": row.owner_name,
        "remark": row.remark,
        "effective_start_date": row.effective_start_date.isoformat(),
        "effective_end_date": row.effective_end_date.isoformat(),
    }


def _load_current_history_rows(
    conn: Any,
    snapshot_date: date,
) -> tuple[list[dict[str, Any]], dict[tuple[str, str], dict[str, Any]], list[dict[str, Any]]]:
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            SELECT
                entity_type AS entity_type_alias,
                entity_id AS entity_id_alias,
                entity_code AS entity_code_alias,
                entity_name AS entity_name_alias,
                owner_name AS owner_name_alias,
                effective_start_date AS effective_start_date_alias,
                effective_end_date AS effective_end_date_alias,
                is_current AS is_current_alias
            FROM {HISTORY_TABLE_NAME}
            WHERE %s BETWEEN effective_start_date AND effective_end_date
            ORDER BY entity_type, entity_code, effective_start_date
            """,
            (snapshot_date,),
        )
        rows = [dict(row) for row in cursor.fetchall()]

    grouped_rows: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for row in rows:
        current_row = {
            "entity_type": row["entity_type_alias"],
            "entity_id": row["entity_id_alias"],
            "entity_code": row["entity_code_alias"],
            "entity_name": row["entity_name_alias"],
            "owner_name": _normalize_owner_name(row["owner_name_alias"]),
            "effective_start_date": row["effective_start_date_alias"],
            "effective_end_date": row["effective_end_date_alias"],
            "is_current": row["is_current_alias"],
        }
        grouped_rows.setdefault((current_row["entity_type"], current_row["entity_code"]), []).append(current_row)

    current_rows: list[dict[str, Any]] = []
    current_map: dict[tuple[str, str], dict[str, Any]] = {}
    overlap_rows: list[dict[str, Any]] = []
    for entity_key, matched_rows in grouped_rows.items():
        if len(matched_rows) > 1:
            overlap_rows.extend(_serialize_history_row(row) for row in matched_rows)
            continue
        current_rows.append(matched_rows[0])
        current_map[entity_key] = matched_rows[0]
    return current_rows, current_map, overlap_rows


def _snapshot_row_equals_history_row(snapshot_row: SnapshotRow, history_row: dict[str, Any]) -> bool:
    if (
        snapshot_row.entity_id != history_row.get("entity_id")
        or snapshot_row.entity_name != history_row.get("entity_name")
        or snapshot_row.owner_name != history_row.get("owner_name")
    ):
        return False

    if snapshot_row.has_explicit_effective_start_date and (
        snapshot_row.effective_start_date != history_row.get("effective_start_date")
    ):
        return False

    if snapshot_row.has_explicit_effective_end_date and (
        snapshot_row.effective_end_date != history_row.get("effective_end_date")
    ):
        return False

    return True


def _snapshot_row_matches_history_payload(snapshot_row: SnapshotRow, history_row: dict[str, Any]) -> bool:
    return (
        snapshot_row.entity_id == history_row.get("entity_id")
        and snapshot_row.entity_name == history_row.get("entity_name")
        and snapshot_row.owner_name == history_row.get("owner_name")
    )


def _classify_history_changes(
    snapshot_rows: list[SnapshotRow],
    current_map: dict[tuple[str, str], dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    snapshot_map = {(row.entity_type, row.entity_code): row for row in snapshot_rows}

    unchanged_rows: list[dict[str, Any]] = []
    changed_rows: list[dict[str, Any]] = []
    new_rows: list[dict[str, Any]] = []
    exited_rows: list[dict[str, Any]] = []

    for entity_key in sorted(set(snapshot_map) | set(current_map)):
        snapshot_row = snapshot_map.get(entity_key)
        current_row = current_map.get(entity_key)
        if snapshot_row is not None and current_row is not None:
            payload = {
                "entity_type": entity_key[0],
                "entity_code": entity_key[1],
                "current": current_row,
                "snapshot": snapshot_row,
            }
            if _snapshot_row_equals_history_row(snapshot_row, current_row):
                unchanged_rows.append(payload)
            else:
                changed_rows.append(payload)
            continue

        if snapshot_row is not None:
            new_rows.append(
                {
                    "entity_type": entity_key[0],
                    "entity_code": entity_key[1],
                    "snapshot": snapshot_row,
                }
            )
            continue

        if current_row is not None:
            exited_rows.append(
                {
                    "entity_type": entity_key[0],
                    "entity_code": entity_key[1],
                    "current": current_row,
                }
            )

    return {
        "unchanged_rows": unchanged_rows,
        "changed_rows": changed_rows,
        "new_rows": new_rows,
        "exited_rows": exited_rows,
    }


def _build_summary(
    file_path: Path,
    file_md5: str,
    snapshot_date: date,
    workbook_summary: dict[str, Any],
    source_rows: list[SourceRow],
    active_store_rows: dict[str, dict[str, Any]],
    duplicate_active_store_rows: list[dict[str, Any]],
    store_subject_map: dict[str, str],
    ambiguous_store_subject_rows: list[dict[str, Any]],
    ambiguous_subject_anchor_rows: list[dict[str, Any]],
    ambiguous_subject_name_rows: list[dict[str, Any]],
    expected_entities: dict[tuple[str, str], ExpectedEntity],
    snapshot_rows: list[SnapshotRow],
    snapshot_validation: dict[str, Any],
    current_history_rows: list[dict[str, Any]],
    history_overlap_rows: list[dict[str, Any]],
    history_changes: dict[str, list[dict[str, Any]]],
    preview_limit: int,
) -> dict[str, Any]:
    preview_snapshot_rows = [_serialize_snapshot_row(row) for row in snapshot_rows[:preview_limit]]
    preview_changed_rows = [
        {
            "entity_type": row["entity_type"],
            "entity_code": row["entity_code"],
            "current": _serialize_history_row(row["current"]),
            "snapshot": _serialize_snapshot_row(row["snapshot"]),
        }
        for row in history_changes["changed_rows"][:preview_limit]
    ]
    preview_new_rows = [
        {
            "entity_type": row["entity_type"],
            "entity_code": row["entity_code"],
            "snapshot": _serialize_snapshot_row(row["snapshot"]),
        }
        for row in history_changes["new_rows"][:preview_limit]
    ]
    preview_exited_rows = [
        {
            "entity_type": row["entity_type"],
            "entity_code": row["entity_code"],
            "current": _serialize_history_row(row["current"]),
        }
        for row in history_changes["exited_rows"][:preview_limit]
    ]
    affected_start_dates = [
        row["snapshot"].effective_start_date
        for row in history_changes["changed_rows"]
    ]
    affected_start_dates.extend(
        row["snapshot"].effective_start_date
        for row in history_changes["new_rows"]
    )
    affected_start_dates.extend([snapshot_date] * len(history_changes["exited_rows"]))
    earliest_history_effective_start_date = min(affected_start_dates).isoformat() if affected_start_dates else None

    return {
        "mode": "dry-run",
        "file_name": file_path.name,
        "file_path": str(file_path),
        "file_md5": file_md5,
        "source_sheet": workbook_summary["source_sheet"],
        "sheet_names": workbook_summary["sheet_names"],
        "snapshot_date": snapshot_date,
        "source_row_count": len(source_rows),
        "blank_owner_count": workbook_summary["blank_owner_count"],
        "date_headers_present": workbook_summary["date_headers_present"],
        "explicit_start_date_row_count": workbook_summary["explicit_start_date_row_count"],
        "explicit_end_date_row_count": workbook_summary["explicit_end_date_row_count"],
        "explicit_interval_row_count": workbook_summary["explicit_interval_row_count"],
        "active_store_count": len(active_store_rows),
        "joint_assessment_store_count": len(store_subject_map),
        "expected_entity_count": len(expected_entities),
        "matched_entity_count": snapshot_validation["matched_entity_count"],
        "missing_entity_count": len(snapshot_validation["missing_entities"]),
        "unexpected_entity_count": len(snapshot_validation["unexpected_entities"]),
        "tolerated_transition_entity_count": len(snapshot_validation["tolerated_transition_entities"]),
        "entity_name_mismatch_count": len(snapshot_validation["entity_name_mismatch_rows"]),
        "invalid_effective_date_count": len(snapshot_validation["invalid_effective_date_rows"]),
        "duplicate_active_store_rows": duplicate_active_store_rows,
        "ambiguous_store_subject_rows": ambiguous_store_subject_rows,
        "ambiguous_subject_anchor_rows": ambiguous_subject_anchor_rows,
        "ambiguous_subject_name_rows": ambiguous_subject_name_rows,
        "missing_entities": snapshot_validation["missing_entities"],
        "unexpected_entities": snapshot_validation["unexpected_entities"],
        "tolerated_transition_entities": snapshot_validation["tolerated_transition_entities"],
        "entity_name_mismatch_rows": snapshot_validation["entity_name_mismatch_rows"],
        "invalid_effective_date_rows": snapshot_validation["invalid_effective_date_rows"],
        "current_history_row_count": len(current_history_rows),
        "history_overlap_rows": history_overlap_rows,
        "history_diff_counts": {
            "unchanged": len(history_changes["unchanged_rows"]),
            "changed": len(history_changes["changed_rows"]),
            "new": len(history_changes["new_rows"]),
            "exited": len(history_changes["exited_rows"]),
        },
        "earliest_history_effective_start_date": earliest_history_effective_start_date,
        "snapshot_rows_inserted": 0,
        "history_rows_opened": 0,
        "history_rows_closed": 0,
        "preview_snapshot_rows": preview_snapshot_rows,
        "preview_changed_rows": preview_changed_rows,
        "preview_new_rows": preview_new_rows,
        "preview_exited_rows": preview_exited_rows,
        "started_at": datetime.now(),
    }


def _build_validation_error_message(summary: dict[str, Any]) -> str:
    if summary["duplicate_active_store_rows"]:
        return f"{STORE_ATTR_TABLE_NAME} 在快照日期存在重复有效 store_code: {summary['duplicate_active_store_rows']}"
    if summary["ambiguous_store_subject_rows"]:
        return (
            f"{ASSESSMENT_ASSIGNMENT_TABLE_NAME} 中同一门店命中了多个共同考核主体，无法确定经营体: "
            f"{summary['ambiguous_store_subject_rows']}"
        )
    if summary["ambiguous_subject_anchor_rows"]:
        return (
            f"{ASSESSMENT_ASSIGNMENT_TABLE_NAME} 中同一主体命中了多个挂靠主店，无法确定经营体ID: "
            f"{summary['ambiguous_subject_anchor_rows']}"
        )
    if summary["ambiguous_subject_name_rows"]:
        return (
            f"{ASSESSMENT_SUBJECT_TARGET_TABLE_NAME} 中同一主体存在多个名称，无法校验负责人快照名称: "
            f"{summary['ambiguous_subject_name_rows']}"
        )
    if summary["missing_entities"]:
        return f"负责人快照缺少当前应维护的经营实体: {summary['missing_entities']}"
    if summary["unexpected_entities"]:
        return f"负责人快照出现不应维护的实体编码: {summary['unexpected_entities']}"
    if summary["entity_name_mismatch_rows"]:
        return f"负责人快照中的实体名称与当前经营实体清单不一致: {summary['entity_name_mismatch_rows']}"
    if summary["invalid_effective_date_rows"]:
        return (
            "负责人快照中的显式生效区间未覆盖当前 snapshot_date，"
            f"请检查 生效日期/失效日期: {summary['invalid_effective_date_rows']}"
        )
    if summary["history_overlap_rows"]:
        return f"{HISTORY_TABLE_NAME} 在快照日期存在多个当前有效版本，需先清理历史重叠: {summary['history_overlap_rows']}"
    return ""


def _build_validation_warning_messages(summary: dict[str, Any]) -> list[str]:
    warning_messages: list[str] = []
    tolerated_entities = summary.get("tolerated_transition_entities") or []
    if tolerated_entities:
        preview_items: list[str] = []
        for row in tolerated_entities[:10]:
            if row.get("reason") == "store_absorbed_by_subject":
                preview_items.append(
                    f"row={row.get('row_number')} {row.get('entity_code')} -> {row.get('related_subject_code')}"
                )
            else:
                related_store_codes = row.get("related_store_codes") or []
                preview_items.append(
                    f"row={row.get('row_number')} {row.get('entity_code')} with stores={related_store_codes[:3]}"
                )
        remaining = len(tolerated_entities) - len(preview_items)
        if remaining > 0:
            preview_items.append(f"...其余{remaining}项")
        warning_messages.append(
            "负责人快照存在共同考核过渡期并存实体，本次仅提示不阻断导入："
            f"{preview_items}"
        )
    return warning_messages


def _to_printable_summary(summary: dict[str, Any]) -> dict[str, Any]:
    printable = dict(summary)
    snapshot_date = printable.get("snapshot_date")
    if isinstance(snapshot_date, date):
        printable["snapshot_date"] = snapshot_date.isoformat()
    started_at = printable.get("started_at")
    if isinstance(started_at, datetime):
        printable["started_at"] = started_at.isoformat(timespec="seconds")
    return printable


def _write_output_json(output_json: str | None, payload: dict[str, Any]) -> None:
    if not output_json:
        return
    output_path = _resolve_output_path(output_json)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_import_log(
    summary: dict[str, Any],
    status: str,
    message: str,
) -> None:
    conn = _connect()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                f"""
                INSERT INTO {LOG_TABLE_NAME} (
                    file_name,
                    file_path,
                    file_md5,
                    source_sheet,
                    snapshot_date,
                    records_total,
                    expected_entity_count,
                    matched_entity_count,
                    missing_entity_count,
                    unexpected_entity_count,
                    snapshot_rows_inserted,
                    history_rows_opened,
                    history_rows_closed,
                    status,
                    message,
                    started_at,
                    finished_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """,
                (
                    summary["file_name"],
                    summary["file_path"],
                    summary["file_md5"],
                    summary["source_sheet"],
                    summary["snapshot_date"],
                    summary["source_row_count"],
                    summary["expected_entity_count"],
                    summary["matched_entity_count"],
                    summary["missing_entity_count"],
                    summary["unexpected_entity_count"],
                    summary["snapshot_rows_inserted"],
                    summary["history_rows_opened"],
                    summary["history_rows_closed"],
                    status,
                    _truncate_message(message),
                    summary["started_at"],
                    datetime.now(),
                ),
            )
        conn.commit()
    finally:
        conn.close()


def _close_or_delete_history_row(
    cursor: Any,
    current_row: dict[str, Any],
    transition_start_date: date,
    updated_by: str,
) -> None:
    if current_row["effective_start_date"] >= transition_start_date:
        cursor.execute(
            f"""
            DELETE FROM {HISTORY_TABLE_NAME}
            WHERE entity_type = %s
              AND entity_code = %s
              AND effective_start_date = %s
              AND effective_end_date = %s
            """,
            (
                current_row["entity_type"],
                current_row["entity_code"],
                current_row["effective_start_date"],
                current_row["effective_end_date"],
            ),
        )
        return

    cursor.execute(
        f"""
        UPDATE {HISTORY_TABLE_NAME}
        SET effective_end_date = %s,
            is_current = 'N',
            updated_by = %s
        WHERE entity_type = %s
          AND entity_code = %s
          AND effective_start_date = %s
          AND effective_end_date = %s
        """,
        (
            transition_start_date - timedelta(days=1),
            updated_by,
            current_row["entity_type"],
            current_row["entity_code"],
            current_row["effective_start_date"],
            current_row["effective_end_date"],
        ),
    )


def _fetch_previous_history_row(
    cursor: Any,
    entity_type: str,
    entity_code: str,
    transition_start_date: date,
) -> dict[str, Any] | None:
    cursor.execute(
        f"""
        SELECT
            entity_type,
            entity_id,
            entity_code,
            entity_name,
            owner_name,
            effective_start_date,
            effective_end_date,
            is_current
        FROM {HISTORY_TABLE_NAME}
        WHERE entity_type = %s
          AND entity_code = %s
          AND effective_end_date = %s
        ORDER BY effective_start_date DESC
        LIMIT 1
        """,
        (entity_type, entity_code, transition_start_date - timedelta(days=1)),
    )
    row = cursor.fetchone()
    if row is None:
        return None
    return {
        "entity_type": row["entity_type"],
        "entity_id": row["entity_id"],
        "entity_code": row["entity_code"],
        "entity_name": row["entity_name"],
        "owner_name": _normalize_owner_name(row["owner_name"]),
        "effective_start_date": row["effective_start_date"],
        "effective_end_date": row["effective_end_date"],
        "is_current": row["is_current"],
    }


def _reopen_or_insert_history_row(
    cursor: Any,
    snapshot_row: SnapshotRow,
    summary: dict[str, Any],
    created_by: str,
) -> None:
    previous_row = _fetch_previous_history_row(
        cursor,
        snapshot_row.entity_type,
        snapshot_row.entity_code,
        snapshot_row.effective_start_date,
    )
    if previous_row is not None and _snapshot_row_matches_history_payload(snapshot_row, previous_row):
        cursor.execute(
            f"""
            UPDATE {HISTORY_TABLE_NAME}
            SET entity_id = %s,
                entity_name = %s,
                owner_name = %s,
                source_snapshot_date = %s,
                source_file_name = %s,
                source_file_md5 = %s,
                effective_end_date = %s,
                is_current = 'Y',
                updated_by = %s
            WHERE entity_type = %s
              AND entity_code = %s
              AND effective_start_date = %s
            """,
            (
                snapshot_row.entity_id,
                snapshot_row.entity_name,
                snapshot_row.owner_name,
                snapshot_row.snapshot_date,
                summary["file_name"],
                summary["file_md5"],
                snapshot_row.effective_end_date,
                created_by,
                snapshot_row.entity_type,
                snapshot_row.entity_code,
                previous_row["effective_start_date"],
            ),
        )
        return

    cursor.execute(
        f"""
        INSERT INTO {HISTORY_TABLE_NAME} (
            entity_type,
            entity_id,
            entity_code,
            entity_name,
            owner_name,
            source_snapshot_date,
            source_file_name,
            source_file_md5,
            effective_start_date,
            effective_end_date,
            is_current,
            created_by,
            updated_by
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'Y', %s, %s)
        """,
        (
            snapshot_row.entity_type,
            snapshot_row.entity_id,
            snapshot_row.entity_code,
            snapshot_row.entity_name,
            snapshot_row.owner_name,
            snapshot_row.snapshot_date,
            summary["file_name"],
            summary["file_md5"],
            snapshot_row.effective_start_date,
            snapshot_row.effective_end_date,
            created_by,
            created_by,
        ),
    )


def _apply_import(
    summary: dict[str, Any],
    snapshot_rows: list[SnapshotRow],
    history_changes: dict[str, list[dict[str, Any]]],
    created_by: str,
) -> dict[str, Any]:
    conn = _connect()
    try:
        with conn.cursor() as cursor:
            cursor.execute(
                f"DELETE FROM {SNAPSHOT_TABLE_NAME} WHERE snapshot_date = %s",
                (summary["snapshot_date"],),
            )
            if snapshot_rows:
                cursor.executemany(
                    f"""
                    INSERT INTO {SNAPSHOT_TABLE_NAME} (
                        snapshot_date,
                        entity_type,
                        entity_id,
                        entity_code,
                        entity_name,
                        owner_name,
                        remark,
                        source_file_name,
                        source_file_md5,
                        created_by,
                        updated_by
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    [
                        (
                            row.snapshot_date,
                            row.entity_type,
                            row.entity_id,
                            row.entity_code,
                            row.entity_name,
                            row.owner_name,
                            row.remark,
                            summary["file_name"],
                            summary["file_md5"],
                            created_by,
                            created_by,
                        )
                        for row in snapshot_rows
                    ],
                )
                summary["snapshot_rows_inserted"] = cursor.rowcount

            history_rows_closed = 0
            for row in history_changes["changed_rows"]:
                _close_or_delete_history_row(
                    cursor,
                    row["current"],
                    row["snapshot"].effective_start_date,
                    created_by,
                )
                history_rows_closed += 1
            for row in history_changes["exited_rows"]:
                _close_or_delete_history_row(cursor, row["current"], summary["snapshot_date"], created_by)
                history_rows_closed += 1

            history_rows_opened = 0
            for row in history_changes["changed_rows"]:
                _reopen_or_insert_history_row(cursor, row["snapshot"], summary, created_by)
                history_rows_opened += 1
            for row in history_changes["new_rows"]:
                _reopen_or_insert_history_row(cursor, row["snapshot"], summary, created_by)
                history_rows_opened += 1

            summary["history_rows_opened"] = history_rows_opened
            summary["history_rows_closed"] = history_rows_closed
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    return summary


def _run_import(
    file_path: Path,
    requested_sheet_name: str | None,
    snapshot_date: date,
    preview_limit: int,
) -> tuple[dict[str, Any], list[SnapshotRow], dict[str, list[dict[str, Any]]], str | None]:
    source_rows, workbook_summary = _parse_workbook(file_path, requested_sheet_name)

    conn = _connect()
    try:
        _ensure_required_tables(conn)
        active_store_rows, duplicate_active_store_rows = _load_active_store_rows(conn, snapshot_date)
        assignment_rows = _load_active_joint_assignment_rows(conn, snapshot_date)
        month_assignment_rows = _load_month_joint_assignment_rows(conn, snapshot_date)
        (
            store_subject_map,
            subject_anchor_map,
            ambiguous_store_subject_rows,
            ambiguous_subject_anchor_rows,
        ) = _resolve_joint_assignment_maps(assignment_rows, set(active_store_rows))
        month_store_subject_map, month_subject_store_map = _build_month_joint_transition_maps(month_assignment_rows)
        subject_name_map, ambiguous_subject_name_rows = _load_subject_name_map(conn, snapshot_date)
        expected_entities = _build_expected_entities(
            active_store_rows,
            store_subject_map,
            subject_anchor_map,
            subject_name_map,
        )
        snapshot_rows, snapshot_validation = _build_snapshot_rows(
            source_rows,
            expected_entities,
            snapshot_date,
            month_store_subject_map,
            month_subject_store_map,
        )
        current_history_rows, current_history_map, history_overlap_rows = _load_current_history_rows(
            conn,
            snapshot_date,
        )
    finally:
        conn.close()

    history_changes = _classify_history_changes(snapshot_rows, current_history_map)
    summary = _build_summary(
        file_path=file_path,
        file_md5=_compute_file_md5(file_path),
        snapshot_date=snapshot_date,
        workbook_summary=workbook_summary,
        source_rows=source_rows,
        active_store_rows=active_store_rows,
        duplicate_active_store_rows=duplicate_active_store_rows,
        store_subject_map=store_subject_map,
        ambiguous_store_subject_rows=ambiguous_store_subject_rows,
        ambiguous_subject_anchor_rows=ambiguous_subject_anchor_rows,
        ambiguous_subject_name_rows=ambiguous_subject_name_rows,
        expected_entities=expected_entities,
        snapshot_rows=snapshot_rows,
        snapshot_validation=snapshot_validation,
        current_history_rows=current_history_rows,
        history_overlap_rows=history_overlap_rows,
        history_changes=history_changes,
        preview_limit=preview_limit,
    )
    error_message = _build_validation_error_message(summary)
    warning_messages = _build_validation_warning_messages(summary)
    summary["warning_messages"] = warning_messages
    summary["warning_count"] = len(warning_messages)
    summary["validation_status"] = "FAILED" if error_message else ("WARNING" if warning_messages else "PASSED")
    return summary, snapshot_rows, history_changes, error_message


def _compute_file_md5(file_path: Path) -> str:
    ensure_nas_path_access(file_path)
    import hashlib

    md5_hash = hashlib.md5()
    with file_path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            md5_hash.update(chunk)
    return md5_hash.hexdigest()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="从 NAS 导入门店经营负责人当前快照")
    parser.add_argument("--file-path", default=None, help="可选：显式指定负责人快照文件路径")
    parser.add_argument("--sheet-name", default=None, help="可选：显式指定工作表名称")
    parser.add_argument(
        "--snapshot-date",
        type=date.fromisoformat,
        default=date.today(),
        help="快照日期，格式 YYYY-MM-DD；默认今天",
    )
    parser.add_argument("--preview-limit", type=int, default=10, help="dry-run 输出预览条数")
    parser.add_argument("--output-json", help="可选：将摘要写入 JSON 文件")
    parser.add_argument("--created-by", default=DEFAULT_CREATED_BY, help="正式导入时写入 created_by / updated_by")
    parser.add_argument("--apply", action="store_true", help="显式启用写库模式；默认只做 dry-run")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    summary: dict[str, Any] | None = None
    try:
        file_path = _resolve_input_file(args.file_path)
        summary, snapshot_rows, history_changes, error_message = _run_import(
            file_path=file_path,
            requested_sheet_name=args.sheet_name,
            snapshot_date=args.snapshot_date,
            preview_limit=max(args.preview_limit, 0),
        )
        printable = _to_printable_summary(summary)

        if error_message:
            printable["error_message"] = error_message
            _write_output_json(args.output_json, printable)
            if args.apply:
                _write_import_log(summary, "FAILED", error_message)
            print(json.dumps(printable, ensure_ascii=False, indent=2))
            return 1

        if args.apply:
            summary = _apply_import(summary, snapshot_rows, history_changes, args.created_by)
            summary["mode"] = "apply"
            printable = _to_printable_summary(summary)
            _write_import_log(
                summary,
                "SUCCESS",
                (
                    f"matched={summary['matched_entity_count']}, snapshot_inserted={summary['snapshot_rows_inserted']}, "
                    f"history_opened={summary['history_rows_opened']}, history_closed={summary['history_rows_closed']}, "
                    f"warnings={summary.get('warning_count', 0)}"
                ),
            )

        _write_output_json(args.output_json, printable)
        print(json.dumps(printable, ensure_ascii=False, indent=2))
        return 0
    except Exception as exc:
        error_payload = {
            "mode": "apply" if args.apply else "dry-run",
            "validation_status": "FAILED",
            "error_message": str(exc),
        }
        if summary is not None:
            error_payload.update(_to_printable_summary(summary))
        _write_output_json(args.output_json, error_payload)
        print(json.dumps(error_payload, ensure_ascii=False, indent=2))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())